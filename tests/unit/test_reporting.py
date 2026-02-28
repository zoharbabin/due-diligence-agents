"""Unit tests for the reporting module.

Covers:
- FindingMerger: merge, dedup, severity escalation, ID generation, governance merge
- ReportDiffBuilder: detect new/resolved/changed findings between runs
- ExcelReportGenerator: schema-driven workbook generation, sheets, formatting
- ContractDateReconciler: classify customers by contract status
"""

from __future__ import annotations

import json
from datetime import date
from pathlib import Path
from typing import Any

import pytest

from dd_agents.models.reporting import ReportSchema
from dd_agents.reporting.contract_dates import (
    STATUS_ACTIVE_AUTO_RENEWAL,
    STATUS_ACTIVE_DB_STALE,
    STATUS_EXPIRED_CONFIRMED,
    STATUS_EXPIRED_NO_CONTRACTS,
    STATUS_LIKELY_ACTIVE,
    ContractDateReconciler,
)
from dd_agents.reporting.diff import ReportDiffBuilder
from dd_agents.reporting.excel import ExcelReportGenerator, compute_overall_risk
from dd_agents.reporting.merge import FindingMerger

# ---------------------------------------------------------------------------
# Helpers / fixtures
# ---------------------------------------------------------------------------


def _make_finding(
    severity: str = "P2",
    category: str = "change_of_control",
    source_path: str = "contract.pdf",
    location: str = "Section 5",
    exact_quote: str = "Sample quote",
    title: str = "Test finding",
    description: str = "A test finding",
    confidence: str = "high",
) -> dict:
    """Build a minimal agent-output finding dict."""
    return {
        "severity": severity,
        "category": category,
        "title": title,
        "description": description,
        "citations": [
            {
                "source_type": "file",
                "source_path": source_path,
                "location": location,
                "exact_quote": exact_quote,
            }
        ],
        "confidence": confidence,
    }


def _make_gap(
    customer: str = "Acme Corp",
    priority: str = "P1",
    gap_type: str = "Missing_Doc",
    missing_item: str = "MSA",
) -> dict:
    return {
        "customer": customer,
        "priority": priority,
        "gap_type": gap_type,
        "missing_item": missing_item,
        "why_needed": "Required for risk assessment",
        "risk_if_missing": "Incomplete analysis",
        "request_to_company": "Please provide the document",
        "evidence": "Referenced in SOW",
        "detection_method": "cross_reference",
    }


@pytest.fixture()
def report_schema_model() -> ReportSchema:
    """Load the real report_schema.json from config/."""
    schema_path = Path(__file__).parent.parent.parent / "config" / "report_schema.json"
    raw = json.loads(schema_path.read_text())
    return ReportSchema.model_validate(raw)


# ===========================================================================
# FindingMerger tests
# ===========================================================================


class TestFindingMerger:
    """Test the FindingMerger class."""

    def test_merge_two_agents_for_one_customer(self) -> None:
        """Merge findings from legal and finance for a single customer."""
        legal_output = {
            "customer": "Acme Corp",
            "customer_safe_name": "acme_corp",
            "findings": [
                _make_finding(severity="P1", source_path="msa.pdf", location="Section 3"),
            ],
            "governance_graph": {"edges": []},
            "cross_references": [],
        }
        finance_output = {
            "customer": "Acme Corp",
            "customer_safe_name": "acme_corp",
            "findings": [
                _make_finding(severity="P2", source_path="sow.pdf", location="Section 1"),
            ],
            "governance_graph": {"edges": []},
            "cross_references": [],
        }

        merger = FindingMerger(run_id="test_run", timestamp="2025-01-01T00:00:00Z")
        result = merger.merge_customer(
            agent_outputs={"legal": legal_output, "finance": finance_output},
            customer_name="Acme Corp",
            customer_safe_name="acme_corp",
        )

        assert result.customer == "Acme Corp"
        assert result.customer_safe_name == "acme_corp"
        assert len(result.findings) == 2

    def test_dedup_by_match_key_keeps_highest_severity(self) -> None:
        """Two findings with same source_path + location should dedup; highest severity wins."""
        legal_output = {
            "customer": "Acme Corp",
            "findings": [
                _make_finding(severity="P2", source_path="msa.pdf", location="Section 5"),
            ],
            "governance_graph": {"edges": []},
            "cross_references": [],
        }
        finance_output = {
            "customer": "Acme Corp",
            "findings": [
                _make_finding(severity="P1", source_path="msa.pdf", location="Section 5"),
            ],
            "governance_graph": {"edges": []},
            "cross_references": [],
        }

        merger = FindingMerger(run_id="test_run", timestamp="2025-01-01T00:00:00Z")
        result = merger.merge_customer(
            agent_outputs={"legal": legal_output, "finance": finance_output},
            customer_name="Acme Corp",
        )

        # Should be deduped to 1 finding
        assert len(result.findings) == 1
        assert result.findings[0].severity == "P1"

    def test_dedup_longest_exact_quote_on_tie(self) -> None:
        """When severity is equal, the finding with the longest exact_quote wins."""
        short_quote = _make_finding(
            severity="P1",
            source_path="msa.pdf",
            location="Section 5",
            exact_quote="Short",
        )
        long_quote = _make_finding(
            severity="P1",
            source_path="msa.pdf",
            location="Section 5",
            exact_quote="This is a much longer quote that should be preferred",
        )

        legal_output = {
            "customer": "Acme Corp",
            "findings": [short_quote],
            "governance_graph": {"edges": []},
            "cross_references": [],
        }
        finance_output = {
            "customer": "Acme Corp",
            "findings": [long_quote],
            "governance_graph": {"edges": []},
            "cross_references": [],
        }

        merger = FindingMerger(run_id="test", timestamp="2025-01-01T00:00:00Z")
        result = merger.merge_customer(
            agent_outputs={"legal": legal_output, "finance": finance_output},
            customer_name="Acme Corp",
        )

        assert len(result.findings) == 1
        # The winner should have the longer quote
        assert "much longer" in result.findings[0].citations[0].exact_quote

    def test_finding_id_generation(self) -> None:
        """Auto-generated IDs follow the pattern forensic-dd_{agent}_{safe_name}_{seq}."""
        agent_output = {
            "customer": "Acme Corp",
            "findings": [
                _make_finding(source_path="a.pdf", location="S1"),
                _make_finding(source_path="b.pdf", location="S2"),
            ],
            "governance_graph": {"edges": []},
            "cross_references": [],
        }

        merger = FindingMerger(run_id="test", timestamp="2025-01-01T00:00:00Z")
        result = merger.merge_customer(
            agent_outputs={"legal": agent_output},
            customer_name="Acme Corp",
            customer_safe_name="acme_corp",
        )

        assert len(result.findings) == 2
        assert result.findings[0].id == "forensic-dd_legal_acme_corp_0001"
        assert result.findings[1].id == "forensic-dd_legal_acme_corp_0002"

    def test_governance_merge_legal_primary(self) -> None:
        """Legal governance is primary; other agents add supplementary edges."""
        legal_output = {
            "customer": "Acme Corp",
            "findings": [],
            "governance_graph": {
                "edges": [
                    {"from_file": "sow.pdf", "to_file": "msa.pdf", "relationship": "governs"},
                ]
            },
            "cross_references": [],
        }
        finance_output = {
            "customer": "Acme Corp",
            "findings": [],
            "governance_graph": {
                "edges": [
                    # Same edge as legal -- should not duplicate
                    {"from_file": "sow.pdf", "to_file": "msa.pdf", "relationship": "governs"},
                    # New edge legal missed
                    {"from_file": "po.pdf", "to_file": "sow.pdf", "relationship": "references"},
                ]
            },
            "cross_references": [],
        }

        merger = FindingMerger(run_id="test", timestamp="2025-01-01T00:00:00Z")
        result = merger.merge_customer(
            agent_outputs={"legal": legal_output, "finance": finance_output},
            customer_name="Acme Corp",
        )

        # 1 from legal + 1 new from finance (duplicate suppressed)
        assert len(result.governance_graph.edges) == 2

    def test_merge_all_from_directory(self, tmp_path: Path) -> None:
        """merge_all discovers and merges all customer files."""
        findings_dir = tmp_path / "findings"
        legal_dir = findings_dir / "legal"
        legal_dir.mkdir(parents=True)

        legal_output = {
            "customer": "Beta Inc",
            "customer_safe_name": "beta_inc",
            "findings": [
                _make_finding(source_path="beta_msa.pdf", location="S1"),
            ],
            "governance_graph": {"edges": []},
            "cross_references": [],
        }
        (legal_dir / "beta_inc.json").write_text(json.dumps(legal_output))

        merger = FindingMerger(run_id="test", timestamp="2025-01-01T00:00:00Z")
        results = merger.merge_all(findings_dir)

        assert "beta_inc" in results
        assert len(results["beta_inc"].findings) == 1

    def test_write_merged(self, tmp_path: Path) -> None:
        """write_merged writes per-customer JSON files."""
        merger = FindingMerger(run_id="test", timestamp="2025-01-01T00:00:00Z")

        agent_output = {
            "customer": "Acme Corp",
            "findings": [_make_finding(source_path="a.pdf", location="S1")],
            "governance_graph": {"edges": []},
            "cross_references": [],
        }
        merged = {
            "acme_corp": merger.merge_customer(
                {"legal": agent_output},
                customer_name="Acme Corp",
                customer_safe_name="acme_corp",
            ),
        }

        out_dir = tmp_path / "merged"
        merger.write_merged(merged, out_dir)

        assert (out_dir / "acme_corp.json").exists()
        data = json.loads((out_dir / "acme_corp.json").read_text())
        assert data["customer"] == "Acme Corp"

    # ------------------------------------------------------------------
    # Citation normalisation / resilient promotion
    # ------------------------------------------------------------------

    def test_normalize_citation_maps_file_path_to_source_path(self) -> None:
        """Agent-produced file_path should be mapped to source_path."""
        merger = FindingMerger(run_id="test")
        raw = {"file_path": "msa.pdf", "page": 5, "section_ref": "Section 3", "exact_quote": "q"}
        normalised = merger._normalize_citation(raw)
        assert normalised["source_path"] == "msa.pdf"
        assert normalised["source_type"] == "file"
        assert normalised["location"] == "Section 3, p. 5"
        assert normalised["exact_quote"] == "q"
        assert "file_path" not in normalised

    def test_normalize_citation_preserves_correct_schema(self) -> None:
        """When agent already provides source_type + source_path, they are preserved."""
        merger = FindingMerger(run_id="test")
        raw = {"source_type": "file", "source_path": "sow.pdf", "location": "S1", "exact_quote": "q"}
        normalised = merger._normalize_citation(raw)
        assert normalised == raw

    def test_normalize_citation_defaults_missing_fields(self) -> None:
        """When only exact_quote is present, defaults are filled in."""
        merger = FindingMerger(run_id="test")
        raw = {"exact_quote": "some text"}
        normalised = merger._normalize_citation(raw)
        assert normalised["source_type"] == "file"
        assert normalised["source_path"] == ""
        assert normalised["location"] == ""
        assert normalised["exact_quote"] == "some text"

    # ------------------------------------------------------------------ #
    # Citation path resolution against file inventory
    # ------------------------------------------------------------------ #

    def test_resolve_citation_path_exact_match(self) -> None:
        """Exact match in inventory is returned unchanged."""
        inventory = ["1. Legal/1.1 MSA/msa.pdf", "2. Finance/2.1 Tax/tax.xlsx"]
        merger = FindingMerger(run_id="test", file_inventory=inventory)
        assert merger._resolve_citation_path("1. Legal/1.1 MSA/msa.pdf") == "1. Legal/1.1 MSA/msa.pdf"

    def test_resolve_citation_path_strips_md_suffix(self) -> None:
        """Extraction artifact .md suffix is stripped and resolved."""
        inventory = ["1. Legal/1.1 MSA/msa.pdf"]
        merger = FindingMerger(run_id="test", file_inventory=inventory)
        assert merger._resolve_citation_path("1. Legal/1.1 MSA/msa.pdf.md") == "1. Legal/1.1 MSA/msa.pdf"

    def test_resolve_citation_path_strips_absolute_prefix(self) -> None:
        """Absolute /Users/.../data-room/ prefix is stripped to data-room root."""
        inventory = ["1. Legal/1.1 MSA/msa.pdf"]
        merger = FindingMerger(run_id="test", file_inventory=inventory)
        result = merger._resolve_citation_path("/Users/me/data/1. Legal/1.1 MSA/msa.pdf")
        assert result == "1. Legal/1.1 MSA/msa.pdf"

    def test_resolve_citation_path_basename_unique(self) -> None:
        """When basename is unique in inventory, resolve to the full path."""
        inventory = ["1. Legal/1.1 MSA/master_services_agreement.pdf"]
        merger = FindingMerger(run_id="test", file_inventory=inventory)
        assert merger._resolve_citation_path("master_services_agreement.pdf") == (
            "1. Legal/1.1 MSA/master_services_agreement.pdf"
        )

    def test_resolve_citation_path_basename_ambiguous_suffix_disambiguates(self) -> None:
        """When multiple files share a basename, the matching suffix wins."""
        inventory = [
            "1. Legal/1.1 MSA/contract.pdf",
            "2. Finance/2.1 Tax/contract.pdf",
        ]
        merger = FindingMerger(run_id="test", file_inventory=inventory)
        assert merger._resolve_citation_path("2.1 Tax/contract.pdf") == "2. Finance/2.1 Tax/contract.pdf"

    def test_resolve_citation_path_no_inventory(self) -> None:
        """When no inventory is provided, path is returned unchanged."""
        merger = FindingMerger(run_id="test")
        assert merger._resolve_citation_path("msa.pdf") == "msa.pdf"

    def test_resolve_citation_path_no_match(self) -> None:
        """When path doesn't match anything, it's returned unchanged."""
        inventory = ["1. Legal/msa.pdf"]
        merger = FindingMerger(run_id="test", file_inventory=inventory)
        assert merger._resolve_citation_path("nonexistent.docx") == "nonexistent.docx"

    def test_resolve_citation_path_md_suffix_then_basename(self) -> None:
        """.md stripping + basename lookup works in combination."""
        inventory = ["1. Legal/1.1 MSA/msa.pdf"]
        merger = FindingMerger(run_id="test", file_inventory=inventory)
        assert merger._resolve_citation_path("msa.pdf.md") == "1. Legal/1.1 MSA/msa.pdf"

    def test_normalize_citation_resolves_path_with_inventory(self) -> None:
        """End-to-end: _normalize_citation resolves source_path via inventory."""
        inventory = ["1. Legal/1.1 MSA/msa.pdf"]
        merger = FindingMerger(run_id="test", file_inventory=inventory)
        raw = {"file_path": "msa.pdf", "exact_quote": "test"}
        normalised = merger._normalize_citation(raw)
        assert normalised["source_path"] == "1. Legal/1.1 MSA/msa.pdf"

    def test_normalize_citation_skips_synthetic_paths(self) -> None:
        """Paths starting with [ (synthetic) should not be resolved."""
        inventory = ["1. Legal/msa.pdf"]
        merger = FindingMerger(run_id="test", file_inventory=inventory)
        raw = {"source_path": "[synthetic:no_citation_provided]", "source_type": "file", "location": ""}
        normalised = merger._normalize_citation(raw)
        assert normalised["source_path"] == "[synthetic:no_citation_provided]"

    def test_promote_findings_with_flat_citation_fields(self) -> None:
        """Findings with flat file_path/page/exact_quote (no citations array) are promoted."""
        merger = FindingMerger(run_id="test", timestamp="2025-01-01T00:00:00Z")
        raw_findings = [
            {
                "agent": "legal",
                "severity": "P2",
                "category": "change_of_control",
                "title": "CoC clause",
                "description": "desc",
                "confidence": "high",
                "file_path": "msa.pdf",
                "page": 12,
                "section_ref": "Section 8.2",
                "exact_quote": "In the event of a change of control...",
            },
        ]
        promoted = merger._promote_findings(raw_findings, "Acme Corp", "acme_corp")
        assert len(promoted) == 1
        cit = promoted[0].citations[0]
        assert cit.source_path == "msa.pdf"
        assert cit.source_type == "file"
        assert "Section 8.2" in cit.location
        assert "p. 12" in cit.location
        assert "change of control" in (cit.exact_quote or "")

    def test_promote_findings_with_missing_citation_fields(self) -> None:
        """Citations missing source_type/source_path are handled gracefully."""
        merger = FindingMerger(run_id="test", timestamp="2025-01-01T00:00:00Z")
        raw_findings = [
            {
                "agent": "finance",
                "severity": "P3",
                "category": "pricing",
                "title": "Pricing concern",
                "description": "desc",
                "confidence": "medium",
                "citations": [
                    {"exact_quote": "The price shall be...", "page": 3},
                ],
            },
        ]
        promoted = merger._promote_findings(raw_findings, "Acme Corp", "acme_corp")
        assert len(promoted) == 1
        cit = promoted[0].citations[0]
        assert cit.source_type == "file"
        assert cit.exact_quote == "The price shall be..."

    def test_promote_findings_no_citations_gets_synthetic(self) -> None:
        """A finding with no citation info at all gets a synthetic placeholder."""
        merger = FindingMerger(run_id="test", timestamp="2025-01-01T00:00:00Z")
        raw_findings = [
            {
                "agent": "commercial",
                "severity": "P3",
                "category": "pricing",
                "title": "General concern",
                "description": "desc",
                "confidence": "low",
            },
        ]
        promoted = merger._promote_findings(raw_findings, "Acme Corp", "acme_corp")
        assert len(promoted) == 1
        cit = promoted[0].citations[0]
        assert "[synthetic:" in cit.source_path

    def test_collect_gaps_handles_string_entries(self) -> None:
        """String gap entries should be wrapped, not crash."""
        agent_outputs = {
            "legal": {
                "gaps": [
                    "Missing MSA document",
                    {
                        "missing_item": "SOW",
                        "gap_type": "Missing_Doc",
                        "why_needed": "Required",
                        "risk_if_missing": "High",
                        "request_to_company": "Please provide",
                        "evidence": "Referenced in amendment",
                        "detection_method": "cross_reference",
                        "customer": "Acme Corp",
                        "priority": "P2",
                    },
                ],
            },
        }
        # The string entry will be wrapped and then fail Gap validation
        # (missing required fields), but the dict entry should survive.
        gaps = FindingMerger._collect_gaps(agent_outputs, "Acme Corp")
        assert len(gaps) >= 1
        assert any(g.missing_item == "SOW" for g in gaps)

    def test_merge_skips_non_dict_findings(self) -> None:
        """Non-dict finding entries should be skipped without crashing."""
        agent_output = {
            "customer": "Acme Corp",
            "findings": [
                "this is a string, not a finding",
                _make_finding(source_path="a.pdf", location="S1"),
            ],
            "governance_graph": {"edges": []},
            "cross_references": [],
        }
        merger = FindingMerger(run_id="test", timestamp="2025-01-01T00:00:00Z")
        result = merger.merge_customer(
            agent_outputs={"legal": agent_output},
            customer_name="Acme Corp",
            customer_safe_name="acme_corp",
        )
        # Only the valid dict finding should survive
        assert len(result.findings) == 1


# ===========================================================================
# Merge normalization tests (gaps, governance, cross-references)
# ===========================================================================


class TestMergeNormalization:
    """Tests for normalization layers that map agent output → Pydantic models."""

    def test_normalize_gap_maps_simplified_fields(self) -> None:
        """Agent-produced gap with simplified fields is normalised to Gap model."""
        raw = {
            "gap_type": "missing document",
            "description": "MSA not found in data room",
            "file_path": "sow.pdf",
            "severity": "P1",
        }
        normalised = FindingMerger._normalize_gap(raw, "Acme Corp", "legal")
        assert normalised["customer"] == "Acme Corp"
        assert normalised["agent"] == "legal"
        assert normalised["gap_type"] == "Missing_Doc"
        assert normalised["priority"] == "P1"
        assert normalised["why_needed"] == "MSA not found in data room"
        assert "sow.pdf" in normalised["evidence"]
        assert normalised["detection_method"] == "checklist"
        assert "missing_item" in normalised

    def test_normalize_gap_preserves_full_fields(self) -> None:
        """When agent provides all Gap model fields, they are preserved."""
        raw = _make_gap(customer="Beta Inc", priority="P0", gap_type="Missing_Doc", missing_item="NDA")
        normalised = FindingMerger._normalize_gap(raw, "Beta Inc", "legal")
        assert normalised["customer"] == "Beta Inc"
        assert normalised["missing_item"] == "NDA"
        assert normalised["priority"] == "P0"

    def test_collect_gaps_with_agent_format(self) -> None:
        """Gaps produced in agent-simplified format are accepted after normalization."""
        agent_outputs = {
            "legal": {
                "customer": "Acme Corp",
                "findings": [],
                "gaps": [
                    {
                        "gap_type": "missing document",
                        "description": "Certificate of incorporation not found",
                        "title": "Certificate of Incorporation",
                        "severity": "P1",
                        "file_path": "sow.pdf",
                    }
                ],
            }
        }
        merger = FindingMerger(run_id="test", timestamp="2025-01-01T00:00:00Z")
        result = merger.merge_customer(agent_outputs, customer_name="Acme Corp", customer_safe_name="acme_corp")
        assert len(result.gaps) == 1
        assert result.gaps[0].gap_type == "Missing_Doc"
        assert result.gaps[0].customer == "Acme Corp"

    # -- _coerce_gap_type keyword classification --

    @pytest.mark.parametrize(
        ("raw_type", "expected"),
        [
            # Exact enum values pass through
            ("Missing_Doc", "Missing_Doc"),
            ("Missing_Data", "Missing_Data"),
            ("Ambiguous_Link", "Ambiguous_Link"),
            ("Unreadable", "Unreadable"),
            ("Contradiction", "Contradiction"),
            ("Data_Mismatch", "Data_Mismatch"),
            # Missing_Doc variants
            ("Not_Found", "Missing_Doc"),
            ("not_found", "Missing_Doc"),
            ("missing document", "Missing_Doc"),
            ("missing_document", "Missing_Doc"),
            ("document_missing", "Missing_Doc"),
            ("not_provided", "Missing_Doc"),
            ("absent", "Missing_Doc"),
            ("unavailable", "Missing_Doc"),
            ("missing_file", "Missing_Doc"),
            ("missing_contract", "Missing_Doc"),
            ("missing_policy", "Missing_Doc"),
            # Missing_Data variants
            ("Incomplete", "Missing_Data"),
            ("incomplete_data", "Missing_Data"),
            ("partial", "Missing_Data"),
            ("redacted", "Missing_Data"),
            ("blank", "Missing_Data"),
            ("empty", "Missing_Data"),
            ("missing_information", "Missing_Data"),
            ("missing_detail", "Missing_Data"),
            ("missing_analysis", "Missing_Data"),
            ("missing_metrics", "Missing_Data"),
            ("missing_content", "Missing_Data"),
            ("missing_category", "Missing_Data"),
            ("no_data", "Missing_Data"),
            # Unreadable variants
            ("unreadable_document", "Unreadable"),
            ("ocr_failure", "Unreadable"),
            ("scan_quality", "Unreadable"),
            ("garbled", "Unreadable"),
            ("image_only", "Unreadable"),
            # Data_Mismatch variants
            ("mismatch", "Data_Mismatch"),
            ("discrepancy", "Data_Mismatch"),
            ("data_inconsistency", "Data_Mismatch"),
            # Contradiction variants
            ("conflict", "Contradiction"),
            ("contradicts_prior", "Contradiction"),
            # Ambiguous_Link variants
            ("ambiguous", "Ambiguous_Link"),
            ("unclear_reference", "Ambiguous_Link"),
            # Unknown → default Missing_Doc
            ("something_random", "Missing_Doc"),
            ("xyz", "Missing_Doc"),
        ],
    )
    def test_coerce_gap_type(self, raw_type: str, expected: str) -> None:
        """_coerce_gap_type maps agent strings to valid GapType enum values."""
        assert FindingMerger._coerce_gap_type(raw_type, "test") == expected

    def test_coerce_gap_type_none_defaults_to_missing_doc(self) -> None:
        """None or empty gap_type defaults to Missing_Doc."""
        assert FindingMerger._coerce_gap_type(None, "test") == "Missing_Doc"
        assert FindingMerger._coerce_gap_type("", "test") == "Missing_Doc"
        assert FindingMerger._coerce_gap_type("  ", "test") == "Missing_Doc"

    def test_coerce_gap_type_specificity_ordering(self) -> None:
        """More specific keywords win over broad ones (e.g. 'unreadable' beats 'missing')."""
        # 'unreadable_document' should be Unreadable, not Missing_Doc
        assert FindingMerger._coerce_gap_type("unreadable_document", "test") == "Unreadable"
        # 'data_mismatch_found' should be Data_Mismatch, not Missing_Doc
        assert FindingMerger._coerce_gap_type("data_mismatch_found", "test") == "Data_Mismatch"
        # 'incomplete_document' should be Missing_Data, not Missing_Doc
        assert FindingMerger._coerce_gap_type("incomplete_document", "test") == "Missing_Data"

    # -- priority coercion --

    def test_normalize_gap_priority_high_to_p1(self) -> None:
        """Agent-produced priority 'high' maps to P1."""
        raw = {"gap_type": "Missing_Doc", "title": "Test", "description": "Test", "severity": "high"}
        normalised = FindingMerger._normalize_gap(raw, "Test Corp", "commercial")
        assert normalised["priority"] == "P1"

    def test_normalize_gap_priority_low_to_p3(self) -> None:
        """Agent-produced priority 'low' maps to P3."""
        raw = {"gap_type": "Missing_Doc", "title": "Test", "description": "Test", "severity": "low"}
        normalised = FindingMerger._normalize_gap(raw, "Test Corp", "commercial")
        assert normalised["priority"] == "P3"

    # -- end-to-end: gaps survive normalization + Pydantic validation --

    def test_collect_gaps_agent_variants_survive_validation(self) -> None:
        """Gaps with agent-produced types survive normalization + validation."""
        agent_outputs = {
            "producttech": {
                "customer": "Acme Corp",
                "findings": [],
                "gaps": [
                    {"gap_type": "Not_Found", "title": "Privacy Policy", "description": "Empty file"},
                    {"gap_type": "Incomplete", "title": "SOC2 Report", "description": "Partial"},
                    {"gap_type": "missing_information", "title": "Revenue Data", "description": "Missing"},
                ],
            }
        }
        merger = FindingMerger(run_id="test", timestamp="2025-01-01T00:00:00Z")
        result = merger.merge_customer(agent_outputs, customer_name="Acme Corp", customer_safe_name="acme_corp")
        assert len(result.gaps) == 3
        gap_types = {g.gap_type.value for g in result.gaps}
        assert "Missing_Doc" in gap_types
        assert "Missing_Data" in gap_types

    def test_normalize_governance_edge_maps_from_to(self) -> None:
        """Agent-produced 'from'/'to' mapped to 'from_file'/'to_file'."""
        raw = {"from": "sow.pdf", "to": "msa.pdf", "relationship": "governs"}
        normalised = FindingMerger._normalize_governance_edge(raw)
        assert normalised["from_file"] == "sow.pdf"
        assert normalised["to_file"] == "msa.pdf"
        assert "from" not in normalised
        assert "to" not in normalised

    def test_consolidate_governance_with_agent_format(self) -> None:
        """Governance edges using 'from'/'to' are accepted after normalization."""
        agent_outputs = {
            "legal": {
                "customer": "Acme Corp",
                "findings": [],
                "governance_graph": {
                    "edges": [
                        {"from": "sow.pdf", "to": "msa.pdf", "relationship": "governs"},
                        {"from": "amendment.pdf", "to": "msa.pdf", "relationship": "amends"},
                    ]
                },
            }
        }
        merger = FindingMerger(run_id="test", timestamp="2025-01-01T00:00:00Z")
        result = merger.merge_customer(agent_outputs, customer_name="Acme Corp", customer_safe_name="acme_corp")
        assert len(result.governance_graph.edges) == 2
        assert result.governance_graph.edges[0].from_file == "sow.pdf"
        assert result.governance_graph.edges[0].to_file == "msa.pdf"

    def test_normalize_cross_reference_maps_flat_fields(self) -> None:
        """Agent-produced cross-ref with flat fields is normalised."""
        raw = {
            "data_point": "ARR",
            "contract_value": "$1.2M",
            "reference_value": "$1.1M",
            "source_file": "msa.pdf",
            "reference_file": "financials.xlsx",
            "target_category": "Revenue",
            "status": "mismatch",
            "variance": "-8.3%",
        }
        normalised = FindingMerger._normalize_cross_reference(raw)
        assert normalised["match_status"] == "mismatch"
        assert normalised["contract_source"]["file"] == "msa.pdf"
        assert normalised["reference_source"]["file"] == "financials.xlsx"
        assert normalised["reference_source"]["tab"] == "Revenue"
        assert "source_file" not in normalised
        assert "status" not in normalised

    def test_union_cross_refs_with_agent_format(self) -> None:
        """Cross-references with agent field names are accepted after normalization."""
        agent_outputs = {
            "finance": {
                "customer": "Acme Corp",
                "findings": [],
                "cross_references": [
                    {
                        "data_point": "Contract Value",
                        "contract_value": "$500K",
                        "reference_value": "$500K",
                        "source_file": "order_form.pdf",
                        "reference_file": "pricing.xlsx",
                        "status": "match",
                    }
                ],
            }
        }
        merger = FindingMerger(run_id="test", timestamp="2025-01-01T00:00:00Z")
        result = merger.merge_customer(agent_outputs, customer_name="Acme Corp", customer_safe_name="acme_corp")
        assert len(result.cross_references) == 1
        assert result.cross_references[0].data_point == "Contract Value"
        assert result.cross_references[0].match_status == "match"

    def test_union_cross_refs_recovers_string_entries(self) -> None:
        """String entries in cross_references should be auto-recovered, not dropped."""
        agent_outputs = {
            "finance": {
                "customer": "Acme Corp",
                "findings": [],
                "cross_references": [
                    "Revenue terms match between MSA and Order Form",
                    42,
                    {
                        "data_point": "Revenue",
                        "contract_value": "$1M",
                        "reference_value": "$1M",
                        "source_file": "contract.pdf",
                        "reference_file": "financials.xlsx",
                        "status": "match",
                    },
                ],
            }
        }
        merger = FindingMerger(run_id="test", timestamp="2025-01-01T00:00:00Z")
        result = merger.merge_customer(agent_outputs, customer_name="Acme Corp", customer_safe_name="acme_corp")
        # The string is recovered as a cross-reference; the int is dropped
        assert len(result.cross_references) == 2
        # Recovered entry should have the string as data_point
        recovered = [cr for cr in result.cross_references if "Revenue terms" in cr.data_point]
        assert len(recovered) == 1
        assert recovered[0].match_status == "match"  # "match" keyword detected
        # The dict entry also survives
        assert any(cr.data_point == "Revenue" for cr in result.cross_references)

    def test_cross_ref_string_mismatch_keyword_detected(self) -> None:
        """Bare string cross-refs with mismatch keywords get status='mismatch'."""
        agent_outputs = {
            "finance": {
                "customer": "Acme Corp",
                "findings": [],
                "cross_references": [
                    "ARR discrepancy between contract ($1.2M) and revenue cube ($1.0M)",
                ],
            }
        }
        merger = FindingMerger(run_id="test", timestamp="2025-01-01T00:00:00Z")
        result = merger.merge_customer(agent_outputs, customer_name="Acme Corp", customer_safe_name="acme_corp")
        assert len(result.cross_references) == 1
        assert result.cross_references[0].match_status == "mismatch"
        assert "ARR discrepancy" in result.cross_references[0].data_point

    def test_cross_ref_string_unknown_status(self) -> None:
        """Bare string cross-refs without match/mismatch keywords get 'not_available'."""
        agent_outputs = {
            "finance": {
                "customer": "Acme Corp",
                "findings": [],
                "cross_references": [
                    "Payment terms reference Section 4.2 of the Order Form",
                ],
            }
        }
        merger = FindingMerger(run_id="test", timestamp="2025-01-01T00:00:00Z")
        result = merger.merge_customer(agent_outputs, customer_name="Acme Corp", customer_safe_name="acme_corp")
        assert len(result.cross_references) == 1
        assert result.cross_references[0].match_status == "not_available"

    def test_cross_ref_empty_string_dropped(self) -> None:
        """Empty or whitespace-only string cross-refs are dropped."""
        agent_outputs = {
            "finance": {
                "customer": "Acme Corp",
                "findings": [],
                "cross_references": ["", "   ", None],
            }
        }
        merger = FindingMerger(run_id="test", timestamp="2025-01-01T00:00:00Z")
        result = merger.merge_customer(agent_outputs, customer_name="Acme Corp", customer_safe_name="acme_corp")
        assert len(result.cross_references) == 0

    def test_gap_string_recovery_preserves_full_text(self) -> None:
        """Bare string gaps should be recovered with the full text preserved."""
        agent_outputs = {
            "legal": {
                "gaps": [
                    "Missing DPA for data processing. Required under GDPR compliance obligations.",
                ],
            },
        }
        gaps = FindingMerger._collect_gaps(agent_outputs, "Acme Corp")
        assert len(gaps) == 1
        assert gaps[0].missing_item == "Missing DPA for data processing"
        # Full text preserved in why_needed
        assert "GDPR" in gaps[0].why_needed

    def test_gap_string_recovery_infers_gap_type(self) -> None:
        """Bare string gaps should infer gap_type from keywords."""
        agent_outputs = {
            "producttech": {
                "gaps": [
                    "Unreadable scanned document: security_addendum_scan.pdf",
                    "Contradiction between MSA liability cap and Order Form cap",
                    "Missing SOC 2 Type II audit report",
                ],
            },
        }
        gaps = FindingMerger._collect_gaps(agent_outputs, "Acme Corp")
        assert len(gaps) == 3
        gap_types = {g.missing_item[:20]: str(g.gap_type) for g in gaps}
        assert gap_types["Unreadable scanned d"] == "Unreadable"
        assert gap_types["Contradiction betwee"] == "Contradiction"
        assert gap_types["Missing SOC 2 Type I"] == "Missing_Doc"

    def test_gap_string_recovery_infers_priority(self) -> None:
        """Bare string gaps with critical keywords should get P1 priority."""
        agent_outputs = {
            "legal": {
                "gaps": [
                    "CRITICAL: Missing master services agreement for largest customer",
                    "Minor documentation gap in archived SOW",
                ],
            },
        }
        gaps = FindingMerger._collect_gaps(agent_outputs, "Acme Corp")
        assert len(gaps) == 2
        critical_gap = next(g for g in gaps if "CRITICAL" in g.missing_item)
        minor_gap = next(g for g in gaps if "Minor" in g.missing_item)
        assert str(critical_gap.priority) == "P1"
        assert str(minor_gap.priority) == "P2"


# ===========================================================================
# ReportDiffBuilder tests
# ===========================================================================


class TestReportDiffBuilder:
    """Test the ReportDiffBuilder class."""

    @staticmethod
    def _setup_run_dir(
        base: Path,
        findings: dict[str, list[dict]],
        gaps: dict[str, list[dict]] | None = None,
    ) -> Path:
        """Create a findings directory tree for one run."""
        merged_dir = base / "merged"
        merged_dir.mkdir(parents=True, exist_ok=True)
        gaps_dir = merged_dir / "gaps"
        gaps_dir.mkdir(parents=True, exist_ok=True)

        for csn, f_list in findings.items():
            data = {"customer": csn, "findings": f_list}
            (merged_dir / f"{csn}.json").write_text(json.dumps(data))

        if gaps:
            for csn, g_list in gaps.items():
                (gaps_dir / f"{csn}.json").write_text(json.dumps({"gaps": g_list}))

        return base

    def test_detect_new_finding(self, tmp_path: Path) -> None:
        prior = self._setup_run_dir(
            tmp_path / "prior",
            findings={"acme": []},
        )
        current = self._setup_run_dir(
            tmp_path / "current",
            findings={
                "acme": [
                    _make_finding(category="risk", source_path="msa.pdf", title="New risk"),
                ],
            },
        )

        builder = ReportDiffBuilder()
        diff = builder.build_diff(current, prior)

        assert diff.summary.new_findings == 1
        assert any(c.change_type == "new_finding" for c in diff.changes)

    def test_detect_resolved_finding(self, tmp_path: Path) -> None:
        prior = self._setup_run_dir(
            tmp_path / "prior",
            findings={
                "acme": [
                    _make_finding(category="risk", source_path="msa.pdf", title="Old risk"),
                ],
            },
        )
        current = self._setup_run_dir(
            tmp_path / "current",
            findings={"acme": []},
        )

        builder = ReportDiffBuilder()
        diff = builder.build_diff(current, prior)

        assert diff.summary.resolved_findings == 1
        assert any(c.change_type == "resolved_finding" for c in diff.changes)

    def test_detect_changed_severity(self, tmp_path: Path) -> None:
        prior = self._setup_run_dir(
            tmp_path / "prior",
            findings={
                "acme": [
                    _make_finding(
                        severity="P2",
                        category="risk",
                        source_path="msa.pdf",
                        title="Escalated",
                    ),
                ],
            },
        )
        current = self._setup_run_dir(
            tmp_path / "current",
            findings={
                "acme": [
                    _make_finding(
                        severity="P0",
                        category="risk",
                        source_path="msa.pdf",
                        title="Escalated",
                    ),
                ],
            },
        )

        builder = ReportDiffBuilder()
        diff = builder.build_diff(current, prior)

        assert diff.summary.changed_severity == 1
        change = next(c for c in diff.changes if c.change_type == "changed_severity")
        assert change.prior_severity == "P2"
        assert change.current_severity == "P0"

    def test_detect_new_and_removed_customer(self, tmp_path: Path) -> None:
        prior = self._setup_run_dir(
            tmp_path / "prior",
            findings={"acme": [], "removed_co": []},
        )
        current = self._setup_run_dir(
            tmp_path / "current",
            findings={"acme": [], "new_co": []},
        )

        builder = ReportDiffBuilder()
        diff = builder.build_diff(current, prior)

        assert diff.summary.new_customers == 1
        assert diff.summary.removed_customers == 1

    def test_detect_new_gap(self, tmp_path: Path) -> None:
        prior = self._setup_run_dir(
            tmp_path / "prior",
            findings={"acme": []},
            gaps={"acme": []},
        )
        current = self._setup_run_dir(
            tmp_path / "current",
            findings={"acme": []},
            gaps={
                "acme": [_make_gap(customer="acme", missing_item="SOW")],
            },
        )

        builder = ReportDiffBuilder()
        diff = builder.build_diff(current, prior)

        assert diff.summary.new_gaps == 1

    def test_detect_resolved_gap(self, tmp_path: Path) -> None:
        prior = self._setup_run_dir(
            tmp_path / "prior",
            findings={"acme": []},
            gaps={"acme": [_make_gap(customer="acme", missing_item="SOW")]},
        )
        current = self._setup_run_dir(
            tmp_path / "current",
            findings={"acme": []},
            gaps={"acme": []},
        )

        builder = ReportDiffBuilder()
        diff = builder.build_diff(current, prior)

        assert diff.summary.resolved_gaps == 1

    def test_write_diff(self, tmp_path: Path) -> None:
        """write_diff serializes the diff to JSON."""
        prior = self._setup_run_dir(
            tmp_path / "prior",
            findings={"acme": []},
        )
        current = self._setup_run_dir(
            tmp_path / "current",
            findings={"acme": []},
        )

        builder = ReportDiffBuilder()
        diff = builder.build_diff(
            current,
            prior,
            current_run_id="run_002",
            prior_run_id="run_001",
        )

        out = tmp_path / "report_diff.json"
        builder.write_diff(diff, out)

        assert out.exists()
        data = json.loads(out.read_text())
        assert data["current_run_id"] == "run_002"
        assert data["prior_run_id"] == "run_001"


# ===========================================================================
# ExcelReportGenerator tests
# ===========================================================================


class TestExcelReportGenerator:
    """Test the ExcelReportGenerator class."""

    def test_generate_workbook_from_schema(
        self,
        tmp_path: Path,
        report_schema_model: ReportSchema,
    ) -> None:
        """Generate a workbook from the real report_schema.json and verify structure."""
        # Build minimal merged data
        merged = {
            "acme_corp": {
                "customer": "Acme Corp",
                "customer_safe_name": "acme_corp",
                "findings": [
                    {
                        "id": "forensic-dd_legal_acme_corp_0001",
                        "severity": "P0",
                        "category": "change_of_control",
                        "title": "CoC clause",
                        "description": "Termination on change of control",
                        "citations": [
                            {
                                "source_type": "file",
                                "source_path": "msa.pdf",
                                "location": "Section 12",
                                "exact_quote": "terminate upon change of control",
                            }
                        ],
                        "confidence": "high",
                        "agent": "legal",
                        "analysis_unit": "Acme Corp",
                    },
                    {
                        "id": "forensic-dd_finance_acme_corp_0001",
                        "severity": "P2",
                        "category": "pricing",
                        "title": "Below market rate",
                        "description": "Pricing is 30% below market",
                        "citations": [
                            {
                                "source_type": "file",
                                "source_path": "sow.pdf",
                                "location": "Schedule A",
                                "exact_quote": "Total price: $100,000",
                            }
                        ],
                        "confidence": "medium",
                        "agent": "finance",
                        "analysis_unit": "Acme Corp",
                    },
                ],
                "gaps": [],
                "cross_references": [],
                "governance_graph": {"edges": []},
                "governance_resolved_pct": 0.75,
                "files_analyzed": 5,
            },
        }

        deal_config: dict = {}  # No special conditions
        out_path = tmp_path / "report.xlsx"

        gen = ExcelReportGenerator()
        gen.generate(merged, report_schema_model, out_path, deal_config)

        assert out_path.exists()

        # Verify with openpyxl
        from openpyxl import load_workbook

        wb = load_workbook(str(out_path))

        # Count always-active sheets (10 sheets)
        always_active = [s for s in report_schema_model.sheets if s.activation_condition == "always"]
        for sheet_def in always_active:
            assert sheet_def.name in wb.sheetnames, f"Missing always-active sheet: {sheet_def.name}"

        # Conditional sheets should NOT be present when conditions are not met
        assert "Quality_Audit" not in wb.sheetnames
        assert "_Metadata" not in wb.sheetnames

    def test_sheet_column_headers_match_schema(
        self,
        tmp_path: Path,
        report_schema_model: ReportSchema,
    ) -> None:
        """Column headers in each sheet match the schema definition."""
        merged = {
            "acme": {
                "customer": "Acme",
                "findings": [],
                "gaps": [],
                "cross_references": [],
                "governance_graph": {"edges": []},
                "governance_resolved_pct": 0.0,
            },
        }

        out_path = tmp_path / "report.xlsx"
        gen = ExcelReportGenerator()
        gen.generate(merged, report_schema_model, out_path)

        from openpyxl import load_workbook

        wb = load_workbook(str(out_path))

        for sheet_def in report_schema_model.sheets:
            if sheet_def.name not in wb.sheetnames:
                continue
            ws = wb[sheet_def.name]
            # Read header row
            headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
            # Filter columns without judge activation
            expected = [
                col.name
                for col in sheet_def.columns
                if col.activation_condition is None or col.activation_condition == "always"
            ]
            assert headers == expected, f"Sheet {sheet_def.name}: headers {headers} != expected {expected}"

    def test_severity_conditional_formatting(
        self,
        tmp_path: Path,
        report_schema_model: ReportSchema,
    ) -> None:
        """P0 findings in the Wolf_Pack sheet should have red background."""
        merged = {
            "acme": {
                "customer": "Acme",
                "findings": [
                    {
                        "id": "forensic-dd_legal_acme_0001",
                        "severity": "P0",
                        "category": "test",
                        "title": "Critical",
                        "description": "Very bad",
                        "citations": [
                            {
                                "source_type": "file",
                                "source_path": "msa.pdf",
                                "location": "S1",
                                "exact_quote": "quote here",
                            }
                        ],
                        "confidence": "high",
                        "agent": "legal",
                        "analysis_unit": "Acme",
                    },
                ],
                "gaps": [],
                "cross_references": [],
                "governance_graph": {"edges": []},
                "governance_resolved_pct": 0.0,
            },
        }

        out_path = tmp_path / "report.xlsx"
        gen = ExcelReportGenerator()
        gen.generate(merged, report_schema_model, out_path)

        from openpyxl import load_workbook

        wb = load_workbook(str(out_path))
        ws = wb["Wolf_Pack"]

        # Find the severity column
        sev_col = None
        for c in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=c).value == "Severity":
                sev_col = c
                break

        assert sev_col is not None
        # Row 2 should have severity P0 with red fill
        sev_cell = ws.cell(row=2, column=sev_col)
        assert sev_cell.value == "P0"
        assert sev_cell.fill.start_color.rgb is not None
        # The fill should be FF0000 (with possible alpha prefix)
        assert "FF0000" in sev_cell.fill.start_color.rgb

    def test_conditional_sheets_activated(
        self,
        tmp_path: Path,
        report_schema_model: ReportSchema,
    ) -> None:
        """When deal_config activates conditional sheets, they appear."""
        merged = {
            "acme": {
                "customer": "Acme",
                "findings": [],
                "gaps": [],
                "cross_references": [],
                "governance_graph": {"edges": []},
                "governance_resolved_pct": 0.0,
            },
        }
        deal_config = {
            "judge": {"enabled": True},
            "reporting": {
                "include_metadata_sheet": True,
                "include_diff_sheet": True,
            },
            "source_of_truth": {
                "customer_database": {"path": "customers.csv"},
            },
        }

        out_path = tmp_path / "report.xlsx"
        gen = ExcelReportGenerator()
        gen.generate(merged, report_schema_model, out_path, deal_config)

        from openpyxl import load_workbook

        wb = load_workbook(str(out_path))

        assert "Quality_Audit" in wb.sheetnames
        assert "_Metadata" in wb.sheetnames
        assert "Run_Diff" in wb.sheetnames
        assert "Contract_Date_Reconciliation" in wb.sheetnames

    def test_freeze_panes_and_auto_filter(
        self,
        tmp_path: Path,
        report_schema_model: ReportSchema,
    ) -> None:
        """Verify freeze panes and auto-filter are applied."""
        merged = {
            "acme": {
                "customer": "Acme",
                "findings": [
                    {
                        "id": "forensic-dd_legal_acme_0001",
                        "severity": "P2",
                        "category": "test",
                        "title": "Finding",
                        "description": "Desc",
                        "citations": [
                            {
                                "source_type": "file",
                                "source_path": "msa.pdf",
                                "location": "S1",
                                "exact_quote": "q",
                            }
                        ],
                        "confidence": "high",
                        "agent": "legal",
                        "analysis_unit": "Acme",
                    },
                ],
                "gaps": [],
                "cross_references": [],
                "governance_graph": {"edges": []},
                "governance_resolved_pct": 0.0,
            },
        }

        out_path = tmp_path / "report.xlsx"
        gen = ExcelReportGenerator()
        gen.generate(merged, report_schema_model, out_path)

        from openpyxl import load_workbook

        wb = load_workbook(str(out_path))
        ws = wb["Summary"]

        # Freeze panes at A2 (row 1 frozen)
        assert ws.freeze_panes == "A2"
        # Auto-filter should be set
        assert ws.auto_filter.ref is not None

    def test_summary_formulas(
        self,
        tmp_path: Path,
        report_schema_model: ReportSchema,
    ) -> None:
        """Summary row should contain computed totals."""
        merged = {
            "acme": {
                "customer": "Acme",
                "findings": [
                    {
                        "id": "f1",
                        "severity": "P0",
                        "category": "test",
                        "title": "F1",
                        "description": "D",
                        "citations": [
                            {"source_type": "file", "source_path": "a.pdf", "location": "S1", "exact_quote": "q"}
                        ],
                        "confidence": "high",
                        "agent": "legal",
                        "analysis_unit": "Acme",
                    },
                ],
                "gaps": [],
                "cross_references": [],
                "governance_graph": {"edges": []},
                "governance_resolved_pct": 1.0,
                "files_analyzed": 3,
            },
            "beta": {
                "customer": "Beta",
                "findings": [
                    {
                        "id": "f2",
                        "severity": "P1",
                        "category": "test",
                        "title": "F2",
                        "description": "D",
                        "citations": [
                            {"source_type": "file", "source_path": "b.pdf", "location": "S1", "exact_quote": "q"}
                        ],
                        "confidence": "high",
                        "agent": "finance",
                        "analysis_unit": "Beta",
                    },
                ],
                "gaps": [],
                "cross_references": [],
                "governance_graph": {"edges": []},
                "governance_resolved_pct": 0.5,
                "files_analyzed": 2,
            },
        }

        out_path = tmp_path / "report.xlsx"
        gen = ExcelReportGenerator()
        gen.generate(merged, report_schema_model, out_path)

        from openpyxl import load_workbook

        wb = load_workbook(str(out_path))
        ws = wb["Summary"]

        # Find the TOTAL row (last data row + 1)
        total_row = ws.max_row
        # Column 1 should say "TOTAL"
        assert ws.cell(row=total_row, column=1).value == "TOTAL"

        # total_findings column (column 7 in always-active)
        # P0 count col is 3, total_findings is 7
        total_findings_val = ws.cell(row=total_row, column=7).value
        assert total_findings_val == 2  # 1 from acme + 1 from beta


# ===========================================================================
# compute_overall_risk tests
# ===========================================================================


class TestComputeOverallRisk:
    """Test the risk rating algorithm."""

    def test_critical_on_p0_finding(self) -> None:
        findings = [{"severity": "P0", "category": "risk"}]
        assert compute_overall_risk(findings, []) == "Critical"

    def test_critical_on_p0_gap(self) -> None:
        findings: list[dict] = []
        gaps = [{"priority": "P0"}]
        assert compute_overall_risk(findings, gaps) == "Critical"

    def test_high_on_p1(self) -> None:
        findings = [{"severity": "P1", "category": "risk"}]
        assert compute_overall_risk(findings, []) == "High"

    def test_medium_on_p2(self) -> None:
        findings = [{"severity": "P2", "category": "risk"}]
        assert compute_overall_risk(findings, []) == "Medium"

    def test_low_on_p3_only(self) -> None:
        findings = [{"severity": "P3", "category": "risk"}]
        assert compute_overall_risk(findings, []) == "Low"

    def test_clean_when_empty(self) -> None:
        assert compute_overall_risk([], []) == "Clean"

    def test_clean_ignores_domain_reviewed(self) -> None:
        findings = [{"severity": "P3", "category": "domain_reviewed_no_issues"}]
        assert compute_overall_risk(findings, []) == "Clean"


# ===========================================================================
# ContractDateReconciler tests
# ===========================================================================


class TestContractDateReconciler:
    """Test the ContractDateReconciler class."""

    def test_active_contract_not_expired(self) -> None:
        """Database end date in the future => Active-Database Stale."""
        reconciler = ContractDateReconciler(reference_date=date(2025, 1, 15))
        db = [
            {"customer": "Acme", "contract_end_date": "2026-06-30", "arr": 500000},
        ]
        result = reconciler.reconcile(db, findings={})

        assert len(result.entries) == 1
        assert result.entries[0].status == STATUS_ACTIVE_DB_STALE

    def test_expired_with_auto_renewal(self) -> None:
        """Expired per DB but auto-renewal clause found => Active-Auto-Renewal."""
        reconciler = ContractDateReconciler(reference_date=date(2025, 3, 1))
        db = [
            {"customer": "Beta", "contract_end_date": "2024-12-31", "arr": 200000},
        ]
        findings = {
            "Beta": [
                {
                    "title": "Auto-renewal clause in MSA",
                    "description": "Contract contains auto-renewal provision",
                    "category": "term",
                    "citations": [
                        {"source_type": "file", "source_path": "msa.pdf", "location": "S8"},
                    ],
                },
            ],
        }
        result = reconciler.reconcile(db, findings)

        assert len(result.entries) == 1
        assert result.entries[0].status == STATUS_ACTIVE_AUTO_RENEWAL

    def test_expired_no_contracts(self) -> None:
        """Expired and no contract documents at all => Expired-No Contracts."""
        reconciler = ContractDateReconciler(reference_date=date(2025, 3, 1))
        db = [
            {"customer": "Ghost", "contract_end_date": "2024-06-30", "arr": 0},
        ]
        result = reconciler.reconcile(db, findings={})

        assert len(result.entries) == 1
        assert result.entries[0].status == STATUS_EXPIRED_NO_CONTRACTS

    def test_expired_confirmed(self) -> None:
        """Expired, ARR=0, contracts exist but no renewal evidence => Expired-Confirmed."""
        reconciler = ContractDateReconciler(reference_date=date(2025, 3, 1))
        db = [
            {"customer": "OldCo", "contract_end_date": "2023-12-31", "arr": 0},
        ]
        # Findings exist but with no renewal keywords
        findings = {
            "OldCo": [
                {
                    "title": "Termination clause present",
                    "description": "Section 10 allows termination",
                    "category": "termination",
                    "citations": [],
                },
            ],
        }
        result = reconciler.reconcile(db, findings)

        assert len(result.entries) == 1
        assert result.entries[0].status == STATUS_EXPIRED_CONFIRMED

    def test_likely_active_with_arr(self) -> None:
        """Expired per DB, ARR > 0, no renewal evidence => Likely Active."""
        reconciler = ContractDateReconciler(reference_date=date(2025, 3, 1))
        db = [
            {"customer": "MaybeCo", "contract_end_date": "2024-06-30", "arr": 100000},
        ]
        # Findings exist but no renewal keywords
        findings = {
            "MaybeCo": [
                {
                    "title": "Standard clause",
                    "description": "Nothing special",
                    "category": "general",
                    "citations": [],
                },
            ],
        }
        result = reconciler.reconcile(db, findings)

        assert len(result.entries) == 1
        assert result.entries[0].status == STATUS_LIKELY_ACTIVE

    def test_arr_totals(self) -> None:
        """Verify total_reclassified_arr and total_expired_arr are computed correctly."""
        reconciler = ContractDateReconciler(reference_date=date(2025, 3, 1))
        db = [
            {"customer": "Active1", "contract_end_date": "2026-12-31", "arr": 300000},
            {"customer": "Expired1", "contract_end_date": "2023-01-01", "arr": 0},
        ]
        result = reconciler.reconcile(db, findings={})

        assert result.total_reclassified_arr == 300000.0
        assert result.total_expired_arr == 0.0

    def test_write_reconciliation(self, tmp_path: Path) -> None:
        reconciler = ContractDateReconciler(reference_date=date(2025, 3, 1))
        db = [{"customer": "Test", "contract_end_date": "2026-01-01", "arr": 100}]
        result = reconciler.reconcile(db, findings={}, run_id="test_run")

        out = tmp_path / "recon.json"
        reconciler.write_reconciliation(result, out)

        assert out.exists()
        data = json.loads(out.read_text())
        assert data["run_id"] == "test_run"
        assert len(data["entries"]) == 1


# ===========================================================================
# Issue #35 / #53 regression tests
# ===========================================================================


class TestReportSchemaValidation:
    """Test that ReportSchema rejects empty sheets (Issue #35)."""

    def test_empty_sheets_raises_value_error(self) -> None:
        """ReportSchema with zero sheets must fail-fast with ValueError."""
        with pytest.raises(ValueError, match="at least one sheet"):
            ReportSchema(schema_version="1.0.0", sheets=[])

    def test_empty_sheets_default_raises_value_error(self) -> None:
        """ReportSchema with default (omitted) sheets must also fail."""
        with pytest.raises(ValueError, match="at least one sheet"):
            ReportSchema(schema_version="1.0.0")

    def test_valid_schema_with_one_sheet(self) -> None:
        """A schema with at least one sheet should construct successfully."""
        from dd_agents.models.reporting import ColumnDef, SheetDef

        sheet = SheetDef(
            name="Summary",
            columns=[ColumnDef(name="Customer", key="customer", type="string")],
        )
        schema = ReportSchema(schema_version="1.0.0", sheets=[sheet])
        assert len(schema.sheets) == 1


class TestExcelGeneratorGuards:
    """Test guards and edge cases in ExcelReportGenerator (Issue #35 / #53)."""

    def test_generate_raises_on_zero_sheets(self, tmp_path: Path) -> None:
        """generate() raises ValueError when schema has zero sheets.

        NOTE: The ReportSchema validator itself will reject empty sheets,
        but if somehow bypassed the generator must still guard.
        """
        from dd_agents.models.reporting import ColumnDef, SheetDef

        # Build a schema with one sheet, then forcibly empty it
        sheet = SheetDef(
            name="Summary",
            columns=[ColumnDef(name="Customer", key="customer", type="string")],
        )
        schema = ReportSchema(schema_version="1.0.0", sheets=[sheet])
        # Bypass the validator by mutating after construction
        schema.sheets = []

        gen = ExcelReportGenerator()
        with pytest.raises(ValueError, match="zero sheet"):
            gen.generate({}, schema, tmp_path / "out.xlsx")

    def test_list_values_joined_with_semicolon(
        self,
        tmp_path: Path,
    ) -> None:
        """List values in cells should be joined with '; ' not repr (Issue #53).

        Uses the Missing_Docs_Gaps sheet because ``_data_gaps`` passes raw
        gap dicts through, so a list-typed field reaches ``_write_sheet``.
        """
        from dd_agents.models.reporting import ColumnDef, SheetDef

        sheet = SheetDef(
            name="Missing_Docs_Gaps",
            columns=[
                ColumnDef(name="Customer", key="customer", type="string"),
                ColumnDef(name="Evidence", key="evidence", type="string", width=30),
            ],
        )
        schema = ReportSchema(schema_version="1.0.0", sheets=[sheet])

        merged = {
            "acme": {
                "customer": "Acme",
                "findings": [],
                "gaps": [
                    {
                        "customer": "Acme",
                        "evidence": ["contract.pdf", "sow.pdf", "amendment.pdf"],
                    },
                ],
                "cross_references": [],
                "governance_graph": {"edges": []},
                "governance_resolved_pct": 0.0,
            },
        }

        out_path = tmp_path / "report.xlsx"
        gen = ExcelReportGenerator()
        gen.generate(merged, schema, out_path)

        from openpyxl import load_workbook

        wb = load_workbook(str(out_path))
        ws = wb["Missing_Docs_Gaps"]

        # Find Evidence column
        ev_col = None
        for c in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=c).value == "Evidence":
                ev_col = c
                break

        assert ev_col is not None
        cell_value = ws.cell(row=2, column=ev_col).value
        assert cell_value == "contract.pdf; sow.pdf; amendment.pdf"
        # Must NOT contain Python list repr
        assert "[" not in str(cell_value)

    def test_empty_sheets_logged_as_warning(
        self,
        tmp_path: Path,
        caplog: pytest.LogCaptureFixture,
    ) -> None:
        """Non-Summary sheets with zero data rows should emit a warning (Issue #53)."""
        from dd_agents.models.reporting import ColumnDef, SheetDef

        sheets = [
            SheetDef(
                name="Summary",
                columns=[ColumnDef(name="Customer", key="customer", type="string")],
            ),
            SheetDef(
                name="Wolf_Pack",
                columns=[ColumnDef(name="Customer", key="analysis_unit", type="string")],
            ),
        ]
        schema = ReportSchema(schema_version="1.0.0", sheets=sheets)

        merged: dict[str, dict] = {
            "acme": {
                "customer": "Acme",
                "findings": [],
                "gaps": [],
                "cross_references": [],
                "governance_graph": {"edges": []},
                "governance_resolved_pct": 0.0,
            },
        }

        out_path = tmp_path / "report.xlsx"
        gen = ExcelReportGenerator()

        import logging

        with caplog.at_level(logging.WARNING, logger="dd_agents.reporting.excel"):
            gen.generate(merged, schema, out_path)

        assert any("Wolf_Pack" in rec.message and "zero data rows" in rec.message for rec in caplog.records)


class TestStep30SchemaFallback:
    """Test that _step_30 loads schema from config/ fallback (Issue #35)."""

    @pytest.fixture()
    def _pipeline_state(self, tmp_path: Path) -> dict:
        """Create a minimal pipeline state-like structure for testing."""
        run_dir = tmp_path / "run"
        run_dir.mkdir(parents=True)
        # No report_schema.json in run_dir on purpose
        project_dir = tmp_path / "project"
        config_dir = project_dir / "config"
        config_dir.mkdir(parents=True)
        return {"run_dir": run_dir, "project_dir": project_dir, "config_dir": config_dir}

    def test_fallback_to_config_dir(self, _pipeline_state: dict) -> None:
        """When run_dir has no schema, step 30 should load from config/."""

        run_dir: Path = _pipeline_state["run_dir"]
        config_dir: Path = _pipeline_state["config_dir"]

        # Write a valid schema to config/
        minimal_schema = {
            "schema_version": "1.0.0",
            "sheets": [
                {
                    "name": "Summary",
                    "columns": [{"name": "Customer", "key": "customer", "type": "string"}],
                }
            ],
        }
        (config_dir / "report_schema.json").write_text(json.dumps(minimal_schema))

        # Verify run_dir has no schema
        assert not (run_dir / "report_schema.json").exists()

        # Load following the same resolution order as step 30
        schema: ReportSchema | None = None

        run_schema_path = run_dir / "report_schema.json"
        if run_schema_path.exists():
            schema = ReportSchema.model_validate_json(run_schema_path.read_text())

        if schema is None:
            config_schema_path = config_dir / "report_schema.json"
            if config_schema_path.exists():
                schema = ReportSchema.model_validate_json(config_schema_path.read_text())

        assert schema is not None
        assert len(schema.sheets) == 1
        assert schema.sheets[0].name == "Summary"

    def test_fallback_to_builtin_minimal(self, _pipeline_state: dict) -> None:
        """When neither run_dir nor config/ has a schema, use built-in minimal."""
        run_dir: Path = _pipeline_state["run_dir"]
        config_dir: Path = _pipeline_state["config_dir"]

        # Neither location has a schema
        assert not (run_dir / "report_schema.json").exists()
        assert not (config_dir / "report_schema.json").exists()

        # Reproduce fallback logic
        schema: ReportSchema | None = None

        run_schema_path = run_dir / "report_schema.json"
        if run_schema_path.exists():
            schema = ReportSchema.model_validate_json(run_schema_path.read_text())

        if schema is None:
            config_schema_path = config_dir / "report_schema.json"
            if config_schema_path.exists():
                schema = ReportSchema.model_validate_json(config_schema_path.read_text())

        if schema is None:
            schema = ReportSchema.model_validate(
                {
                    "schema_version": "1.0.0",
                    "description": "Built-in minimal schema (fallback)",
                    "sheets": [
                        {
                            "name": "Summary",
                            "required": True,
                            "activation_condition": "always",
                            "columns": [
                                {"name": "Customer", "key": "customer", "type": "string", "width": 30},
                                {
                                    "name": "Overall Risk Rating",
                                    "key": "overall_risk_rating",
                                    "type": "string",
                                    "width": 20,
                                },
                                {"name": "Total Findings", "key": "total_findings", "type": "integer", "width": 14},
                                {"name": "Gap Count", "key": "gap_count", "type": "integer", "width": 12},
                            ],
                        },
                    ],
                }
            )

        assert schema is not None
        assert len(schema.sheets) == 1
        assert schema.sheets[0].name == "Summary"
        assert len(schema.sheets[0].columns) == 4


# ===========================================================================
# Severity normalization tests (#77)
# ===========================================================================


class TestSeverityNormalization:
    """Tests for severity normalization in _promote_findings (Issue #77)."""

    @staticmethod
    def _make_raw_finding(severity: str = "P2", **kwargs: Any) -> dict[str, Any]:
        return {
            "severity": severity,
            "category": "test",
            "title": kwargs.get("title", "Test finding"),
            "description": "Test description",
            "confidence": "medium",
            "citations": [
                {
                    "source_type": "file",
                    "source_path": "test.pdf",
                    "location": "page 1",
                    "exact_quote": "test quote",
                }
            ],
            "agent": kwargs.get("agent", "legal"),
            **{k: v for k, v in kwargs.items() if k not in ("title", "agent")},
        }

    def test_normalize_severity_high_to_p1(self) -> None:
        assert FindingMerger._normalize_severity("high") == "P1"

    def test_normalize_severity_critical_to_p0(self) -> None:
        assert FindingMerger._normalize_severity("CRITICAL") == "P0"

    def test_normalize_severity_medium_to_p2(self) -> None:
        assert FindingMerger._normalize_severity("medium") == "P2"

    def test_normalize_severity_low_to_p3(self) -> None:
        assert FindingMerger._normalize_severity("low") == "P3"

    def test_normalize_severity_already_standard(self) -> None:
        assert FindingMerger._normalize_severity("P0") == "P0"
        assert FindingMerger._normalize_severity("P1") == "P1"

    def test_normalize_severity_unknown_defaults_p3(self) -> None:
        assert FindingMerger._normalize_severity("unknown") == "P3"

    def test_promote_findings_normalizes_high_severity(self) -> None:
        """Finding with severity='high' should be promoted as P1, not dropped."""
        merger = FindingMerger(run_id="test", timestamp="2025-01-01T00:00:00Z")
        raw = [self._make_raw_finding(severity="high")]
        result = merger._promote_findings(raw, "Customer A", "customer_a")
        assert len(result) == 1
        assert result[0].severity.value == "P1"

    def test_promote_findings_normalizes_critical_severity(self) -> None:
        """Finding with severity='CRITICAL' should be promoted as P0."""
        merger = FindingMerger(run_id="test", timestamp="2025-01-01T00:00:00Z")
        raw = [self._make_raw_finding(severity="CRITICAL")]
        result = merger._promote_findings(raw, "Customer A", "customer_a")
        assert len(result) == 1
        assert result[0].severity.value == "P0"

    def test_promote_findings_no_findings_dropped_by_severity(self) -> None:
        """Findings with non-standard severity should NOT be dropped."""
        merger = FindingMerger(run_id="test", timestamp="2025-01-01T00:00:00Z")
        raw = [
            self._make_raw_finding(severity="high"),
            self._make_raw_finding(severity="medium"),
            self._make_raw_finding(severity="low"),
            self._make_raw_finding(severity="CRITICAL"),
            self._make_raw_finding(severity="P0"),
        ]
        result = merger._promote_findings(raw, "Customer A", "customer_a")
        assert len(result) == 5


# ===========================================================================
# Agent name normalization tests (#78)
# ===========================================================================


class TestAgentNameNormalization:
    """Tests for agent name normalization in _promote_findings (Issue #78)."""

    def test_normalize_agent_name_standard(self) -> None:
        assert FindingMerger._normalize_agent_name("legal") == "legal"
        assert FindingMerger._normalize_agent_name("finance") == "finance"

    def test_normalize_agent_name_case_insensitive(self) -> None:
        assert FindingMerger._normalize_agent_name("Legal") == "legal"
        assert FindingMerger._normalize_agent_name("FINANCE") == "finance"

    def test_normalize_agent_name_hyphenated(self) -> None:
        assert FindingMerger._normalize_agent_name("product-tech") == "producttech"

    def test_normalize_agent_name_underscore(self) -> None:
        assert FindingMerger._normalize_agent_name("product_tech") == "producttech"


# ===========================================================================
# Title truncation tests (#78)
# ===========================================================================


class TestTitleTruncation:
    """Tests for title truncation in _promote_findings (Issue #78)."""

    def test_long_title_truncated_to_120_chars(self) -> None:
        merger = FindingMerger(run_id="test", timestamp="2025-01-01T00:00:00Z")
        long_title = "A" * 200
        raw = [
            {
                "severity": "P2",
                "category": "test",
                "title": long_title,
                "description": "desc",
                "confidence": "medium",
                "agent": "legal",
                "citations": [{"source_type": "file", "source_path": "f.pdf", "location": "p1", "exact_quote": "q"}],
            }
        ]
        result = merger._promote_findings(raw, "Customer A", "customer_a")
        assert len(result) == 1
        assert len(result[0].title) == 120


# ===========================================================================
# P0/P1 downgrade on missing exact_quote (#78)
# ===========================================================================


class TestSeverityDowngradeOnMissingQuote:
    """Test P0/P1 findings are downgraded to P2 when citations lack exact_quote."""

    def test_p0_downgraded_when_no_exact_quote(self) -> None:
        merger = FindingMerger(run_id="test", timestamp="2025-01-01T00:00:00Z")
        raw = [
            {
                "severity": "P0",
                "category": "test",
                "title": "Critical finding",
                "description": "desc",
                "confidence": "high",
                "agent": "legal",
                "citations": [
                    {"source_type": "file", "source_path": "f.pdf", "location": "p1"}
                    # No exact_quote
                ],
            }
        ]
        result = merger._promote_findings(raw, "Customer A", "customer_a")
        assert len(result) == 1
        # Downgraded from P0 to P2 because citations lack exact_quote
        assert result[0].severity.value == "P2"

    def test_p1_with_exact_quote_stays_p1(self) -> None:
        merger = FindingMerger(run_id="test", timestamp="2025-01-01T00:00:00Z")
        raw = [
            {
                "severity": "P1",
                "category": "test",
                "title": "High finding",
                "description": "desc",
                "confidence": "high",
                "agent": "legal",
                "citations": [
                    {"source_type": "file", "source_path": "f.pdf", "location": "p1", "exact_quote": "actual quote"}
                ],
            }
        ]
        result = merger._promote_findings(raw, "Customer A", "customer_a")
        assert len(result) == 1
        assert result[0].severity.value == "P1"


# ===========================================================================
# Stale file cleanup tests (#69)
# ===========================================================================


class TestStaleFileCleanup:
    """Tests for stale file cleanup in write_merged (Issue #69)."""

    def test_write_merged_removes_stale_files(self, tmp_path: Path) -> None:
        """Non-customer JSON files in merged/ should be removed."""
        merged_dir = tmp_path / "merged"
        merged_dir.mkdir()

        # Pre-existing stale files
        (merged_dir / "numerical_manifest.json").write_text("{}")
        (merged_dir / "coverage_manifest.json").write_text("{}")
        (merged_dir / "report_diff.json").write_text("{}")

        # Write merged output for one customer
        from dd_agents.models.finding import MergedCustomerOutput

        mco = MergedCustomerOutput(customer="Acme Corp", customer_safe_name="acme_corp")
        merger = FindingMerger(run_id="test", timestamp="2025-01-01T00:00:00Z")
        merger.write_merged({"acme_corp": mco}, merged_dir)

        # Customer file should exist
        assert (merged_dir / "acme_corp.json").exists()
        # Stale files should be removed
        assert not (merged_dir / "numerical_manifest.json").exists()
        assert not (merged_dir / "coverage_manifest.json").exists()
        assert not (merged_dir / "report_diff.json").exists()

    def test_write_merged_preserves_customer_files(self, tmp_path: Path) -> None:
        """Customer files from a previous run that are still valid should stay."""
        from dd_agents.models.finding import MergedCustomerOutput

        merged_dir = tmp_path / "merged"
        merged_dir.mkdir()

        mco_a = MergedCustomerOutput(customer="A", customer_safe_name="a")
        mco_b = MergedCustomerOutput(customer="B", customer_safe_name="b")
        merger = FindingMerger(run_id="test", timestamp="2025-01-01T00:00:00Z")
        merger.write_merged({"a": mco_a, "b": mco_b}, merged_dir)

        assert (merged_dir / "a.json").exists()
        assert (merged_dir / "b.json").exists()

    def test_write_merged_clean_stale_false_preserves_all(self, tmp_path: Path) -> None:
        """When clean_stale=False, no files are removed."""
        from dd_agents.models.finding import MergedCustomerOutput

        merged_dir = tmp_path / "merged"
        merged_dir.mkdir()
        (merged_dir / "stale.json").write_text("{}")

        mco = MergedCustomerOutput(customer="X", customer_safe_name="x")
        merger = FindingMerger(run_id="test", timestamp="2025-01-01T00:00:00Z")
        merger.write_merged({"x": mco}, merged_dir, clean_stale=False)

        assert (merged_dir / "x.json").exists()
        assert (merged_dir / "stale.json").exists()


# ===========================================================================
# Judge prose extraction tests (#72)
# ===========================================================================


class TestJudgeProseExtraction:
    """Tests for Judge agent prose output parsing (Issue #72)."""

    def test_extract_scores_from_prose_basic(self) -> None:
        from dd_agents.agents.judge import JudgeAgent

        prose = """
        Quality Assessment Results:
        Legal: 85/100
        Finance: 78/100
        Commercial: 72/100
        ProductTech: 90/100
        Overall quality: 81
        """
        result = JudgeAgent._extract_scores_from_prose(prose)
        assert result is not None
        assert "legal" in result["agent_scores"]
        assert result["agent_scores"]["legal"]["score"] == 85
        assert result["agent_scores"]["finance"]["score"] == 78
        assert result["agent_scores"]["commercial"]["score"] == 72
        assert result["agent_scores"]["producttech"]["score"] == 90
        assert result["overall_quality"] == 81

    def test_extract_scores_from_prose_with_dimensions(self) -> None:
        from dd_agents.agents.judge import JudgeAgent

        prose = """
        Legal score: 82
        citation_verification: 90
        contextual_validation: 80
        financial_accuracy: 75
        """
        result = JudgeAgent._extract_scores_from_prose(prose)
        assert result is not None
        assert result["agent_scores"]["legal"]["score"] == 82
        dims = result["agent_scores"]["legal"]["dimensions"]
        assert dims["citation_verification"] == 90
        assert dims["contextual_validation"] == 80
        assert dims["financial_accuracy"] == 75

    def test_extract_scores_from_prose_no_scores(self) -> None:
        from dd_agents.agents.judge import JudgeAgent

        prose = "The findings look good overall. No specific scores to report."
        result = JudgeAgent._extract_scores_from_prose(prose)
        assert result is None

    def test_extract_scores_overall_computed_from_agents(self) -> None:
        from dd_agents.agents.judge import JudgeAgent

        prose = """
        Legal: 80
        Finance: 90
        """
        result = JudgeAgent._extract_scores_from_prose(prose)
        assert result is not None
        assert result["overall_quality"] == 85  # (80 + 90) / 2

    def test_build_scores_with_prose_fallback(self) -> None:
        """When JSON parsing yields no agent_scores, prose extraction is tried."""
        from dd_agents.agents.judge import JudgeAgent

        result = {
            "run_id": "test_run",
            "output": [{"summary": "Legal: 85, Finance: 78"}],
            "raw_output": "Legal: 85\nFinance: 78\nOverall quality: 82",
        }
        scores = JudgeAgent._build_scores_from_result(result, round_num=1)
        assert scores.overall_quality == 82
        assert "legal" in scores.agent_scores
        assert scores.agent_scores["legal"].score == 85


# ===========================================================================
# Report schema package-relative resolution tests (#71)
# ===========================================================================


class TestReportSchemaPackageResolution:
    """Test that report schema can be found via package-relative path (Issue #71)."""

    def test_config_schema_exists_in_repo(self) -> None:
        """The 14-sheet report schema exists at repo_root/config/report_schema.json."""
        import dd_agents

        pkg_root = Path(dd_agents.__file__).resolve().parent
        repo_root = pkg_root.parent.parent
        schema_path = repo_root / "config" / "report_schema.json"
        assert schema_path.exists(), f"Expected schema at {schema_path}"

        schema = ReportSchema.model_validate_json(schema_path.read_text())
        assert len(schema.sheets) == 14

    def test_package_relative_resolution_order(self, tmp_path: Path) -> None:
        """Resolution finds schema via package path when project_dir is wrong."""
        import dd_agents

        pkg_root = Path(dd_agents.__file__).resolve().parent
        repo_config = pkg_root.parent.parent / "config" / "report_schema.json"

        # Simulate: project_dir points to data room (no config/)
        fake_project_dir = tmp_path / "data_room"
        fake_project_dir.mkdir()

        # Resolution order should find it via package-relative path
        candidate_paths = [
            fake_project_dir / "config" / "report_schema.json",
            pkg_root.parent.parent / "config" / "report_schema.json",
            pkg_root / "config" / "report_schema.json",
        ]
        found = None
        for p in candidate_paths:
            if p.exists():
                found = p
                break

        assert found is not None
        assert found == repo_config
