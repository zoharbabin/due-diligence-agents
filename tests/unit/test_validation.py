"""Unit tests for the dd_agents.validation module.

Tests cover:
- CoverageValidator: correct count passes, missing customer fails, empty dir
- NumericalAuditor: source traceability, cross-source consistency, semantic reasonableness
- NumericalManifest: generated_at field validation
- NumericalAuditor._rederive: N003-N010 rederivation from source files
- QAAuditor: basic audit with sample data
- DefinitionOfDoneChecker: check count, conditional check grouping
- SchemaValidator: validate matching Excel passes, mismatched fails
"""

from __future__ import annotations

import json
from typing import TYPE_CHECKING

import pytest

from dd_agents.models.audit import AuditCheck, AuditReport
from dd_agents.models.numerical import ManifestEntry, NumericalManifest
from dd_agents.validation.coverage import CoverageValidator
from dd_agents.validation.dod import DefinitionOfDoneChecker
from dd_agents.validation.numerical_audit import NumericalAuditor
from dd_agents.validation.qa_audit import QAAuditor
from dd_agents.validation.schema_validator import SchemaValidator

if TYPE_CHECKING:
    from pathlib import Path

# ===================================================================== #
# Helpers for building sample data
# ===================================================================== #

CUSTOMERS = ["acme_corp", "globex", "initech"]
AGENTS = ["legal", "finance", "commercial", "producttech"]


def _make_customer_json(customer: str, agent: str) -> dict:
    """Minimal valid per-customer agent output."""
    return {
        "customer": customer,
        "customer_safe_name": customer,
        "agent": agent,
        "run_id": "run_001",
        "timestamp": "2025-02-18T00:00:00Z",
        "files_analyzed": 1,
        "findings": [
            {
                "id": f"forensic-dd_{agent}_{customer}_0001",
                "severity": "P2",
                "category": "test_category",
                "title": "Test finding",
                "description": "Test description",
                "citations": [
                    {
                        "source_type": "file",
                        "source_path": f"{customer}/contract.pdf",
                        "exact_quote": "test quote",
                    }
                ],
                "confidence": "high",
                "agent": agent,
            }
        ],
        "gaps": [],
    }


def _make_merged_json(customer: str) -> dict:
    """Minimal merged per-customer output."""
    return {
        "customer": customer,
        "customer_safe_name": customer,
        "findings": [
            {
                "id": f"forensic-dd_legal_{customer}_0001",
                "severity": "P2",
                "category": "test_category",
                "title": "Test finding",
                "description": "Test description",
                "citations": [
                    {
                        "source_type": "file",
                        "source_path": f"{customer}/contract.pdf",
                        "exact_quote": "test quote",
                    }
                ],
                "confidence": "high",
                "agent": "legal",
            }
        ],
        "gaps": [],
        "cross_references": [],
        "governance_graph": {"edges": []},
    }


def _make_coverage_manifest(agent: str, customers: list[str]) -> dict:
    """Minimal coverage manifest."""
    return {
        "agent": agent,
        "skill": "forensic-dd",
        "run_id": "run_001",
        "files_assigned": [f"{c}/contract.pdf" for c in customers],
        "files_read": [{"path": f"{c}/contract.pdf"} for c in customers],
        "files_skipped": [],
        "files_failed": [],
        "coverage_pct": 1.0,
        "analysis_units_assigned": len(customers),
        "analysis_units_completed": len(customers),
        "customers": [
            {
                "name": c,
                "files_assigned": [f"{c}/contract.pdf"],
                "files_processed": [f"{c}/contract.pdf"],
                "status": "complete",
            }
            for c in customers
        ],
    }


def _make_manifest(
    n001: int = 3,
    n003: int = 10,
    n004: int = 1,
    n005: int = 2,
    n006: int = 3,
    n007: int = 4,
    n008: int = 0,
    n009: int = 5,
) -> NumericalManifest:
    """Build a NumericalManifest with the given values."""
    return NumericalManifest(
        generated_at="2025-02-18T00:00:00Z",
        numbers=[
            ManifestEntry(
                id="N001", label="total_customers", value=n001, source_file="customers.csv", derivation="COUNT(rows)"
            ),
            ManifestEntry(id="N002", label="total_files", value=20, source_file="files.txt", derivation="wc -l"),
            ManifestEntry(
                id="N003", label="total_findings", value=n003, source_file="findings/*.json", derivation="SUM(findings)"
            ),
            ManifestEntry(
                id="N004", label="findings_p0", value=n004, source_file="findings/*.json", derivation="COUNT(P0)"
            ),
            ManifestEntry(
                id="N005", label="findings_p1", value=n005, source_file="findings/*.json", derivation="COUNT(P1)"
            ),
            ManifestEntry(
                id="N006", label="findings_p2", value=n006, source_file="findings/*.json", derivation="COUNT(P2)"
            ),
            ManifestEntry(
                id="N007", label="findings_p3", value=n007, source_file="findings/*.json", derivation="COUNT(P3)"
            ),
            ManifestEntry(
                id="N008",
                label="clean_result_count",
                value=n008,
                source_file="findings/*.json",
                derivation="COUNT(clean)",
            ),
            ManifestEntry(
                id="N009", label="total_gaps", value=n009, source_file="gaps/*.json", derivation="COUNT(gaps)"
            ),
            ManifestEntry(
                id="N010",
                label="total_reference_files",
                value=5,
                source_file="reference_files.json",
                derivation="len(files)",
            ),
        ],
    )


def _populate_run_dir(
    run_dir: Path,
    inventory_dir: Path,
    customers: list[str] | None = None,
) -> None:
    """Populate a run directory with minimal valid data for QA checks."""
    customers = customers or CUSTOMERS

    # Inventory files
    inventory_dir.mkdir(parents=True, exist_ok=True)
    files_list = [f"{c}/contract.pdf" for c in customers]
    (inventory_dir / "files.txt").write_text("\n".join(files_list) + "\n")
    (inventory_dir / "customers.csv").write_text(
        "group,name,safe_name,path,file_count\n" + "\n".join(f"group1,{c},{c},/data/{c},1" for c in customers) + "\n"
    )
    (inventory_dir / "counts.json").write_text(
        json.dumps({"total_customers": len(customers), "total_files": len(files_list)})
    )
    (inventory_dir / "extraction_quality.json").write_text("[]")

    # Per-agent output directories
    for agent in AGENTS:
        agent_dir = run_dir / "findings" / agent
        agent_dir.mkdir(parents=True, exist_ok=True)
        for customer in customers:
            (agent_dir / f"{customer}.json").write_text(json.dumps(_make_customer_json(customer, agent)))
        (agent_dir / "coverage_manifest.json").write_text(json.dumps(_make_coverage_manifest(agent, customers)))

    # Merged output
    merged_dir = run_dir / "findings" / "merged"
    merged_dir.mkdir(parents=True, exist_ok=True)
    for customer in customers:
        (merged_dir / f"{customer}.json").write_text(json.dumps(_make_merged_json(customer)))

    # Audit logs
    for agent in [*AGENTS, "reporting_lead"]:
        log_dir = run_dir / "audit" / agent
        log_dir.mkdir(parents=True, exist_ok=True)
        (log_dir / "audit_log.jsonl").write_text(
            json.dumps(
                {
                    "ts": "2025-02-18T00:00:00Z",
                    "agent": agent,
                    "skill": "forensic-dd",
                    "action": "phase_complete",
                    "target": "analysis",
                    "result": "complete",
                }
            )
            + "\n"
        )

    # Report directory with dummy Excel
    report_dir = run_dir / "report"
    report_dir.mkdir(parents=True, exist_ok=True)


# ===================================================================== #
# CoverageValidator Tests
# ===================================================================== #


class TestCoverageValidator:
    """Tests for CoverageValidator."""

    def test_correct_count_passes(self, tmp_path: Path) -> None:
        """When every customer has output from every agent, all checks pass."""
        agent_dirs: dict[str, Path] = {}
        for agent in AGENTS:
            d = tmp_path / agent
            d.mkdir()
            for customer in CUSTOMERS:
                (d / f"{customer}.json").write_text(json.dumps(_make_customer_json(customer, agent)))
            agent_dirs[agent] = d

        validator = CoverageValidator()
        results = validator.validate(agent_dirs, CUSTOMERS)

        # 4 per-agent checks + 1 aggregate = 5
        assert len(results) == 5
        for check in results:
            assert isinstance(check, AuditCheck)
            assert check.passed is True

    def test_missing_customer_fails(self, tmp_path: Path) -> None:
        """Missing a customer JSON for one agent causes failure."""
        agent_dirs: dict[str, Path] = {}
        for agent in AGENTS:
            d = tmp_path / agent
            d.mkdir()
            for customer in CUSTOMERS:
                if agent == "legal" and customer == "initech":
                    continue  # skip one
                (d / f"{customer}.json").write_text(json.dumps(_make_customer_json(customer, agent)))
            agent_dirs[agent] = d

        validator = CoverageValidator()
        results = validator.validate(agent_dirs, CUSTOMERS)

        # The legal agent check should fail
        legal_check = next(r for r in results if r.details.get("agent") == "legal")
        assert legal_check.passed is False
        assert "initech" in legal_check.details["missing_customers"]

        # Aggregate should also fail
        aggregate = next(r for r in results if r.details.get("aggregate"))
        assert aggregate.passed is False

    def test_empty_dir_fails(self, tmp_path: Path) -> None:
        """An empty agent output directory means all customers are missing."""
        agent_dirs: dict[str, Path] = {}
        for agent in AGENTS:
            d = tmp_path / agent
            d.mkdir()
            # Do not create any JSON files
            agent_dirs[agent] = d

        validator = CoverageValidator()
        results = validator.validate(agent_dirs, CUSTOMERS)

        for check in results:
            if check.details.get("aggregate"):
                assert check.passed is False
            elif "agent" in check.details:
                assert check.passed is False
                assert len(check.details["missing_customers"]) == len(CUSTOMERS)

    def test_missing_agent_dir(self, tmp_path: Path) -> None:
        """When an agent directory does not exist, check fails."""
        agent_dirs: dict[str, Path] = {}
        for agent in AGENTS:
            if agent == "finance":
                agent_dirs[agent] = tmp_path / "nonexistent"
            else:
                d = tmp_path / agent
                d.mkdir()
                for customer in CUSTOMERS:
                    (d / f"{customer}.json").write_text(json.dumps(_make_customer_json(customer, agent)))
                agent_dirs[agent] = d

        validator = CoverageValidator()
        results = validator.validate(agent_dirs, CUSTOMERS)

        finance_check = next(r for r in results if r.details.get("agent") == "finance")
        assert finance_check.passed is False
        assert finance_check.details["error"] == "output directory missing"

    def test_empty_json_file_detected(self, tmp_path: Path) -> None:
        """A JSON file with no findings/gaps is flagged as empty."""
        d = tmp_path / "legal"
        d.mkdir()
        for customer in CUSTOMERS:
            data = {"customer": customer, "findings": [], "gaps": []}
            (d / f"{customer}.json").write_text(json.dumps(data))

        agent_dirs = {"legal": d}
        # Only check legal to isolate the test
        validator = CoverageValidator()
        results = validator.validate(agent_dirs, CUSTOMERS)

        legal_check = next(r for r in results if r.details.get("agent") == "legal")
        assert len(legal_check.details["empty_files"]) == len(CUSTOMERS)


# ===================================================================== #
# NumericalAuditor Tests
# ===================================================================== #


class TestNumericalAuditor:
    """Tests for NumericalAuditor."""

    def test_source_traceability_passes_with_existing_files(self, tmp_path: Path) -> None:
        """Layer 1 passes when all source files exist."""
        # Create source files that the manifest entries reference
        (tmp_path / "customers.csv").write_text("header\nrow1\nrow2\nrow3\n")
        (tmp_path / "files.txt").write_text("file1\nfile2\n")
        (tmp_path / "reference_files.json").write_text("[]")
        # Create directories with JSON files for glob patterns
        findings_dir = tmp_path / "findings"
        findings_dir.mkdir()
        (findings_dir / "sample.json").write_text("{}")
        gaps_dir = tmp_path / "gaps"
        gaps_dir.mkdir()
        (gaps_dir / "sample.json").write_text("{}")

        manifest = _make_manifest(n001=3)
        # Override source_file to use tmp_path
        for entry in manifest.numbers:
            entry.source_file = str(tmp_path / entry.source_file)

        auditor = NumericalAuditor(run_dir=tmp_path, inventory_dir=tmp_path)
        check = auditor.check_source_traceability(manifest)
        assert check.passed is True
        assert check.details["layer"] == 1

    def test_source_traceability_fails_for_missing_file(self, tmp_path: Path) -> None:
        """Layer 1 fails when a source file is missing."""
        manifest = _make_manifest()
        # Point to nonexistent files
        for entry in manifest.numbers:
            entry.source_file = str(tmp_path / "nonexistent" / entry.source_file)

        auditor = NumericalAuditor(run_dir=tmp_path, inventory_dir=tmp_path)
        check = auditor.check_source_traceability(manifest)
        assert check.passed is False
        assert len(check.details["failures"]) > 0

    def test_source_traceability_fails_for_missing_derivation(self, tmp_path: Path) -> None:
        """Layer 1 fails when derivation is empty."""
        manifest = _make_manifest()
        manifest.numbers[0].derivation = ""
        manifest.numbers[0].source_file = str(tmp_path)  # exists
        for entry in manifest.numbers[1:]:
            entry.source_file = str(tmp_path)

        auditor = NumericalAuditor(run_dir=tmp_path, inventory_dir=tmp_path)
        check = auditor.check_source_traceability(manifest)
        assert check.passed is False
        assert any("missing derivation" in f for f in check.details["failures"])

    def test_source_traceability_accepts_empty_glob_for_zero_value(self, tmp_path: Path) -> None:
        """Layer 1 accepts a glob matching 0 files when the entry value is 0."""
        # Create all non-glob source files so only the glob entry is tested
        (tmp_path / "customers.csv").write_text("header\nrow1\nrow2\nrow3\n")
        (tmp_path / "files.txt").write_text("file1\nfile2\n")
        (tmp_path / "reference_files.json").write_text("[]")
        findings_dir = tmp_path / "findings"
        findings_dir.mkdir()
        (findings_dir / "sample.json").write_text("{}")

        # gaps directory exists but has no JSON files -- N009 value is 0
        gaps_dir = tmp_path / "gaps"
        gaps_dir.mkdir()

        manifest = _make_manifest(n009=0)
        for entry in manifest.numbers:
            entry.source_file = str(tmp_path / entry.source_file)

        auditor = NumericalAuditor(run_dir=tmp_path, inventory_dir=tmp_path)
        check = auditor.check_source_traceability(manifest)
        assert check.passed is True

    def test_source_traceability_rejects_empty_glob_for_nonzero_value(self, tmp_path: Path) -> None:
        """Layer 1 rejects a glob matching 0 files when the entry value is nonzero."""
        # Only create the gaps dir (no JSON files) but N009 value is 5
        manifest = _make_manifest(n009=5)
        for entry in manifest.numbers:
            entry.source_file = str(tmp_path / "nonexistent" / entry.source_file)

        auditor = NumericalAuditor(run_dir=tmp_path, inventory_dir=tmp_path)
        check = auditor.check_source_traceability(manifest)
        assert check.passed is False
        assert any("does not exist" in f for f in check.details["failures"])

    def test_cross_source_consistency_passes(self, tmp_path: Path) -> None:
        """Layer 3 passes when severity sum equals total findings."""
        manifest = _make_manifest(n003=10, n004=1, n005=2, n006=3, n007=4)
        (tmp_path / "customers.csv").write_text("header\n" + "\n".join(f"row{i}" for i in range(3)) + "\n")
        (tmp_path / "counts.json").write_text(json.dumps({"total_customers": 3}))

        auditor = NumericalAuditor(run_dir=tmp_path, inventory_dir=tmp_path)
        check = auditor.check_cross_source_consistency(manifest)
        assert check.passed is True

    def test_cross_source_consistency_fails_on_mismatch(self, tmp_path: Path) -> None:
        """Layer 3 fails when severity sum does not equal total findings."""
        # N003=10 but N004+N005+N006+N007 = 1+2+3+5 = 11
        manifest = _make_manifest(n003=10, n004=1, n005=2, n006=3, n007=5)
        (tmp_path / "customers.csv").write_text("header\n" + "\n".join(f"row{i}" for i in range(3)) + "\n")
        (tmp_path / "counts.json").write_text(json.dumps({"total_customers": 3}))

        auditor = NumericalAuditor(run_dir=tmp_path, inventory_dir=tmp_path)
        check = auditor.check_cross_source_consistency(manifest)
        assert check.passed is False
        assert any("N004+N005+N006+N007" in f for f in check.details["failures"])

    def test_cross_source_csv_vs_counts_mismatch(self, tmp_path: Path) -> None:
        """Layer 3 catches customers.csv vs counts.json mismatch."""
        manifest = _make_manifest(n001=5)
        (tmp_path / "customers.csv").write_text(
            "header\nrow1\nrow2\nrow3\n"  # 3 data rows
        )
        (tmp_path / "counts.json").write_text(
            json.dumps({"total_customers": 5})  # different from CSV
        )

        auditor = NumericalAuditor(run_dir=tmp_path, inventory_dir=tmp_path)
        check = auditor.check_cross_source_consistency(manifest)
        assert check.passed is False

    def test_semantic_reasonableness_passes(self, tmp_path: Path) -> None:
        """Layer 5 passes for reasonable values."""
        manifest = _make_manifest(n003=10, n004=1)
        auditor = NumericalAuditor(run_dir=tmp_path, inventory_dir=tmp_path)
        check = auditor.check_semantic_reasonableness(manifest)
        assert check.passed is True

    def test_semantic_reasonableness_negative_values(self, tmp_path: Path) -> None:
        """Layer 5 fails on negative values."""
        manifest = _make_manifest()
        manifest.numbers[0].value = -5

        auditor = NumericalAuditor(run_dir=tmp_path, inventory_dir=tmp_path)
        check = auditor.check_semantic_reasonableness(manifest)
        assert check.passed is False
        assert any("negative" in f for f in check.details["failures"])

    def test_semantic_reasonableness_p0_exceeds_total(self, tmp_path: Path) -> None:
        """Layer 5 flags P0 count exceeding total findings."""
        manifest = _make_manifest(n003=5, n004=10)  # P0=10 > total=5
        auditor = NumericalAuditor(run_dir=tmp_path, inventory_dir=tmp_path)
        check = auditor.check_semantic_reasonableness(manifest)
        assert check.passed is False
        assert any("P0 findings count > total findings" in f for f in check.details["failures"])

    def test_semantic_reasonableness_customer_count_spike(self, tmp_path: Path) -> None:
        """Layer 5 flags >20% customer count change between runs."""
        prior = _make_manifest(n001=100)
        current = _make_manifest(n001=150)  # 50% change

        auditor = NumericalAuditor(
            run_dir=tmp_path,
            inventory_dir=tmp_path,
            prior_manifest=prior,
        )
        check = auditor.check_semantic_reasonableness(current)
        assert check.passed is False
        assert any("Customer count changed" in f for f in check.details["failures"])

    def test_run_full_audit(self, tmp_path: Path) -> None:
        """Full audit returns 4 checks (layers 1, 2, 3, 5)."""
        manifest = _make_manifest(n003=10, n004=1, n005=2, n006=3, n007=4)
        (tmp_path / "customers.csv").write_text("header\n" + "\n".join(f"row{i}" for i in range(3)) + "\n")
        (tmp_path / "counts.json").write_text(json.dumps({"total_customers": 3}))

        auditor = NumericalAuditor(run_dir=tmp_path, inventory_dir=tmp_path)
        checks = auditor.run_full_audit(manifest)
        assert len(checks) == 4
        layers = [c.details["layer"] for c in checks]
        assert sorted(layers) == [1, 2, 3, 5]

    def test_manifest_with_generated_at_validates(self) -> None:
        """A manifest dict that includes generated_at validates successfully."""
        from pydantic import ValidationError

        manifest_dict = {
            "generated_at": "2025-02-18T00:00:00Z",
            "numbers": [
                {
                    "id": f"N{str(i).zfill(3)}",
                    "label": f"metric_{i}",
                    "value": i,
                    "source_file": f"file_{i}.csv",
                    "derivation": "count",
                }
                for i in range(1, 11)
            ],
        }
        try:
            m = NumericalManifest.model_validate(manifest_dict)
        except ValidationError:
            pytest.fail("NumericalManifest should validate when generated_at is present")
        assert m.generated_at == "2025-02-18T00:00:00Z"
        assert len(m.numbers) == 10

    def test_manifest_without_generated_at_raises_validation_error(self) -> None:
        """A manifest dict missing generated_at raises ValidationError."""
        from pydantic import ValidationError

        manifest_dict = {
            "numbers": [
                {
                    "id": f"N{str(i).zfill(3)}",
                    "label": f"metric_{i}",
                    "value": i,
                    "source_file": f"file_{i}.csv",
                    "derivation": "count",
                }
                for i in range(1, 11)
            ],
        }
        with pytest.raises(ValidationError, match="generated_at"):
            NumericalManifest.model_validate(manifest_dict)

    def test_rederive_n003_counts_merged_findings(self, tmp_path: Path) -> None:
        """N003 rederivation counts total findings from merged dir."""
        merged_dir = tmp_path / "findings" / "merged"
        merged_dir.mkdir(parents=True)
        (merged_dir / "customer_a.json").write_text(
            json.dumps(
                {
                    "findings": [
                        {"severity": "P0", "category": "risk"},
                        {"severity": "P1", "category": "risk"},
                    ]
                }
            )
        )
        (merged_dir / "customer_b.json").write_text(json.dumps({"findings": [{"severity": "P2", "category": "risk"}]}))
        auditor = NumericalAuditor(run_dir=tmp_path, inventory_dir=tmp_path)
        entry = ManifestEntry(
            id="N003", label="Total Findings", value=0, source_file="findings/*.json", derivation="count"
        )
        assert auditor._rederive(entry) == 3

    def test_rederive_n003_excludes_domain_reviewed_no_issues(self, tmp_path: Path) -> None:
        """N003 rederivation excludes findings with category domain_reviewed_no_issues."""
        merged_dir = tmp_path / "findings" / "merged"
        merged_dir.mkdir(parents=True)
        (merged_dir / "customer_a.json").write_text(
            json.dumps(
                {
                    "findings": [
                        {"severity": "P0", "category": "risk"},
                        {"severity": "P1", "category": "domain_reviewed_no_issues"},
                    ]
                }
            )
        )
        auditor = NumericalAuditor(run_dir=tmp_path, inventory_dir=tmp_path)
        entry = ManifestEntry(
            id="N003", label="Total Findings", value=0, source_file="findings/*.json", derivation="count"
        )
        assert auditor._rederive(entry) == 1

    def test_rederive_n004_n007_counts_by_severity(self, tmp_path: Path) -> None:
        """N004-N007 rederivation counts findings by severity P0-P3."""
        merged_dir = tmp_path / "findings" / "merged"
        merged_dir.mkdir(parents=True)
        (merged_dir / "customer_a.json").write_text(
            json.dumps(
                {
                    "findings": [
                        {"severity": "P0", "category": "risk"},
                        {"severity": "P0", "category": "risk"},
                        {"severity": "P1", "category": "risk"},
                        {"severity": "P2", "category": "risk"},
                        {"severity": "P2", "category": "risk"},
                        {"severity": "P2", "category": "risk"},
                        {"severity": "P3", "category": "risk"},
                    ]
                }
            )
        )
        auditor = NumericalAuditor(run_dir=tmp_path, inventory_dir=tmp_path)
        assert (
            auditor._rederive(
                ManifestEntry(id="N004", label="P0", value=0, source_file="x", derivation="count_by_severity")
            )
            == 2
        )
        assert (
            auditor._rederive(
                ManifestEntry(id="N005", label="P1", value=0, source_file="x", derivation="count_by_severity")
            )
            == 1
        )
        assert (
            auditor._rederive(
                ManifestEntry(id="N006", label="P2", value=0, source_file="x", derivation="count_by_severity")
            )
            == 3
        )
        assert (
            auditor._rederive(
                ManifestEntry(id="N007", label="P3", value=0, source_file="x", derivation="count_by_severity")
            )
            == 1
        )

    def test_rederive_n008_counts_clean_results(self, tmp_path: Path) -> None:
        """N008 rederivation counts findings with category=domain_reviewed_no_issues."""
        merged_dir = tmp_path / "findings" / "merged"
        merged_dir.mkdir(parents=True)
        (merged_dir / "customer_a.json").write_text(
            json.dumps(
                {
                    "findings": [
                        {"severity": "P3", "category": "domain_reviewed_no_issues"},
                        {"severity": "P2", "category": "contractual_risk"},
                        {"severity": "P3", "category": "domain_reviewed_no_issues"},
                    ]
                }
            )
        )
        auditor = NumericalAuditor(run_dir=tmp_path, inventory_dir=tmp_path)
        entry = ManifestEntry(
            id="N008", label="Clean Results", value=0, source_file="findings/*.json", derivation="count_clean"
        )
        assert auditor._rederive(entry) == 2

    def test_rederive_n009_counts_total_gaps(self, tmp_path: Path) -> None:
        """N009 rederivation counts total gaps from gaps dir."""
        gaps_dir = tmp_path / "findings" / "merged" / "gaps"
        gaps_dir.mkdir(parents=True)
        (gaps_dir / "customer_a.json").write_text(json.dumps([{"gap": "missing clause"}, {"gap": "no termination"}]))
        (gaps_dir / "customer_b.json").write_text(json.dumps({"gaps": [{"gap": "no renewal"}]}))
        auditor = NumericalAuditor(run_dir=tmp_path, inventory_dir=tmp_path)
        entry = ManifestEntry(id="N009", label="Total Gaps", value=0, source_file="gaps/*.json", derivation="count")
        assert auditor._rederive(entry) == 3

    def test_rederive_n010_counts_reference_files(self, tmp_path: Path) -> None:
        """N010 rederivation counts reference files."""
        (tmp_path / "reference_files.json").write_text(
            json.dumps(["file_1.pdf", "file_2.pdf", "file_3.pdf", "file_4.pdf", "file_5.pdf"])
        )
        auditor = NumericalAuditor(run_dir=tmp_path, inventory_dir=tmp_path)
        entry = ManifestEntry(
            id="N010", label="Reference Files", value=0, source_file="reference_files.json", derivation="count"
        )
        assert auditor._rederive(entry) == 5

    def test_rederive_n008_no_merged_dir_returns_zero(self, tmp_path: Path) -> None:
        """N008 returns 0 when merged dir does not exist."""
        auditor = NumericalAuditor(run_dir=tmp_path, inventory_dir=tmp_path)
        entry = ManifestEntry(
            id="N008", label="Clean Results", value=0, source_file="findings/*.json", derivation="count_clean"
        )
        assert auditor._rederive(entry) == 0

    def test_rederive_n009_no_gaps_dir_returns_zero(self, tmp_path: Path) -> None:
        """N009 returns 0 when gaps directory does not exist."""
        auditor = NumericalAuditor(run_dir=tmp_path, inventory_dir=tmp_path)
        entry = ManifestEntry(id="N009", label="Total Gaps", value=0, source_file="gaps/*.json", derivation="count")
        assert auditor._rederive(entry) == 0

    def test_rederive_n003_no_merged_dir_returns_zero(self, tmp_path: Path) -> None:
        """N003 returns 0 when merged dir does not exist."""
        auditor = NumericalAuditor(run_dir=tmp_path, inventory_dir=tmp_path)
        entry = ManifestEntry(
            id="N003", label="Total Findings", value=0, source_file="findings/*.json", derivation="count"
        )
        assert auditor._rederive(entry) == 0


# ===================================================================== #
# QAAuditor Tests
# ===================================================================== #


class TestQAAuditor:
    """Tests for QAAuditor."""

    def test_basic_audit_with_sample_data(self, tmp_path: Path) -> None:
        """QAAuditor runs all 17 checks and produces an AuditReport."""
        run_dir = tmp_path / "run"
        inventory_dir = tmp_path / "inventory"
        _populate_run_dir(run_dir, inventory_dir)

        # Add report_schema.json for consistency check
        (run_dir / "report_schema.json").write_text("{}")

        # Add numerical_manifest.json
        manifest = _make_manifest()
        (run_dir / "numerical_manifest.json").write_text(manifest.model_dump_json())

        auditor = QAAuditor(
            run_dir=run_dir,
            inventory_dir=inventory_dir,
            customer_safe_names=CUSTOMERS,
        )
        report = auditor.run_full_audit(run_id="test_run")

        assert isinstance(report, AuditReport)
        assert report.run_id == "test_run"
        # Should have 18 check entries (17 original + p0_p1_citation_quality)
        assert len(report.checks) == 18

    def test_customer_coverage_fails_when_missing(self, tmp_path: Path) -> None:
        """Customer coverage check fails when a customer file is missing."""
        run_dir = tmp_path / "run"
        inventory_dir = tmp_path / "inventory"
        _populate_run_dir(run_dir, inventory_dir)

        # Remove one customer file from one agent
        (run_dir / "findings" / "legal" / "initech.json").unlink()

        auditor = QAAuditor(
            run_dir=run_dir,
            inventory_dir=inventory_dir,
            customer_safe_names=CUSTOMERS,
        )
        report = auditor.run_full_audit()

        cov = report.checks.get("customer_coverage")
        assert cov is not None
        assert cov.passed is False

    def test_write_audit_json(self, tmp_path: Path) -> None:
        """write_audit_json creates a valid JSON file."""
        run_dir = tmp_path / "run"
        inventory_dir = tmp_path / "inventory"
        _populate_run_dir(run_dir, inventory_dir)
        (run_dir / "report_schema.json").write_text("{}")
        manifest = _make_manifest()
        (run_dir / "numerical_manifest.json").write_text(manifest.model_dump_json())

        auditor = QAAuditor(
            run_dir=run_dir,
            inventory_dir=inventory_dir,
            customer_safe_names=CUSTOMERS,
        )
        report = auditor.run_full_audit(run_id="test_run")

        output_path = tmp_path / "audit.json"
        auditor.write_audit_json(report, output_path)
        assert output_path.exists()

        data = json.loads(output_path.read_text())
        assert "audit_passed" in data
        assert "checks" in data
        assert "summary" in data

    def test_audit_summary_counts(self, tmp_path: Path) -> None:
        """AuditReport summary reflects merged findings."""
        run_dir = tmp_path / "run"
        inventory_dir = tmp_path / "inventory"
        _populate_run_dir(run_dir, inventory_dir)
        (run_dir / "report_schema.json").write_text("{}")
        manifest = _make_manifest()
        (run_dir / "numerical_manifest.json").write_text(manifest.model_dump_json())

        auditor = QAAuditor(
            run_dir=run_dir,
            inventory_dir=inventory_dir,
            customer_safe_names=CUSTOMERS,
        )
        report = auditor.run_full_audit()

        assert report.summary.total_customers == len(CUSTOMERS)
        assert report.summary.total_findings >= 0


class TestQAAuditorDeferredChecks:
    """Tests for QA audit deferred checks when artifacts don't exist."""

    @staticmethod
    def _populate_minimal(run_dir: Path, inventory_dir: Path) -> None:
        """Populate with ONLY the artifacts the pipeline actually produces.

        No coverage manifests, no audit logs, no report_schema.json.
        """
        customers = CUSTOMERS
        inventory_dir.mkdir(parents=True, exist_ok=True)
        files_list = [f"{c}/contract.pdf" for c in customers]
        (inventory_dir / "files.txt").write_text("\n".join(files_list) + "\n")
        (inventory_dir / "customers.csv").write_text(
            "group,name,safe_name,path,file_count\n"
            + "\n".join(f"group1,{c},{c},/data/{c},1" for c in customers)
            + "\n"
        )
        (inventory_dir / "counts.json").write_text(
            json.dumps({"total_customers": len(customers), "total_files": len(files_list)})
        )
        (inventory_dir / "extraction_quality.json").write_text("[]")

        # Per-agent outputs (no coverage manifests)
        for agent in AGENTS:
            agent_dir = run_dir / "findings" / agent
            agent_dir.mkdir(parents=True, exist_ok=True)
            for customer in customers:
                (agent_dir / f"{customer}.json").write_text(json.dumps(_make_customer_json(customer, agent)))

        # Merged output with ALL agents represented
        merged_dir = run_dir / "findings" / "merged"
        merged_dir.mkdir(parents=True, exist_ok=True)
        for customer in customers:
            merged = _make_merged_json(customer)
            # Add findings from all 4 agents so domain_coverage passes
            merged["findings"] = []
            for agent in AGENTS:
                merged["findings"].append(
                    {
                        "id": f"forensic-dd_{agent}_{customer}_0001",
                        "severity": "P2",
                        "category": "test_category",
                        "title": "Test finding",
                        "description": "Test description",
                        "citations": [
                            {
                                "source_type": "file",
                                "source_path": f"{customer}/contract.pdf",
                                "exact_quote": "test quote",
                            }
                        ],
                        "confidence": "high",
                        "agent": agent,
                    }
                )
            (merged_dir / f"{customer}.json").write_text(json.dumps(merged))

        # Numerical manifest
        manifest = _make_manifest()
        (run_dir / "numerical_manifest.json").write_text(manifest.model_dump_json())

        # Report dir (no Excel)
        (run_dir / "report").mkdir(parents=True, exist_ok=True)

    def test_deferred_agent_manifest(self, tmp_path: Path) -> None:
        """agent_manifest_reconciliation passes when no manifests exist but files do."""
        run_dir = tmp_path / "run"
        inventory_dir = tmp_path / "inventory"
        self._populate_minimal(run_dir, inventory_dir)

        auditor = QAAuditor(run_dir=run_dir, inventory_dir=inventory_dir, customer_safe_names=CUSTOMERS)
        name, check = auditor.check_agent_manifest_reconciliation()
        assert check.passed is True
        # No manifests, but file-counting fallback finds all customer files.
        for agent_detail in check.details.values():
            if isinstance(agent_detail, dict):
                assert agent_detail.get("match") is True

    def test_deferred_file_coverage(self, tmp_path: Path) -> None:
        """file_coverage passes when no coverage manifests exist."""
        run_dir = tmp_path / "run"
        inventory_dir = tmp_path / "inventory"
        self._populate_minimal(run_dir, inventory_dir)

        auditor = QAAuditor(run_dir=run_dir, inventory_dir=inventory_dir, customer_safe_names=CUSTOMERS)
        name, check = auditor.check_file_coverage()
        assert check.passed is True
        assert "deferred" in check.details.get("note", "").lower()

    def test_deferred_audit_logs(self, tmp_path: Path) -> None:
        """audit_logs passes when no logs exist."""
        run_dir = tmp_path / "run"
        inventory_dir = tmp_path / "inventory"
        self._populate_minimal(run_dir, inventory_dir)

        auditor = QAAuditor(run_dir=run_dir, inventory_dir=inventory_dir, customer_safe_names=CUSTOMERS)
        name, check = auditor.check_audit_logs()
        assert check.passed is True
        assert "deferred" in check.details.get("note", "").lower()

    def test_deferred_report_consistency(self, tmp_path: Path) -> None:
        """report_consistency passes when report_schema.json doesn't exist."""
        run_dir = tmp_path / "run"
        inventory_dir = tmp_path / "inventory"
        self._populate_minimal(run_dir, inventory_dir)

        auditor = QAAuditor(run_dir=run_dir, inventory_dir=inventory_dir, customer_safe_names=CUSTOMERS)
        name, check = auditor.check_report_consistency()
        assert check.passed is True
        assert "deferred" in check.details.get("note", "").lower()

    def test_citation_integrity_basename_match(self, tmp_path: Path) -> None:
        """Citation with different prefix but same basename passes."""
        run_dir = tmp_path / "run"
        inventory_dir = tmp_path / "inventory"
        self._populate_minimal(run_dir, inventory_dir)

        # Overwrite merged with citation using full path (different prefix)
        merged_dir = run_dir / "findings" / "merged"
        merged = {
            "customer": "acme_corp",
            "findings": [
                {
                    "id": "test_1",
                    "severity": "P2",
                    "category": "test",
                    "title": "Test",
                    "description": "Test",
                    "citations": [
                        {
                            "source_type": "file",
                            "source_path": "/full/path/to/acme_corp/contract.pdf",
                            "exact_quote": "test",
                        }
                    ],
                    "agent": "legal",
                }
            ],
            "gaps": [],
            "cross_references": [],
            "governance_graph": {"edges": []},
        }
        (merged_dir / "acme_corp.json").write_text(json.dumps(merged))

        auditor = QAAuditor(run_dir=run_dir, inventory_dir=inventory_dir, customer_safe_names=CUSTOMERS)
        name, check = auditor.check_citation_integrity()
        assert check.passed is True

    def test_citation_integrity_skips_synthetic(self, tmp_path: Path) -> None:
        """Synthetic citations are skipped by citation_integrity check."""
        run_dir = tmp_path / "run"
        inventory_dir = tmp_path / "inventory"
        self._populate_minimal(run_dir, inventory_dir)

        merged_dir = run_dir / "findings" / "merged"
        merged = {
            "customer": "acme_corp",
            "findings": [
                {
                    "id": "test_1",
                    "severity": "P3",
                    "category": "test",
                    "title": "Test",
                    "description": "Test",
                    "citations": [
                        {
                            "source_type": "file",
                            "source_path": "[synthetic:no_citation_provided]",
                            "exact_quote": "",
                        }
                    ],
                    "agent": "legal",
                }
            ],
            "gaps": [],
            "cross_references": [],
            "governance_graph": {"edges": []},
        }
        (merged_dir / "acme_corp.json").write_text(json.dumps(merged))

        auditor = QAAuditor(run_dir=run_dir, inventory_dir=inventory_dir, customer_safe_names=CUSTOMERS)
        name, check = auditor.check_citation_integrity()
        assert check.passed is True

    def test_citation_integrity_skips_directory_and_description_refs(self, tmp_path: Path) -> None:
        """Directory paths and description-style references are skipped."""
        run_dir = tmp_path / "run"
        inventory_dir = tmp_path / "inventory"
        self._populate_minimal(run_dir, inventory_dir)

        merged_dir = run_dir / "findings" / "merged"
        merged = {
            "customer": "acme_corp",
            "findings": [
                {
                    "id": "test_dir",
                    "severity": "P2",
                    "category": "test",
                    "title": "Dir ref",
                    "description": "Test",
                    "citations": [
                        {
                            "source_type": "file",
                            "source_path": "2. Legal Due Diligence/2.7. Vertu Management Agreement",
                            "exact_quote": "test",
                        },
                        {
                            "source_type": "file",
                            "source_path": "4.1 Tax Returns folder",
                            "exact_quote": "test",
                        },
                        {
                            "source_type": "file",
                            "source_path": "5.4.3.1.1 Canadian Benefits Overview",
                            "exact_quote": "test",
                        },
                    ],
                    "agent": "legal",
                }
            ],
            "gaps": [],
            "cross_references": [],
            "governance_graph": {"edges": []},
        }
        (merged_dir / "acme_corp.json").write_text(json.dumps(merged))

        auditor = QAAuditor(run_dir=run_dir, inventory_dir=inventory_dir, customer_safe_names=CUSTOMERS)
        name, check = auditor.check_citation_integrity()
        assert check.passed is True

    def test_merge_dedup_tolerance(self, tmp_path: Path) -> None:
        """merge_dedup passes when merged count is within 90% of expected."""
        run_dir = tmp_path / "run"
        inventory_dir = tmp_path / "inventory"
        self._populate_minimal(run_dir, inventory_dir)

        # Remove one merged file (simulating a skipped customer)
        (run_dir / "findings" / "merged" / "initech.json").unlink()

        auditor = QAAuditor(run_dir=run_dir, inventory_dir=inventory_dir, customer_safe_names=CUSTOMERS)
        name, check = auditor.check_merge_dedup()
        # 2/3 = 0.667 < 0.90 threshold, so this should fail
        assert check.passed is False

    def test_merge_dedup_exact_match_still_passes(self, tmp_path: Path) -> None:
        """merge_dedup passes when all customers are merged."""
        run_dir = tmp_path / "run"
        inventory_dir = tmp_path / "inventory"
        self._populate_minimal(run_dir, inventory_dir)

        auditor = QAAuditor(run_dir=run_dir, inventory_dir=inventory_dir, customer_safe_names=CUSTOMERS)
        name, check = auditor.check_merge_dedup()
        assert check.passed is True

    def test_domain_coverage_threshold_fails_when_very_low(self, tmp_path: Path) -> None:
        """domain_coverage fails when coverage is below 20% threshold."""
        run_dir = tmp_path / "run"
        inventory_dir = tmp_path / "inventory"
        self._populate_minimal(run_dir, inventory_dir)

        # Remove ALL merged files -- 0/3 = 0% < 20%, should fail
        for jf in (run_dir / "findings" / "merged").glob("*.json"):
            jf.unlink()

        auditor = QAAuditor(run_dir=run_dir, inventory_dir=inventory_dir, customer_safe_names=CUSTOMERS)
        name, check = auditor.check_domain_coverage()
        assert check.passed is False

    def test_full_audit_passes_without_optional_artifacts(self, tmp_path: Path) -> None:
        """Full QA audit passes with only the artifacts the pipeline produces."""
        run_dir = tmp_path / "run"
        inventory_dir = tmp_path / "inventory"
        self._populate_minimal(run_dir, inventory_dir)

        auditor = QAAuditor(run_dir=run_dir, inventory_dir=inventory_dir, customer_safe_names=CUSTOMERS)
        report = auditor.run_full_audit(run_id="test_minimal")

        failed = [name for name, check in report.checks.items() if not check.passed]
        assert report.audit_passed, f"QA audit should pass but these checks failed: {failed}"

    def test_manifest_reconciliation_file_counting_fallback(self, tmp_path: Path) -> None:
        """Manifest with wrong counts passes when actual files exist on disk."""
        run_dir = tmp_path / "run"
        inventory_dir = tmp_path / "inventory"
        self._populate_minimal(run_dir, inventory_dir)

        # Write a manifest with wrong counts (simulating batch overwrite)
        agent_dir = run_dir / "findings" / "legal"
        agent_dir.mkdir(parents=True, exist_ok=True)
        manifest = {
            "analysis_units_assigned": 1,
            "analysis_units_completed": 1,
        }
        (agent_dir / "coverage_manifest.json").write_text(json.dumps(manifest))

        auditor = QAAuditor(run_dir=run_dir, inventory_dir=inventory_dir, customer_safe_names=CUSTOMERS)
        name, check = auditor.check_agent_manifest_reconciliation()
        # File-counting fallback finds all customer files → passes
        assert check.passed is True

    def test_citation_integrity_tolerance(self, tmp_path: Path) -> None:
        """citation_integrity passes when failure ratio is within threshold."""
        run_dir = tmp_path / "run"
        inventory_dir = tmp_path / "inventory"
        self._populate_minimal(run_dir, inventory_dir)

        # Create 20 findings: 1 with bad citation, 19 with good citations.
        # 1/20 = 5% < 10% threshold → should pass.
        merged_dir = run_dir / "findings" / "merged"
        findings = []
        for i in range(19):
            findings.append(
                {
                    "id": f"test_{i}",
                    "severity": "P2",
                    "category": "test",
                    "title": f"Good finding {i}",
                    "description": "Test",
                    "citations": [
                        {
                            "source_type": "file",
                            "source_path": "acme_corp/contract.pdf",
                            "exact_quote": "test",
                        }
                    ],
                    "agent": "legal",
                }
            )
        # One finding with unmatched source_path
        findings.append(
            {
                "id": "test_bad",
                "severity": "P2",
                "category": "test",
                "title": "Bad citation",
                "description": "Test",
                "citations": [
                    {
                        "source_type": "file",
                        "source_path": "nonexistent/file.pdf",
                        "exact_quote": "test",
                    }
                ],
                "agent": "legal",
            }
        )
        merged = {
            "customer": "acme_corp",
            "findings": findings,
            "gaps": [],
            "cross_references": [],
            "governance_graph": {"edges": []},
        }
        (merged_dir / "acme_corp.json").write_text(json.dumps(merged))

        auditor = QAAuditor(run_dir=run_dir, inventory_dir=inventory_dir, customer_safe_names=CUSTOMERS)
        name, check = auditor.check_citation_integrity()
        assert check.passed is True
        assert check.details.get("failure_ratio", 1.0) <= 0.10

    def test_report_sheets_always_defers(self, tmp_path: Path) -> None:
        """report_sheets always defers (step 28 is pre-generation)."""
        run_dir = tmp_path / "run"
        inventory_dir = tmp_path / "inventory"
        self._populate_minimal(run_dir, inventory_dir)

        # Even if a stale Excel exists, report_sheets should defer
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "WrongSheet"  # Not one of the required sheets
        report_dir = run_dir / "report"
        report_dir.mkdir(parents=True, exist_ok=True)
        wb.save(report_dir / "stale_report.xlsx")

        auditor = QAAuditor(run_dir=run_dir, inventory_dir=inventory_dir, customer_safe_names=CUSTOMERS)
        name, check = auditor.check_report_sheets()
        assert check.passed is True
        assert "deferred" in check.details.get("note", "").lower()

    def test_p0p1_citation_quality_tolerance(self, tmp_path: Path) -> None:
        """p0_p1_citation_quality passes when violation ratio is within threshold."""
        run_dir = tmp_path / "run"
        inventory_dir = tmp_path / "inventory"
        self._populate_minimal(run_dir, inventory_dir)

        # Create 20 P0 findings: 1 with bad citation, 19 with good citations.
        # 1/20 = 5% < 10% threshold → should pass.
        merged_dir = run_dir / "findings" / "merged"
        findings = []
        for i in range(19):
            findings.append(
                {
                    "id": f"test_p0_{i}",
                    "severity": "P0",
                    "category": "test",
                    "title": f"Good P0 {i}",
                    "description": "Test",
                    "citations": [
                        {
                            "source_type": "file",
                            "source_path": "acme_corp/contract.pdf",
                            "exact_quote": "real quote text",
                        }
                    ],
                    "agent": "legal",
                }
            )
        # One P0 finding with synthetic citation
        findings.append(
            {
                "id": "test_p0_bad",
                "severity": "P0",
                "category": "test",
                "title": "Bad P0",
                "description": "Test",
                "citations": [
                    {
                        "source_type": "file",
                        "source_path": "[synthetic:no_citation_provided]",
                        "exact_quote": "some quote",
                    }
                ],
                "agent": "legal",
            }
        )
        merged = {
            "customer": "acme_corp",
            "findings": findings,
            "gaps": [],
            "cross_references": [],
            "governance_graph": {"edges": []},
        }
        (merged_dir / "acme_corp.json").write_text(json.dumps(merged))

        auditor = QAAuditor(run_dir=run_dir, inventory_dir=inventory_dir, customer_safe_names=CUSTOMERS)
        name, check = auditor.check_p0_p1_citation_quality()
        assert check.passed is True
        assert check.details.get("violation_ratio", 1.0) <= 0.10


# ===================================================================== #
# DefinitionOfDoneChecker Tests
# ===================================================================== #


class TestDefinitionOfDoneChecker:
    """Tests for DefinitionOfDoneChecker."""

    def test_check_count_without_conditionals(self, tmp_path: Path) -> None:
        """Without conditional features, expect 30 - 8 conditional = 22 checks."""
        run_dir = tmp_path / "run"
        inventory_dir = tmp_path / "inventory"
        _populate_run_dir(run_dir, inventory_dir)

        checker = DefinitionOfDoneChecker(
            run_dir=run_dir,
            inventory_dir=inventory_dir,
            customer_safe_names=CUSTOMERS,
            deal_config={},
        )
        results = checker.check_all()

        # Core (12) + Reporting (5) + Contract Dates (1) + Extraction (1)
        # + Report Consistency (3) = 22 (no Judge, no Incremental)
        assert len(results) == 22

    def test_check_count_with_judge(self, tmp_path: Path) -> None:
        """With judge.enabled, 4 additional checks appear."""
        run_dir = tmp_path / "run"
        inventory_dir = tmp_path / "inventory"
        _populate_run_dir(run_dir, inventory_dir)

        checker = DefinitionOfDoneChecker(
            run_dir=run_dir,
            inventory_dir=inventory_dir,
            customer_safe_names=CUSTOMERS,
            deal_config={"judge": {"enabled": True, "threshold": 70}},
        )
        results = checker.check_all()
        assert len(results) == 26  # 22 + 4 judge

    def test_check_count_with_incremental(self, tmp_path: Path) -> None:
        """With incremental mode, 4 additional checks appear."""
        run_dir = tmp_path / "run"
        inventory_dir = tmp_path / "inventory"
        _populate_run_dir(run_dir, inventory_dir)

        checker = DefinitionOfDoneChecker(
            run_dir=run_dir,
            inventory_dir=inventory_dir,
            customer_safe_names=CUSTOMERS,
            deal_config={"execution": {"mode": "incremental"}},
        )
        results = checker.check_all()
        assert len(results) == 26  # 22 + 4 incremental

    def test_check_count_with_all_conditionals(self, tmp_path: Path) -> None:
        """With all conditionals, expect full 30 checks."""
        run_dir = tmp_path / "run"
        inventory_dir = tmp_path / "inventory"
        _populate_run_dir(run_dir, inventory_dir)

        checker = DefinitionOfDoneChecker(
            run_dir=run_dir,
            inventory_dir=inventory_dir,
            customer_safe_names=CUSTOMERS,
            deal_config={
                "judge": {"enabled": True, "threshold": 70},
                "execution": {"mode": "incremental"},
            },
        )
        results = checker.check_all()
        assert len(results) == 30

    def test_all_checks_return_audit_check(self, tmp_path: Path) -> None:
        """Every check returns an AuditCheck instance."""
        run_dir = tmp_path / "run"
        inventory_dir = tmp_path / "inventory"
        _populate_run_dir(run_dir, inventory_dir)

        checker = DefinitionOfDoneChecker(
            run_dir=run_dir,
            inventory_dir=inventory_dir,
            customer_safe_names=CUSTOMERS,
        )
        results = checker.check_all()
        for check in results:
            assert isinstance(check, AuditCheck)
            assert isinstance(check.dod_checks, list)
            assert len(check.dod_checks) >= 1

    def test_customer_outputs_pass_with_full_data(self, tmp_path: Path) -> None:
        """DoD check 1 passes when all customer outputs exist."""
        run_dir = tmp_path / "run"
        inventory_dir = tmp_path / "inventory"
        _populate_run_dir(run_dir, inventory_dir)

        checker = DefinitionOfDoneChecker(
            run_dir=run_dir,
            inventory_dir=inventory_dir,
            customer_safe_names=CUSTOMERS,
        )
        check = checker.check_1_customer_outputs_complete()
        assert check.passed is True
        assert 1 in check.dod_checks

    def test_customer_outputs_fail_missing_file(self, tmp_path: Path) -> None:
        """DoD check 1 fails when a customer file is missing."""
        run_dir = tmp_path / "run"
        inventory_dir = tmp_path / "inventory"
        _populate_run_dir(run_dir, inventory_dir)
        (run_dir / "findings" / "legal" / "acme_corp.json").unlink()

        checker = DefinitionOfDoneChecker(
            run_dir=run_dir,
            inventory_dir=inventory_dir,
            customer_safe_names=CUSTOMERS,
        )
        check = checker.check_1_customer_outputs_complete()
        assert check.passed is False

    def test_dod_groups_cover_all_30(self) -> None:
        """Verify that check group ranges cover exactly 1..30."""
        all_nums = set()
        all_nums.update(DefinitionOfDoneChecker.CORE_ANALYSIS)
        all_nums.update(DefinitionOfDoneChecker.REPORTING)
        all_nums.update(DefinitionOfDoneChecker.CONTRACT_DATES)
        all_nums.update(DefinitionOfDoneChecker.EXTRACTION)
        all_nums.update(DefinitionOfDoneChecker.JUDGE)
        all_nums.update(DefinitionOfDoneChecker.INCREMENTAL)
        all_nums.update(DefinitionOfDoneChecker.REPORT_CONSISTENCY)
        assert all_nums == set(range(1, 31))


# ===================================================================== #
# SchemaValidator Tests
# ===================================================================== #


class TestSchemaValidator:
    """Tests for SchemaValidator."""

    @pytest.fixture()
    def _excel_with_matching_schema(self, tmp_path: Path):
        """Create a matching Excel + ReportSchema pair."""
        from dd_agents.models.reporting import (
            ColumnDef,
            ReportSchema,
            SheetDef,
            SortOrder,
        )

        schema = ReportSchema(
            schema_version="1.0",
            sheets=[
                SheetDef(
                    name="Summary",
                    required=True,
                    columns=[
                        ColumnDef(name="Customer", key="customer", type="string"),
                        ColumnDef(name="Findings", key="findings", type="integer"),
                    ],
                    sort_order=[SortOrder(column="Customer", direction="asc")],
                ),
                SheetDef(
                    name="Wolf_Pack",
                    required=True,
                    columns=[
                        ColumnDef(name="ID", key="id", type="string"),
                        ColumnDef(name="Severity", key="severity", type="string"),
                    ],
                ),
            ],
        )

        # Create matching Excel
        import openpyxl

        wb = openpyxl.Workbook()

        ws_summary = wb.active
        ws_summary.title = "Summary"
        ws_summary.append(["Customer", "Findings"])
        ws_summary.append(["Acme Corp", 10])
        ws_summary.append(["Globex", 5])
        ws_summary.append(["Initech", 8])

        ws_wolf = wb.create_sheet("Wolf_Pack")
        ws_wolf.append(["ID", "Severity"])
        ws_wolf.append(["F001", "P0"])

        excel_path = tmp_path / "report.xlsx"
        wb.save(excel_path)

        return schema, excel_path

    def test_matching_excel_passes(self, _excel_with_matching_schema: tuple, tmp_path: Path) -> None:
        """A workbook matching the schema passes all checks."""
        schema, excel_path = _excel_with_matching_schema

        validator = SchemaValidator(schema)
        checks = validator.validate_report(excel_path)

        assert len(checks) >= 3  # sheets, columns, sort, formatting
        for check in checks:
            assert isinstance(check, AuditCheck)
            assert check.passed is True, f"Check failed: {check.details}"

    def test_missing_sheet_fails(self, tmp_path: Path) -> None:
        """When a required sheet is missing, the sheets_exist check fails."""
        from dd_agents.models.reporting import (
            ColumnDef,
            ReportSchema,
            SheetDef,
        )

        schema = ReportSchema(
            schema_version="1.0",
            sheets=[
                SheetDef(
                    name="Summary",
                    required=True,
                    columns=[
                        ColumnDef(name="Customer", key="customer", type="string"),
                    ],
                ),
                SheetDef(
                    name="Details",
                    required=True,
                    columns=[
                        ColumnDef(name="ID", key="id", type="string"),
                    ],
                ),
            ],
        )

        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Summary"
        ws.append(["Customer"])
        # "Details" sheet is missing
        excel_path = tmp_path / "report.xlsx"
        wb.save(excel_path)

        validator = SchemaValidator(schema)
        checks = validator.validate_report(excel_path)

        sheets_check = next(c for c in checks if c.details.get("check") == "sheets_exist")
        assert sheets_check.passed is False
        assert "Details" in sheets_check.details["missing_sheets"]

    def test_column_mismatch_fails(self, tmp_path: Path) -> None:
        """When column names do not match, the columns_match check fails."""
        from dd_agents.models.reporting import (
            ColumnDef,
            ReportSchema,
            SheetDef,
        )

        schema = ReportSchema(
            schema_version="1.0",
            sheets=[
                SheetDef(
                    name="Summary",
                    required=True,
                    columns=[
                        ColumnDef(name="Customer", key="customer", type="string"),
                        ColumnDef(name="Findings Count", key="findings", type="integer"),
                    ],
                ),
            ],
        )

        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Summary"
        ws.append(["Customer", "Total Issues"])  # wrong column name
        excel_path = tmp_path / "report.xlsx"
        wb.save(excel_path)

        validator = SchemaValidator(schema)
        checks = validator.validate_report(excel_path)

        col_check = next(c for c in checks if c.details.get("check") == "columns_match")
        assert col_check.passed is False

    def test_nonexistent_excel_fails(self, tmp_path: Path) -> None:
        """When the Excel file does not exist, validation fails immediately."""
        from dd_agents.models.reporting import ColumnDef, ReportSchema, SheetDef

        schema = ReportSchema(
            schema_version="1.0",
            sheets=[
                SheetDef(
                    name="Summary",
                    columns=[ColumnDef(name="Customer", key="customer", type="string")],
                ),
            ],
        )
        validator = SchemaValidator(schema)
        checks = validator.validate_report(tmp_path / "nonexistent.xlsx")

        assert len(checks) == 1
        assert checks[0].passed is False

    def test_sort_order_validated(self, tmp_path: Path) -> None:
        """Sort order check detects incorrectly sorted data."""
        from dd_agents.models.reporting import (
            ColumnDef,
            ReportSchema,
            SheetDef,
            SortOrder,
        )

        schema = ReportSchema(
            schema_version="1.0",
            sheets=[
                SheetDef(
                    name="Summary",
                    required=True,
                    columns=[
                        ColumnDef(name="Customer", key="customer", type="string"),
                    ],
                    sort_order=[SortOrder(column="Customer", direction="asc")],
                ),
            ],
        )

        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Summary"
        ws.append(["Customer"])
        ws.append(["Zebra"])  # wrong order
        ws.append(["Acme"])
        ws.append(["Middle"])
        excel_path = tmp_path / "report.xlsx"
        wb.save(excel_path)

        validator = SchemaValidator(schema)
        checks = validator.validate_report(excel_path)

        sort_check = next(c for c in checks if c.details.get("check") == "sort_orders")
        assert sort_check.passed is False

    def test_sheet_activation_never(self, tmp_path: Path) -> None:
        """A sheet with activation_condition='never' should not be required."""
        from dd_agents.models.reporting import (
            ColumnDef,
            ReportSchema,
            SheetDef,
        )

        schema = ReportSchema(
            schema_version="1.0",
            sheets=[
                SheetDef(
                    name="Summary",
                    required=True,
                    columns=[
                        ColumnDef(name="Customer", key="customer", type="string"),
                    ],
                ),
                SheetDef(
                    name="Conditional_Sheet",
                    required=True,
                    activation_condition="never",
                    columns=[
                        ColumnDef(name="Data", key="data", type="string"),
                    ],
                ),
            ],
        )

        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Summary"
        ws.append(["Customer"])
        # Conditional_Sheet is NOT present but should not be required
        excel_path = tmp_path / "report.xlsx"
        wb.save(excel_path)

        validator = SchemaValidator(schema)
        checks = validator.validate_report(excel_path)

        sheets_check = next(c for c in checks if c.details.get("check") == "sheets_exist")
        assert sheets_check.passed is True

    def test_column_activation_never(self, tmp_path: Path) -> None:
        """A column with activation_condition='never' should not be checked."""
        from dd_agents.models.reporting import (
            ColumnDef,
            ReportSchema,
            SheetDef,
        )

        schema = ReportSchema(
            schema_version="1.0",
            sheets=[
                SheetDef(
                    name="Summary",
                    required=True,
                    columns=[
                        ColumnDef(name="Customer", key="customer", type="string"),
                        ColumnDef(
                            name="Optional_Col",
                            key="optional",
                            type="string",
                            activation_condition="never",
                        ),
                    ],
                ),
            ],
        )

        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Summary"
        ws.append(["Customer"])  # Only "Customer", no "Optional_Col"
        ws.append(["Acme Corp"])
        excel_path = tmp_path / "report.xlsx"
        wb.save(excel_path)

        validator = SchemaValidator(schema)
        checks = validator.validate_report(excel_path)

        col_check = next(c for c in checks if c.details.get("check") == "columns_match")
        assert col_check.passed is True


# ===================================================================== #
# Issue #50 -- DoD hardcoded passes removed
# ===================================================================== #


class TestDoDHardcodedPassesRemoved:
    """Tests verifying that DoD checks no longer use hardcoded passed=True."""

    def test_check_4_fails_without_audit_json(self, tmp_path: Path) -> None:
        """check_4 should fail when audit.json doesn't exist."""
        run_dir = tmp_path / "run"
        inventory_dir = tmp_path / "inventory"
        _populate_run_dir(run_dir, inventory_dir)

        checker = DefinitionOfDoneChecker(
            run_dir=run_dir,
            inventory_dir=inventory_dir,
            customer_safe_names=CUSTOMERS,
        )
        check = checker.check_4_governance_resolved()
        assert check.passed is False
        assert check.details.get("governance_check") == "qa_audit_not_run_yet"

    def test_check_5_fails_without_audit_json(self, tmp_path: Path) -> None:
        """check_5 should fail when audit.json doesn't exist."""
        run_dir = tmp_path / "run"
        inventory_dir = tmp_path / "inventory"
        _populate_run_dir(run_dir, inventory_dir)

        checker = DefinitionOfDoneChecker(
            run_dir=run_dir,
            inventory_dir=inventory_dir,
            customer_safe_names=CUSTOMERS,
        )
        check = checker.check_5_citations_valid()
        assert check.passed is False

    def test_check_6_fails_without_merged_dir(self, tmp_path: Path) -> None:
        """check_6 should fail when merged directory has no JSON files."""
        run_dir = tmp_path / "run"
        run_dir.mkdir(parents=True)
        inventory_dir = tmp_path / "inventory"
        inventory_dir.mkdir(parents=True)

        checker = DefinitionOfDoneChecker(
            run_dir=run_dir,
            inventory_dir=inventory_dir,
            customer_safe_names=CUSTOMERS,
        )
        check = checker.check_6_gaps_tracked()
        assert check.passed is False

    def test_check_10_fails_without_reference_files(self, tmp_path: Path) -> None:
        """check_10 should fail when reference_files.json is missing but _reference/ dir exists."""
        run_dir = tmp_path / "run"
        run_dir.mkdir(parents=True)
        inventory_dir = tmp_path / "inventory"
        inventory_dir.mkdir(parents=True)
        # Create a _reference dir so the check doesn't pass vacuously
        (inventory_dir.parent / "_reference").mkdir(parents=True)

        checker = DefinitionOfDoneChecker(
            run_dir=run_dir,
            inventory_dir=inventory_dir,
            customer_safe_names=CUSTOMERS,
        )
        check = checker.check_10_reference_files_processed()
        assert check.passed is False

    def test_check_10_passes_vacuously_without_reference_dir(self, tmp_path: Path) -> None:
        """check_10 passes when neither reference_files.json nor _reference/ dir exist."""
        run_dir = tmp_path / "run"
        run_dir.mkdir(parents=True)
        inventory_dir = tmp_path / "inventory"
        inventory_dir.mkdir(parents=True)

        checker = DefinitionOfDoneChecker(
            run_dir=run_dir,
            inventory_dir=inventory_dir,
            customer_safe_names=CUSTOMERS,
        )
        check = checker.check_10_reference_files_processed()
        assert check.passed is True

    def test_check_16_fails_without_entity_matches(self, tmp_path: Path) -> None:
        """check_16 should fail when entity_matches.json is missing."""
        run_dir = tmp_path / "run"
        run_dir.mkdir(parents=True)
        inventory_dir = tmp_path / "inventory"
        inventory_dir.mkdir(parents=True)

        checker = DefinitionOfDoneChecker(
            run_dir=run_dir,
            inventory_dir=inventory_dir,
            customer_safe_names=CUSTOMERS,
        )
        check = checker.check_16_entity_resolution_log()
        assert check.passed is False

    def test_check_16_fails_with_unmatched_aliases(self, tmp_path: Path) -> None:
        """check_16 should fail when unmatched entities have aliases available."""
        run_dir = tmp_path / "run"
        run_dir.mkdir(parents=True)
        inventory_dir = tmp_path / "inventory"
        inventory_dir.mkdir(parents=True)
        (run_dir / "entity_matches.json").write_text(
            json.dumps(
                {
                    "entries": [
                        {"name": "Customer A", "matched": True},
                        {"name": "Customer B", "matched": False, "aliases": ["Cust B Inc"]},
                    ]
                }
            )
        )
        checker = DefinitionOfDoneChecker(run_dir=run_dir, inventory_dir=inventory_dir, customer_safe_names=CUSTOMERS)
        check = checker.check_16_entity_resolution_log()
        assert check.passed is False
        assert check.details["unmatched_with_aliases"] == 1

    def test_check_16_passes_with_all_matched(self, tmp_path: Path) -> None:
        """check_16 should pass when all entities are matched."""
        run_dir = tmp_path / "run"
        run_dir.mkdir(parents=True)
        inventory_dir = tmp_path / "inventory"
        inventory_dir.mkdir(parents=True)
        (run_dir / "entity_matches.json").write_text(json.dumps({"entries": [{"name": "Customer A", "matched": True}]}))
        checker = DefinitionOfDoneChecker(run_dir=run_dir, inventory_dir=inventory_dir, customer_safe_names=CUSTOMERS)
        check = checker.check_16_entity_resolution_log()
        assert check.passed is True

    def test_check_10_passes_when_all_processed(self, tmp_path: Path) -> None:
        """check_10 should pass when all reference files appear in agent manifests."""
        run_dir = tmp_path / "run"
        run_dir.mkdir(parents=True)
        inventory_dir = tmp_path / "inventory"
        inventory_dir.mkdir(parents=True)
        (inventory_dir / "reference_files.json").write_text(json.dumps(["ref_a.pdf", "ref_b.pdf"]))
        agents_dir = run_dir / "findings" / "agents" / "legal"
        agents_dir.mkdir(parents=True)
        (agents_dir / "reference_files_processed.json").write_text(json.dumps(["ref_a.pdf", "ref_b.pdf"]))
        checker = DefinitionOfDoneChecker(run_dir=run_dir, inventory_dir=inventory_dir, customer_safe_names=CUSTOMERS)
        check = checker.check_10_reference_files_processed()
        assert check.passed is True

    def test_check_10_passes_with_dict_format_ref_files(self, tmp_path: Path) -> None:
        """check_10 should handle dict-format reference_files.json (file_path key)."""
        run_dir = tmp_path / "run"
        run_dir.mkdir(parents=True)
        inventory_dir = tmp_path / "inventory"
        inventory_dir.mkdir(parents=True)
        (inventory_dir / "reference_files.json").write_text(
            json.dumps([{"file_path": "docs/ref_a.pdf", "category": "Financial"}])
        )
        agents_dir = run_dir / "findings" / "agents" / "finance"
        agents_dir.mkdir(parents=True)
        (agents_dir / "reference_files_processed.json").write_text(json.dumps(["ref_a.pdf"]))
        checker = DefinitionOfDoneChecker(run_dir=run_dir, inventory_dir=inventory_dir, customer_safe_names=CUSTOMERS)
        check = checker.check_10_reference_files_processed()
        assert check.passed is True

    def test_check_10_fails_when_file_unprocessed(self, tmp_path: Path) -> None:
        """check_10 should fail when a reference file is not processed by any agent."""
        run_dir = tmp_path / "run"
        run_dir.mkdir(parents=True)
        inventory_dir = tmp_path / "inventory"
        inventory_dir.mkdir(parents=True)
        (inventory_dir / "reference_files.json").write_text(json.dumps(["ref_a.pdf", "ref_b.pdf"]))
        agents_dir = run_dir / "findings" / "agents" / "legal"
        agents_dir.mkdir(parents=True)
        (agents_dir / "reference_files_processed.json").write_text(json.dumps(["ref_a.pdf"]))
        checker = DefinitionOfDoneChecker(run_dir=run_dir, inventory_dir=inventory_dir, customer_safe_names=CUSTOMERS)
        check = checker.check_10_reference_files_processed()
        assert check.passed is False
        assert check.details["unprocessed_count"] == 1

    def test_check_25_passes_when_no_carried_forward(self, tmp_path: Path) -> None:
        """check_25 passes when merged findings have no _carried_forward metadata."""
        run_dir = tmp_path / "run"
        inventory_dir = tmp_path / "inventory"
        _populate_run_dir(run_dir, inventory_dir)

        checker = DefinitionOfDoneChecker(
            run_dir=run_dir,
            inventory_dir=inventory_dir,
            customer_safe_names=CUSTOMERS,
            deal_config={"execution": {"mode": "incremental"}},
        )
        check = checker.check_25_carried_forward_metadata()
        assert check.passed is True

    def test_check_30_fails_with_prior_run_no_diff(self, tmp_path: Path) -> None:
        """check_30 should fail when prior_run_id is set but report_diff.json is missing."""
        run_dir = tmp_path / "run"
        run_dir.mkdir(parents=True)
        inventory_dir = tmp_path / "inventory"
        inventory_dir.mkdir(parents=True)

        checker = DefinitionOfDoneChecker(
            run_dir=run_dir,
            inventory_dir=inventory_dir,
            customer_safe_names=CUSTOMERS,
            deal_config={"execution": {"prior_run_id": "run_previous"}},
        )
        check = checker.check_30_report_diff()
        assert check.passed is False

    def test_check_30_passes_without_prior_run(self, tmp_path: Path) -> None:
        """check_30 should pass when no prior_run_id is configured."""
        run_dir = tmp_path / "run"
        run_dir.mkdir(parents=True)
        inventory_dir = tmp_path / "inventory"
        inventory_dir.mkdir(parents=True)

        checker = DefinitionOfDoneChecker(
            run_dir=run_dir,
            inventory_dir=inventory_dir,
            customer_safe_names=CUSTOMERS,
            deal_config={},
        )
        check = checker.check_30_report_diff()
        assert check.passed is True


# ===================================================================== #
# Issue #48 -- P0/P1 citation quality check
# ===================================================================== #


class TestP0P1CitationQuality:
    """Tests for the P0/P1 citation quality QA check."""

    def test_p0_p1_with_valid_citations_passes(self, tmp_path: Path) -> None:
        """P0/P1 findings with real citations should pass."""
        run_dir = tmp_path / "run"
        inventory_dir = tmp_path / "inventory"
        _populate_run_dir(run_dir, inventory_dir)

        # Overwrite merged files with P1 findings that have real citations
        merged_dir = run_dir / "findings" / "merged"
        for customer in CUSTOMERS:
            data = _make_merged_json(customer)
            data["findings"][0]["severity"] = "P1"
            (merged_dir / f"{customer}.json").write_text(json.dumps(data))

        auditor = QAAuditor(
            run_dir=run_dir,
            inventory_dir=inventory_dir,
            customer_safe_names=CUSTOMERS,
        )
        _name, check = auditor.check_p0_p1_citation_quality()
        assert check.passed is True

    def test_p0_with_synthetic_citation_fails(self, tmp_path: Path) -> None:
        """P0 finding with synthetic citation should fail."""
        run_dir = tmp_path / "run"
        inventory_dir = tmp_path / "inventory"
        _populate_run_dir(run_dir, inventory_dir)

        merged_dir = run_dir / "findings" / "merged"
        data = _make_merged_json("acme_corp")
        data["findings"][0]["severity"] = "P0"
        data["findings"][0]["citations"] = [
            {
                "source_type": "file",
                "source_path": "[synthetic:no_citation_provided]",
                "exact_quote": "some quote",
            }
        ]
        (merged_dir / "acme_corp.json").write_text(json.dumps(data))

        auditor = QAAuditor(
            run_dir=run_dir,
            inventory_dir=inventory_dir,
            customer_safe_names=CUSTOMERS,
        )
        _name, check = auditor.check_p0_p1_citation_quality()
        assert check.passed is False
        assert len(check.details["violations"]) > 0

    def test_p0_missing_exact_quote_fails(self, tmp_path: Path) -> None:
        """P0 finding missing exact_quote should fail."""
        run_dir = tmp_path / "run"
        inventory_dir = tmp_path / "inventory"
        _populate_run_dir(run_dir, inventory_dir)

        merged_dir = run_dir / "findings" / "merged"
        data = _make_merged_json("acme_corp")
        data["findings"][0]["severity"] = "P0"
        data["findings"][0]["citations"] = [
            {
                "source_type": "file",
                "source_path": "acme_corp/contract.pdf",
                # No exact_quote
            }
        ]
        (merged_dir / "acme_corp.json").write_text(json.dumps(data))

        auditor = QAAuditor(
            run_dir=run_dir,
            inventory_dir=inventory_dir,
            customer_safe_names=CUSTOMERS,
        )
        _name, check = auditor.check_p0_p1_citation_quality()
        assert check.passed is False


# ===================================================================== #
# Issue #47 -- Merge/dedup empty citation collision
# ===================================================================== #


class TestMergeDedupEmptyCitation:
    """Tests for the merge dedup fix with empty citations."""

    def test_distinct_findings_not_collapsed_when_citations_empty(self) -> None:
        """Two findings with empty citations should NOT be deduped into one."""
        from dd_agents.reporting.merge import FindingMerger

        f1 = {
            "severity": "P2",
            "category": "test",
            "title": "Finding A",
            "description": "Description A",
            "citations": [],
            "confidence": "medium",
            "agent": "legal",
        }
        f2 = {
            "severity": "P2",
            "category": "test",
            "title": "Finding B",
            "description": "Description B",
            "citations": [],
            "confidence": "medium",
            "agent": "finance",
        }
        merger = FindingMerger(run_id="test_run", timestamp="2025-01-01T00:00:00Z")
        deduped = merger._deduplicate([f1, f2])
        assert len(deduped) == 2, "Distinct findings with empty citations must NOT be collapsed"

    def test_same_citation_findings_still_deduped(self) -> None:
        """Two findings pointing to the same citation should still be deduped."""
        from dd_agents.reporting.merge import FindingMerger

        cit = [{"source_type": "file", "source_path": "contract.pdf", "location": "Section 5"}]
        f1 = {
            "severity": "P1",
            "category": "test",
            "title": "Finding A",
            "description": "A",
            "citations": cit,
            "confidence": "high",
            "agent": "legal",
        }
        f2 = {
            "severity": "P2",
            "category": "test",
            "title": "Finding B",
            "description": "B",
            "citations": cit,
            "confidence": "medium",
            "agent": "finance",
        }
        merger = FindingMerger(run_id="test_run", timestamp="2025-01-01T00:00:00Z")
        deduped = merger._deduplicate([f1, f2])
        assert len(deduped) == 1, "Findings with same citation should be deduped"


# ===================================================================== #
# Issue #60 -- Gaps collected through merge pipeline
# ===================================================================== #


class TestGapsMerge:
    """Tests for gap collection through merge pipeline."""

    def test_gaps_collected_from_agent_outputs(self) -> None:
        """Gaps from specialist agents should appear in merged output."""
        from dd_agents.reporting.merge import FindingMerger

        agent_outputs = {
            "legal": {
                "customer": "Customer A",
                "customer_safe_name": "customer_a",
                "findings": [],
                "gaps": [
                    {
                        "customer": "Customer A",
                        "priority": "P1",
                        "gap_type": "Missing_Doc",
                        "missing_item": "Renewal terms",
                        "why_needed": "Required for term analysis",
                        "risk_if_missing": "Cannot assess renewal risk",
                        "request_to_company": "Please provide renewal terms",
                        "evidence": "Referenced in MSA Section 12",
                        "detection_method": "cross_reference",
                    }
                ],
            },
            "finance": {
                "customer": "Customer A",
                "customer_safe_name": "customer_a",
                "findings": [],
                "gaps": [
                    {
                        "customer": "Customer A",
                        "priority": "P2",
                        "gap_type": "Missing_Data",
                        "missing_item": "Revenue schedule",
                        "why_needed": "Financial analysis",
                        "risk_if_missing": "Incomplete financial picture",
                        "request_to_company": "Please provide revenue data",
                        "evidence": "Not found in data room",
                        "detection_method": "checklist",
                    }
                ],
            },
        }

        merger = FindingMerger(run_id="test_run", timestamp="2025-01-01T00:00:00Z")
        result = merger.merge_customer(
            agent_outputs,
            customer_name="Customer A",
            customer_safe_name="customer_a",
        )

        assert len(result.gaps) == 2, "Both gaps should be collected"
        gap_items = {g.missing_item for g in result.gaps}
        assert "Renewal terms" in gap_items
        assert "Revenue schedule" in gap_items

    def test_merged_output_model_has_gaps_field(self) -> None:
        """MergedCustomerOutput should have a gaps field."""
        from dd_agents.models.finding import MergedCustomerOutput

        mco = MergedCustomerOutput(
            customer="Test",
            customer_safe_name="test",
        )
        assert hasattr(mco, "gaps")
        assert mco.gaps == []


# ===================================================================== #
# Priority 1: DoD checks 7, 8, 9, 12, 21, 30 -- pass and fail
# ===================================================================== #


class TestDoDCheck7CrossCustomerPatterns:
    """Tests for DoD check 7 (cross-customer pattern check)."""

    def test_check_7_passes_when_audit_json_has_passing_cross_reference(self, tmp_path: Path) -> None:
        """check_7 should pass when audit.json has a passing cross_reference_completeness check."""
        run_dir = tmp_path / "run"
        run_dir.mkdir(parents=True)
        inventory_dir = tmp_path / "inventory"
        inventory_dir.mkdir(parents=True)

        audit_data = {
            "checks": {
                "cross_reference_completeness": {"passed": True, "details": {}},
            }
        }
        (run_dir / "audit.json").write_text(json.dumps(audit_data))

        checker = DefinitionOfDoneChecker(
            run_dir=run_dir,
            inventory_dir=inventory_dir,
            customer_safe_names=CUSTOMERS,
        )
        check = checker.check_7_cross_customer_patterns()
        assert check.passed is True
        assert 7 in check.dod_checks

    def test_check_7_fails_when_audit_json_missing(self, tmp_path: Path) -> None:
        """check_7 should fail when audit.json does not exist."""
        run_dir = tmp_path / "run"
        run_dir.mkdir(parents=True)
        inventory_dir = tmp_path / "inventory"
        inventory_dir.mkdir(parents=True)

        checker = DefinitionOfDoneChecker(
            run_dir=run_dir,
            inventory_dir=inventory_dir,
            customer_safe_names=CUSTOMERS,
        )
        check = checker.check_7_cross_customer_patterns()
        assert check.passed is False
        assert check.details.get("cross_customer_check") == "qa_audit_not_run_yet"

    def test_check_7_fails_when_cross_reference_not_passed(self, tmp_path: Path) -> None:
        """check_7 should fail when cross_reference_completeness is false."""
        run_dir = tmp_path / "run"
        run_dir.mkdir(parents=True)
        inventory_dir = tmp_path / "inventory"
        inventory_dir.mkdir(parents=True)

        audit_data = {
            "checks": {
                "cross_reference_completeness": {"passed": False, "details": {"issues": ["gap"]}},
            }
        }
        (run_dir / "audit.json").write_text(json.dumps(audit_data))

        checker = DefinitionOfDoneChecker(
            run_dir=run_dir,
            inventory_dir=inventory_dir,
            customer_safe_names=CUSTOMERS,
        )
        check = checker.check_7_cross_customer_patterns()
        assert check.passed is False

    def test_check_7_fails_on_malformed_audit_json(self, tmp_path: Path) -> None:
        """check_7 should fail gracefully when audit.json is malformed."""
        run_dir = tmp_path / "run"
        run_dir.mkdir(parents=True)
        inventory_dir = tmp_path / "inventory"
        inventory_dir.mkdir(parents=True)

        (run_dir / "audit.json").write_text("NOT VALID JSON {{")

        checker = DefinitionOfDoneChecker(
            run_dir=run_dir,
            inventory_dir=inventory_dir,
            customer_safe_names=CUSTOMERS,
        )
        check = checker.check_7_cross_customer_patterns()
        assert check.passed is False


class TestDoDCheck8CrossReferenceReconciliation:
    """Tests for DoD check 8 (cross-reference reconciliation)."""

    def test_check_8_passes_when_audit_json_has_passing_check(self, tmp_path: Path) -> None:
        """check_8 should pass when cross_reference_completeness is true."""
        run_dir = tmp_path / "run"
        run_dir.mkdir(parents=True)
        inventory_dir = tmp_path / "inventory"
        inventory_dir.mkdir(parents=True)

        audit_data = {
            "checks": {
                "cross_reference_completeness": {"passed": True},
            }
        }
        (run_dir / "audit.json").write_text(json.dumps(audit_data))

        checker = DefinitionOfDoneChecker(
            run_dir=run_dir,
            inventory_dir=inventory_dir,
            customer_safe_names=CUSTOMERS,
        )
        check = checker.check_8_cross_reference_reconciliation()
        assert check.passed is True
        assert 8 in check.dod_checks

    def test_check_8_fails_when_audit_json_missing(self, tmp_path: Path) -> None:
        """check_8 should fail when audit.json does not exist."""
        run_dir = tmp_path / "run"
        run_dir.mkdir(parents=True)
        inventory_dir = tmp_path / "inventory"
        inventory_dir.mkdir(parents=True)

        checker = DefinitionOfDoneChecker(
            run_dir=run_dir,
            inventory_dir=inventory_dir,
            customer_safe_names=CUSTOMERS,
        )
        check = checker.check_8_cross_reference_reconciliation()
        assert check.passed is False
        assert check.details.get("reconciliation_check") == "qa_audit_not_run_yet"

    def test_check_8_fails_when_reconciliation_not_passed(self, tmp_path: Path) -> None:
        """check_8 should fail when cross_reference_completeness is false."""
        run_dir = tmp_path / "run"
        run_dir.mkdir(parents=True)
        inventory_dir = tmp_path / "inventory"
        inventory_dir.mkdir(parents=True)

        audit_data = {
            "checks": {
                "cross_reference_completeness": {"passed": False},
            }
        }
        (run_dir / "audit.json").write_text(json.dumps(audit_data))

        checker = DefinitionOfDoneChecker(
            run_dir=run_dir,
            inventory_dir=inventory_dir,
            customer_safe_names=CUSTOMERS,
        )
        check = checker.check_8_cross_reference_reconciliation()
        assert check.passed is False


class TestDoDCheck9GhostCustomers:
    """Tests for DoD check 9 (ghost customers logged as P0 gaps)."""

    def test_check_9_passes_when_gap_completeness_passes(self, tmp_path: Path) -> None:
        """check_9 should pass when audit.json gap_completeness is true."""
        run_dir = tmp_path / "run"
        run_dir.mkdir(parents=True)
        inventory_dir = tmp_path / "inventory"
        inventory_dir.mkdir(parents=True)

        audit_data = {
            "checks": {
                "gap_completeness": {"passed": True},
            }
        }
        (run_dir / "audit.json").write_text(json.dumps(audit_data))

        checker = DefinitionOfDoneChecker(
            run_dir=run_dir,
            inventory_dir=inventory_dir,
            customer_safe_names=CUSTOMERS,
        )
        check = checker.check_9_ghost_customers()
        assert check.passed is True
        assert 9 in check.dod_checks

    def test_check_9_fails_when_audit_json_missing(self, tmp_path: Path) -> None:
        """check_9 should fail when audit.json does not exist."""
        run_dir = tmp_path / "run"
        run_dir.mkdir(parents=True)
        inventory_dir = tmp_path / "inventory"
        inventory_dir.mkdir(parents=True)

        checker = DefinitionOfDoneChecker(
            run_dir=run_dir,
            inventory_dir=inventory_dir,
            customer_safe_names=CUSTOMERS,
        )
        check = checker.check_9_ghost_customers()
        assert check.passed is False
        assert check.details.get("ghost_check") == "qa_audit_not_run_yet"

    def test_check_9_fails_when_gap_completeness_false(self, tmp_path: Path) -> None:
        """check_9 should fail when gap_completeness.passed is false."""
        run_dir = tmp_path / "run"
        run_dir.mkdir(parents=True)
        inventory_dir = tmp_path / "inventory"
        inventory_dir.mkdir(parents=True)

        audit_data = {
            "checks": {
                "gap_completeness": {"passed": False, "ghost_customers": 3},
            }
        }
        (run_dir / "audit.json").write_text(json.dumps(audit_data))

        checker = DefinitionOfDoneChecker(
            run_dir=run_dir,
            inventory_dir=inventory_dir,
            customer_safe_names=CUSTOMERS,
        )
        check = checker.check_9_ghost_customers()
        assert check.passed is False


class TestDoDCheck12DomainCoverage:
    """Tests for DoD check 12 (domain coverage)."""

    def test_check_12_passes_when_domain_coverage_passes(self, tmp_path: Path) -> None:
        """check_12 should pass when audit.json domain_coverage is true."""
        run_dir = tmp_path / "run"
        run_dir.mkdir(parents=True)
        inventory_dir = tmp_path / "inventory"
        inventory_dir.mkdir(parents=True)

        audit_data = {
            "checks": {
                "domain_coverage": {"passed": True},
            }
        }
        (run_dir / "audit.json").write_text(json.dumps(audit_data))

        checker = DefinitionOfDoneChecker(
            run_dir=run_dir,
            inventory_dir=inventory_dir,
            customer_safe_names=CUSTOMERS,
        )
        check = checker.check_12_domain_coverage()
        assert check.passed is True
        assert 12 in check.dod_checks

    def test_check_12_fails_when_audit_json_missing(self, tmp_path: Path) -> None:
        """check_12 should fail when audit.json does not exist."""
        run_dir = tmp_path / "run"
        run_dir.mkdir(parents=True)
        inventory_dir = tmp_path / "inventory"
        inventory_dir.mkdir(parents=True)

        checker = DefinitionOfDoneChecker(
            run_dir=run_dir,
            inventory_dir=inventory_dir,
            customer_safe_names=CUSTOMERS,
        )
        check = checker.check_12_domain_coverage()
        assert check.passed is False
        assert check.details.get("domain_coverage_check") == "qa_audit_not_run_yet"

    def test_check_12_fails_when_domain_coverage_false(self, tmp_path: Path) -> None:
        """check_12 should fail when domain_coverage.passed is false."""
        run_dir = tmp_path / "run"
        run_dir.mkdir(parents=True)
        inventory_dir = tmp_path / "inventory"
        inventory_dir.mkdir(parents=True)

        audit_data = {
            "checks": {
                "domain_coverage": {"passed": False, "missing_domains": ["legal"]},
            }
        }
        (run_dir / "audit.json").write_text(json.dumps(audit_data))

        checker = DefinitionOfDoneChecker(
            run_dir=run_dir,
            inventory_dir=inventory_dir,
            customer_safe_names=CUSTOMERS,
        )
        check = checker.check_12_domain_coverage()
        assert check.passed is False


class TestDoDCheck21P0SpotChecked:
    """Tests for DoD check 21 (P0 findings spot-checked by Judge)."""

    def test_check_21_passes_with_p0_spot_checks(self, tmp_path: Path) -> None:
        """check_21 passes when quality_scores.json has P0 spot checks."""
        run_dir = tmp_path / "run"
        judge_dir = run_dir / "judge"
        judge_dir.mkdir(parents=True)
        inventory_dir = tmp_path / "inventory"
        inventory_dir.mkdir(parents=True)

        scores_data = {
            "agent_scores": {
                "legal": {"score": 85},
            },
            "spot_checks": [
                {"severity": "P0", "finding_id": "F1", "result": "pass"},
                {"severity": "P1", "finding_id": "F2", "result": "pass"},
            ],
        }
        (judge_dir / "quality_scores.json").write_text(json.dumps(scores_data))

        checker = DefinitionOfDoneChecker(
            run_dir=run_dir,
            inventory_dir=inventory_dir,
            customer_safe_names=CUSTOMERS,
            deal_config={"judge": {"enabled": True}},
        )
        check = checker.check_21_p0_spot_checked()
        assert check.passed is True
        assert 21 in check.dod_checks

    def test_check_21_passes_when_no_p0_findings_exist(self, tmp_path: Path) -> None:
        """check_21 passes when there are no P0 findings at all."""
        run_dir = tmp_path / "run"
        judge_dir = run_dir / "judge"
        judge_dir.mkdir(parents=True)
        inventory_dir = tmp_path / "inventory"
        inventory_dir.mkdir(parents=True)

        # No P0 findings or spot checks
        scores_data = {
            "agent_scores": {
                "legal": {"score": 85},
            },
            "spot_checks": [
                {"severity": "P2", "finding_id": "F1", "result": "pass"},
            ],
        }
        (judge_dir / "quality_scores.json").write_text(json.dumps(scores_data))

        checker = DefinitionOfDoneChecker(
            run_dir=run_dir,
            inventory_dir=inventory_dir,
            customer_safe_names=CUSTOMERS,
            deal_config={"judge": {"enabled": True}},
        )
        check = checker.check_21_p0_spot_checked()
        assert check.passed is True

    def test_check_21_fails_when_quality_scores_missing(self, tmp_path: Path) -> None:
        """check_21 fails when quality_scores.json does not exist."""
        run_dir = tmp_path / "run"
        run_dir.mkdir(parents=True)
        inventory_dir = tmp_path / "inventory"
        inventory_dir.mkdir(parents=True)

        checker = DefinitionOfDoneChecker(
            run_dir=run_dir,
            inventory_dir=inventory_dir,
            customer_safe_names=CUSTOMERS,
            deal_config={"judge": {"enabled": True}},
        )
        check = checker.check_21_p0_spot_checked()
        assert check.passed is False
        assert check.details.get("error") == "quality_scores.json missing"

    def test_check_21_fails_on_malformed_quality_scores(self, tmp_path: Path) -> None:
        """check_21 fails gracefully on invalid JSON."""
        run_dir = tmp_path / "run"
        judge_dir = run_dir / "judge"
        judge_dir.mkdir(parents=True)
        inventory_dir = tmp_path / "inventory"
        inventory_dir.mkdir(parents=True)

        (judge_dir / "quality_scores.json").write_text("INVALID JSON")

        checker = DefinitionOfDoneChecker(
            run_dir=run_dir,
            inventory_dir=inventory_dir,
            customer_safe_names=CUSTOMERS,
            deal_config={"judge": {"enabled": True}},
        )
        check = checker.check_21_p0_spot_checked()
        assert check.passed is False
        assert "invalid" in check.details.get("error", "")


class TestDoDCheck30ReportDiff:
    """Tests for DoD check 30 (report diff)."""

    def test_check_30_passes_when_prior_run_and_diff_exist(self, tmp_path: Path) -> None:
        """check_30 passes when prior_run_id is set and report_diff.json exists."""
        run_dir = tmp_path / "run"
        run_dir.mkdir(parents=True)
        inventory_dir = tmp_path / "inventory"
        inventory_dir.mkdir(parents=True)

        (run_dir / "report_diff.json").write_text("{}")

        checker = DefinitionOfDoneChecker(
            run_dir=run_dir,
            inventory_dir=inventory_dir,
            customer_safe_names=CUSTOMERS,
            deal_config={"execution": {"prior_run_id": "run_prev"}},
        )
        check = checker.check_30_report_diff()
        assert check.passed is True
        assert 30 in check.dod_checks
        assert check.details["prior_run_id"] == "run_prev"

    def test_check_30_fails_when_prior_run_but_no_diff(self, tmp_path: Path) -> None:
        """check_30 fails when prior_run_id is set but report_diff.json is missing."""
        run_dir = tmp_path / "run"
        run_dir.mkdir(parents=True)
        inventory_dir = tmp_path / "inventory"
        inventory_dir.mkdir(parents=True)

        checker = DefinitionOfDoneChecker(
            run_dir=run_dir,
            inventory_dir=inventory_dir,
            customer_safe_names=CUSTOMERS,
            deal_config={"execution": {"prior_run_id": "run_prev"}},
        )
        check = checker.check_30_report_diff()
        assert check.passed is False
        assert check.details["prior_run_id"] == "run_prev"

    def test_check_30_passes_when_no_prior_run(self, tmp_path: Path) -> None:
        """check_30 passes by default when no prior_run_id is configured."""
        run_dir = tmp_path / "run"
        run_dir.mkdir(parents=True)
        inventory_dir = tmp_path / "inventory"
        inventory_dir.mkdir(parents=True)

        checker = DefinitionOfDoneChecker(
            run_dir=run_dir,
            inventory_dir=inventory_dir,
            customer_safe_names=CUSTOMERS,
            deal_config={},
        )
        check = checker.check_30_report_diff()
        assert check.passed is True
        assert check.details["prior_run_id"] == ""

    def test_check_30_passes_when_empty_prior_run_id(self, tmp_path: Path) -> None:
        """check_30 passes when prior_run_id is an empty string."""
        run_dir = tmp_path / "run"
        run_dir.mkdir(parents=True)
        inventory_dir = tmp_path / "inventory"
        inventory_dir.mkdir(parents=True)

        checker = DefinitionOfDoneChecker(
            run_dir=run_dir,
            inventory_dir=inventory_dir,
            customer_safe_names=CUSTOMERS,
            deal_config={"execution": {"prior_run_id": ""}},
        )
        check = checker.check_30_report_diff()
        assert check.passed is True

    # ------------------------------------------------------------------ #
    # DoD [13] -- merged file count filters non-customer files
    # ------------------------------------------------------------------ #

    def test_check_13_filters_non_customer_files(self, tmp_path: Path) -> None:
        """check_13 passes when customer files exist even with extra non-customer files."""
        run_dir = tmp_path / "run"
        inventory_dir = tmp_path / "inventory"
        inventory_dir.mkdir(parents=True)
        merged_dir = run_dir / "findings" / "merged"
        merged_dir.mkdir(parents=True)
        # Write customer files
        for c in CUSTOMERS:
            (merged_dir / f"{c}.json").write_text(json.dumps(_make_merged_json(c)))
        # Write stale non-customer files (e.g. Reporting Lead artefacts)
        (merged_dir / "coverage_manifest.json").write_text("{}")
        (merged_dir / "numerical_audit.json").write_text("{}")
        (merged_dir / "report_diff.json").write_text("{}")

        checker = DefinitionOfDoneChecker(
            run_dir=run_dir,
            inventory_dir=inventory_dir,
            customer_safe_names=CUSTOMERS,
        )
        check = checker.check_13_merge_dedup_complete()
        assert check.passed is True
        assert check.details["merged_count"] == 3
        assert check.details["total_json_files"] == 6  # 3 customer + 3 stale

    def test_check_13_fails_when_customer_missing(self, tmp_path: Path) -> None:
        """check_13 fails when a customer file is missing."""
        run_dir = tmp_path / "run"
        inventory_dir = tmp_path / "inventory"
        inventory_dir.mkdir(parents=True)
        merged_dir = run_dir / "findings" / "merged"
        merged_dir.mkdir(parents=True)
        # Only write 2 of 3 customers
        for c in CUSTOMERS[:2]:
            (merged_dir / f"{c}.json").write_text(json.dumps(_make_merged_json(c)))

        checker = DefinitionOfDoneChecker(
            run_dir=run_dir,
            inventory_dir=inventory_dir,
            customer_safe_names=CUSTOMERS,
        )
        check = checker.check_13_merge_dedup_complete()
        assert check.passed is False
        assert check.details["merged_count"] == 2

    # ------------------------------------------------------------------ #
    # DoD [11] -- audit log fallback to QA audit
    # ------------------------------------------------------------------ #

    def test_check_11_passes_with_qa_audit_fallback(self, tmp_path: Path) -> None:
        """check_11 passes when audit logs are missing but QA audit passed."""
        run_dir = tmp_path / "run"
        run_dir.mkdir(parents=True)
        inventory_dir = tmp_path / "inventory"
        inventory_dir.mkdir(parents=True)
        # Write audit.json with audit_passed=True
        (run_dir / "audit.json").write_text(json.dumps({"audit_passed": True, "checks": {}}))
        # Do NOT create audit log directories

        checker = DefinitionOfDoneChecker(
            run_dir=run_dir,
            inventory_dir=inventory_dir,
            customer_safe_names=CUSTOMERS,
        )
        check = checker.check_11_audit_logs_exist()
        assert check.passed is True
        assert check.details.get("fallback") == "qa_audit_passed"

    def test_check_11_fails_when_no_logs_and_no_audit(self, tmp_path: Path) -> None:
        """check_11 fails when both audit logs and audit.json are missing."""
        run_dir = tmp_path / "run"
        run_dir.mkdir(parents=True)
        inventory_dir = tmp_path / "inventory"
        inventory_dir.mkdir(parents=True)

        checker = DefinitionOfDoneChecker(
            run_dir=run_dir,
            inventory_dir=inventory_dir,
            customer_safe_names=CUSTOMERS,
        )
        check = checker.check_11_audit_logs_exist()
        assert check.passed is False

    # ------------------------------------------------------------------ #
    # DoD [19] -- extraction quality fallback path
    # ------------------------------------------------------------------ #

    def test_check_19_finds_alt_path(self, tmp_path: Path) -> None:
        """check_19 passes when extraction_quality.json is in index/text/."""
        run_dir = tmp_path / "run"
        run_dir.mkdir(parents=True)
        inventory_dir = tmp_path / "data" / "inventory"
        inventory_dir.mkdir(parents=True)
        # Write to index/text/ (sibling of inventory's parent)
        alt_dir = tmp_path / "data" / "index" / "text"
        alt_dir.mkdir(parents=True)
        (alt_dir / "extraction_quality.json").write_text("[]")

        checker = DefinitionOfDoneChecker(
            run_dir=run_dir,
            inventory_dir=inventory_dir,
            customer_safe_names=CUSTOMERS,
        )
        check = checker.check_19_extraction_quality()
        assert check.passed is True

    # ------------------------------------------------------------------ #
    # DoD [10] -- reference files defers to QA audit
    # ------------------------------------------------------------------ #

    def test_check_10_defers_to_qa_audit(self, tmp_path: Path) -> None:
        """check_10 uses QA audit cross_reference_completeness when available."""
        run_dir = tmp_path / "run"
        run_dir.mkdir(parents=True)
        inventory_dir = tmp_path / "inventory"
        inventory_dir.mkdir(parents=True)
        (run_dir / "audit.json").write_text(
            json.dumps(
                {
                    "audit_passed": True,
                    "checks": {"cross_reference_completeness": {"passed": True, "details": {}}},
                }
            )
        )

        checker = DefinitionOfDoneChecker(
            run_dir=run_dir,
            inventory_dir=inventory_dir,
            customer_safe_names=CUSTOMERS,
        )
        check = checker.check_10_reference_files_processed()
        assert check.passed is True
        assert check.details.get("reference_files_check") == "from_qa_audit"

    # ------------------------------------------------------------------ #
    # DoD [16] -- entity resolution with empty result
    # ------------------------------------------------------------------ #

    def test_check_16_passes_with_empty_entity_matches(self, tmp_path: Path) -> None:
        """check_16 passes when entity_matches.json is an empty list."""
        run_dir = tmp_path / "run"
        run_dir.mkdir(parents=True)
        inventory_dir = tmp_path / "inventory"
        inventory_dir.mkdir(parents=True)
        (run_dir / "entity_matches.json").write_text("[]")

        checker = DefinitionOfDoneChecker(
            run_dir=run_dir,
            inventory_dir=inventory_dir,
            customer_safe_names=CUSTOMERS,
        )
        check = checker.check_16_entity_resolution_log()
        assert check.passed is True
        assert check.details.get("note") == "empty_result"

    def test_check_16_passes_with_empty_dict(self, tmp_path: Path) -> None:
        """check_16 passes when entity_matches.json is an empty dict."""
        run_dir = tmp_path / "run"
        run_dir.mkdir(parents=True)
        inventory_dir = tmp_path / "inventory"
        inventory_dir.mkdir(parents=True)
        (run_dir / "entity_matches.json").write_text("{}")

        checker = DefinitionOfDoneChecker(
            run_dir=run_dir,
            inventory_dir=inventory_dir,
            customer_safe_names=CUSTOMERS,
        )
        check = checker.check_16_entity_resolution_log()
        assert check.passed is True
