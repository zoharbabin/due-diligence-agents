"""Tests for Excel formula extraction + model-integrity audit (Issue #194).

Builds synthetic .xlsx fixtures in-test with openpyxl (already a dependency) —
no committed binary. Covers: formula capture, hardcoded-override detection,
broken external link, error literal, circular self-reference, and the
no-formula / non-xlsx fall-through.
"""

from __future__ import annotations

from typing import TYPE_CHECKING

from dd_agents.extraction.formula_audit import (
    audit_formulas,
    extract_formula_map,
    format_formula_audit,
)

if TYPE_CHECKING:
    from pathlib import Path


def _write_xlsx(path: Path, sheets: dict[str, dict[str, object]]) -> None:
    """Write an .xlsx where each sheet maps cell-coord -> value/formula."""
    from openpyxl import Workbook

    wb = Workbook()
    # Remove the default sheet so we control names deterministically.
    default = wb.active
    first = True
    for name, cells in sheets.items():
        ws = default if first else wb.create_sheet(title=name)
        if first:
            ws.title = name
            first = False
        for coord, val in cells.items():
            ws[coord] = val
    wb.save(path)
    wb.close()


class TestExtractFormulaMap:
    def test_captures_formula_text(self, tmp_path: Path) -> None:
        p = tmp_path / "model.xlsx"
        _write_xlsx(p, {"P&L": {"A1": 10, "A2": 20, "A3": "=SUM(A1:A2)"}})
        fmap = extract_formula_map(p)
        assert fmap == {"P&L": {"A3": "=SUM(A1:A2)"}}

    def test_non_xlsx_returns_empty(self, tmp_path: Path) -> None:
        p = tmp_path / "notes.txt"
        p.write_text("hello")
        assert extract_formula_map(p) == {}

    def test_no_formulas_returns_empty(self, tmp_path: Path) -> None:
        p = tmp_path / "values.xlsx"
        _write_xlsx(p, {"Data": {"A1": 1, "A2": 2}})
        assert extract_formula_map(p) == {}


class TestAuditFormulas:
    def test_hardcoded_override_detected(self) -> None:
        # A column of real formulas with one typed-in numeric literal.
        fmap = {
            "Model": {
                "B2": "=A2*1.1",
                "B3": "=A3*1.1",
                "B4": "=A4*1.1",
                "B5": "=1234",  # hardcoded override
            }
        }
        issues = audit_formulas(fmap)
        overrides = [i for i in issues if i.kind == "hardcoded_override"]
        assert len(overrides) == 1
        assert overrides[0].cell == "B5"
        assert overrides[0].sheet == "Model"

    def test_constant_column_not_flagged(self) -> None:
        # All-literal column must NOT be flagged as overrides (no real formulas).
        fmap = {"Const": {"A1": "=1", "A2": "=2", "A3": "=3", "A4": "=4"}}
        issues = audit_formulas(fmap)
        assert [i for i in issues if i.kind == "hardcoded_override"] == []

    def test_broken_external_link_detected(self) -> None:
        fmap = {"Links": {"C1": "=[1]Sheet1!A1*2", "C2": "='[Budget.xlsx]Q1'!B2"}}
        issues = audit_formulas(fmap)
        links = [i for i in issues if i.kind == "broken_external_link"]
        assert {i.cell for i in links} == {"C1", "C2"}

    def test_error_literal_detected(self) -> None:
        fmap = {"Errs": {"D1": "=A1/#REF!", "D2": "=1/0+#DIV/0!"}}
        issues = audit_formulas(fmap)
        errs = [i for i in issues if i.kind == "error_literal"]
        assert {i.cell for i in errs} == {"D1", "D2"}

    def test_circular_self_reference_detected(self) -> None:
        fmap = {"Circ": {"E5": "=E5+1"}}
        issues = audit_formulas(fmap)
        circ = [i for i in issues if i.kind == "circular_reference"]
        assert len(circ) == 1 and circ[0].cell == "E5"

    def test_no_false_circular_on_substring_coord(self) -> None:
        # E5 referencing E50 must NOT be flagged circular (word-boundary guard).
        fmap = {"S": {"E5": "=E50+E51"}}
        issues = audit_formulas(fmap)
        assert [i for i in issues if i.kind == "circular_reference"] == []

    def test_issues_sorted_stable(self) -> None:
        # C9 self-references (circular); A1 links externally (no self-ref).
        fmap = {"Z": {"C9": "=C9+1", "A1": "=[1]X!B2"}}
        issues = audit_formulas(fmap)
        # Sorted by (sheet, col, row): A1 (col A) before C9 (col C).
        assert [i.cell for i in issues] == ["A1", "C9"]

    def test_multi_letter_columns_sort_numerically(self) -> None:
        # Regression: "AA" (col 27) must sort AFTER "B"/"Z", not lexicographically first.
        fmap = {"S": {"AA1": "=[1]X!Q1", "B2": "=[1]X!Q2", "Z3": "=[1]X!Q3"}}
        issues = audit_formulas(fmap)
        assert [i.cell for i in issues] == ["B2", "Z3", "AA1"]


class TestFormatFormulaAudit:
    def test_empty_map_renders_nothing(self) -> None:
        assert format_formula_audit({}, []) == ""

    def test_renders_issues_and_counts(self) -> None:
        fmap = {"P&L": {"B2": "=A2*1.1", "B3": "=A3*1.1", "B4": "=A4*1.1", "B5": "=99"}}
        issues = audit_formulas(fmap)
        out = format_formula_audit(fmap, issues)
        assert "Formula Integrity" in out
        assert "hardcoded_override" in out
        assert "P&L!B5" in out
        assert "Formula cell counts by sheet: P&L (4)" in out

    def test_renders_clean_when_no_issues(self) -> None:
        fmap = {"S": {"A1": "=SUM(B1:B9)"}}
        out = format_formula_audit(fmap, audit_formulas(fmap))
        assert "No formula-integrity issues detected" in out


class TestReadOfficeIntegration:
    def test_read_office_appends_formula_section(self, tmp_path: Path) -> None:
        from dd_agents.tools.read_office import read_office

        p = tmp_path / "deal_model.xlsx"
        _write_xlsx(
            p,
            {"Rev": {"A1": 100, "A2": 200, "B1": "=A1*1.2", "B2": "=A2*1.2", "B3": "=A1*1.2", "B4": "=5000"}},
        )
        result = read_office(str(p))
        assert result["status"] == "ok"
        # Value content present AND the formula-integrity section appended.
        assert "## Sheet: Rev" in result["content"]
        assert "Formula Integrity" in result["content"]
        assert "hardcoded_override" in result["content"]

    def test_read_office_no_formula_section_for_value_only(self, tmp_path: Path) -> None:
        from dd_agents.tools.read_office import read_office

        p = tmp_path / "values_only.xlsx"
        _write_xlsx(p, {"Data": {"A1": 1, "A2": 2, "A3": 3}})
        result = read_office(str(p))
        assert result["status"] == "ok"
        # No formulas → no audit section appended (non-model spreadsheets stay clean).
        assert "Formula Integrity" not in result["content"]


class TestAuditDataRoom:
    """Data-room-level aggregation reused by the report (Issue #238)."""

    def test_empty_when_no_xlsx(self, tmp_path: Path) -> None:
        from dd_agents.extraction.formula_audit import audit_data_room

        (tmp_path / "a.txt").write_text("x")
        report = audit_data_room([tmp_path / "a.txt"])
        assert report["files_scanned"] == 0
        assert report["files_with_formulas"] == 0
        assert report["issues"] == []

    def test_aggregates_issues_with_file_attribution(self, tmp_path: Path) -> None:
        from dd_agents.extraction.formula_audit import audit_data_room

        good = tmp_path / "values.xlsx"
        _write_xlsx(good, {"Data": {"A1": 1, "A2": 2}})  # no formulas
        bad = tmp_path / "model.xlsx"
        _write_xlsx(
            bad,
            {"Model": {"B2": "=A2*1.1", "B3": "=A3*1.1", "B4": "=A4*1.1", "B5": "=1234"}},
        )
        report = audit_data_room([good, bad])
        assert report["files_scanned"] == 2
        assert report["files_with_formulas"] == 1
        assert report["files_with_issues"] == 1
        assert report["total_issues"] >= 1
        # Every issue cites its originating file.
        assert all(r["file"] == str(bad) for r in report["issues"])
        assert "hardcoded_override" in report["by_kind"]

    def test_truncation_flagged_not_silent(self, tmp_path: Path) -> None:
        from dd_agents.extraction.formula_audit import audit_data_room

        p = tmp_path / "model.xlsx"
        # Many overrides in one column → exceed a tiny max_issues cap.
        cells = {f"B{r}": f"=A{r}*1.1" for r in range(2, 8)}
        cells["B20"] = "=11"
        cells["B21"] = "=22"
        cells["B22"] = "=33"
        _write_xlsx(p, {"M": cells})
        report = audit_data_room([p], max_issues=1)
        assert report["truncated"] is True
        assert len(report["issues"]) == 1

    def test_result_is_json_serializable(self, tmp_path: Path) -> None:
        import json

        from dd_agents.extraction.formula_audit import audit_data_room

        p = tmp_path / "m.xlsx"
        _write_xlsx(p, {"S": {"C1": "=[1]Ext!A1"}})
        report = audit_data_room([p])
        json.dumps(report)  # must not raise
