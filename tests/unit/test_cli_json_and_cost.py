"""Tests for --json parity (Issue #241, #256) + the cost reader / assess pre-flight (#246)."""

from __future__ import annotations

import json
from typing import TYPE_CHECKING

from click.testing import CliRunner

from dd_agents.cli import main

if TYPE_CHECKING:
    from pathlib import Path


def _valid_config(tmp_path: Path) -> Path:
    cfg = {
        "config_version": "1.0.0",
        "buyer": {"name": "Buyer"},
        "target": {"name": "Target"},
        "deal": {"type": "acquisition", "focus_areas": ["ip_ownership"]},
    }
    p = tmp_path / "deal-config.json"
    p.write_text(json.dumps(cfg))
    return p


class TestValidateJson:
    def test_valid_emits_json(self, tmp_path: Path) -> None:
        r = CliRunner().invoke(main, ["validate", str(_valid_config(tmp_path)), "--json"])
        assert r.exit_code == 0
        out = json.loads(r.output)
        assert out["valid"] is True
        assert out["target"] == "Target"

    def test_invalid_emits_errors_json_and_exit_1(self, tmp_path: Path) -> None:
        bad = tmp_path / "bad.json"
        bad.write_text(json.dumps({"config_version": "1.0.0"}))  # missing buyer/target/deal
        r = CliRunner().invoke(main, ["validate", str(bad), "--json"])
        assert r.exit_code == 1
        out = json.loads(r.output)
        assert out["valid"] is False
        assert len(out["errors"]) >= 1
        assert "loc" in out["errors"][0]

    def test_missing_file_emits_json(self, tmp_path: Path) -> None:
        r = CliRunner().invoke(main, ["validate", str(tmp_path / "nope.json"), "--json"])
        assert r.exit_code == 1
        assert json.loads(r.output)["valid"] is False


class TestAssessJson:
    def _data_room(self, tmp_path: Path) -> Path:
        dr = tmp_path / "dr"
        (dr / "Acme").mkdir(parents=True)
        (dr / "Acme" / "msa.pdf").write_text("contract text " * 50)
        (dr / "Acme" / "more.pdf").write_text("more text " * 50)
        (dr / "Acme" / "third.pdf").write_text("third " * 50)
        (dr / "Acme" / "fourth.pdf").write_text("fourth " * 50)
        (dr / "Acme" / "fifth.pdf").write_text("fifth " * 50)
        return dr

    def test_assess_emits_json_report(self, tmp_path: Path) -> None:
        r = CliRunner().invoke(main, ["assess", str(self._data_room(tmp_path)), "--json"])
        assert r.exit_code == 0, r.output
        out = json.loads(r.output)
        assert "overall_score" in out
        assert "vdr_convention" in out

    def test_assess_formula_preflight_flags_bad_model(self, tmp_path: Path) -> None:
        from openpyxl import Workbook

        dr = self._data_room(tmp_path)
        wb = Workbook()
        ws = wb.active
        ws.title = "Model"
        ws["B2"] = "=A2*1.1"
        ws["B3"] = "=A3*1.1"
        ws["B4"] = "=A4*1.1"
        ws["B5"] = "=1234"  # hardcoded override
        wb.save(dr / "Acme" / "model.xlsx")
        wb.close()

        r = CliRunner().invoke(main, ["assess", str(dr), "--json"])
        assert r.exit_code == 0, r.output
        out = json.loads(r.output)
        assert out.get("formula_audit", {}).get("total_issues", 0) >= 1
        # File path is data-room-relative (no local-path leak).
        issues = out["formula_audit"]["issues"]
        assert all(not str(i["file"]).startswith("/") for i in issues)


class TestCostCommand:
    def _run_dir(self, tmp_path: Path, summary: dict) -> Path:  # type: ignore[type-arg]
        rd = tmp_path / "run"
        rd.mkdir()
        (rd / "cost_summary.json").write_text(json.dumps(summary))
        return rd

    _SUMMARY = {
        "total_cost": 1.2345,
        "total_tokens": 50000,
        "budget_limit_usd": None,
        "by_agent": {"legal": 0.5, "finance": 0.7345},
        "by_step": {"16_spawn": 1.2345},
        "by_model": {"claude-sonnet-4-6": {"cost": 1.2345, "estimated": False}},
        "by_provider": {"(run default)": {"cost": 1.2345, "base_url": None}},
    }

    def test_cost_json(self, tmp_path: Path) -> None:
        rd = self._run_dir(tmp_path, self._SUMMARY)
        r = CliRunner().invoke(main, ["cost", str(rd), "--json"])
        assert r.exit_code == 0, r.output
        out = json.loads(r.output)
        assert out["total_cost"] == 1.2345
        assert "by_provider" in out

    def test_cost_rich_output(self, tmp_path: Path) -> None:
        rd = self._run_dir(tmp_path, self._SUMMARY)
        r = CliRunner().invoke(main, ["cost", str(rd)])
        assert r.exit_code == 0, r.output
        assert "By Provider" in r.output
        assert "By Model" in r.output

    def test_cost_missing_summary_exits_1(self, tmp_path: Path) -> None:
        empty = tmp_path / "empty"
        empty.mkdir()
        r = CliRunner().invoke(main, ["cost", str(empty), "--json"])
        assert r.exit_code == 1
        assert "error" in json.loads(r.output)

    def test_cost_parent_dir_picks_newest_run(self, tmp_path: Path) -> None:
        # Audit fix: a parent dir must resolve to the NEWEST run, not the oldest.
        runs = tmp_path / "runs"
        for name, cost_val in (("run_20260101_010000_a", 1.0), ("run_20260615_120000_c", 9.0)):
            rd = runs / name
            rd.mkdir(parents=True)
            (rd / "cost_summary.json").write_text(json.dumps({**self._SUMMARY, "total_cost": cost_val}))
        r = CliRunner().invoke(main, ["cost", str(runs), "--json"])
        assert r.exit_code == 0, r.output
        assert json.loads(r.output)["total_cost"] == 9.0  # newest, not oldest (1.0)

    def test_cost_rich_shows_step_and_routing(self, tmp_path: Path) -> None:
        summary = {**self._SUMMARY, "routing": {"provider": "bedrock", "base_url": None, "models_used": ["m"]}}
        rd = self._run_dir(tmp_path, summary)
        r = CliRunner().invoke(main, ["cost", str(rd)])
        assert "By Step" in r.output
        assert "Routing:" in r.output


class TestHealthJson:
    def test_health_json_no_knowledge_base(self, tmp_path: Path) -> None:
        r = CliRunner().invoke(main, ["health", "--data-room", str(tmp_path), "--json"])
        assert r.exit_code == 0, r.output
        assert "error" in json.loads(r.output)

    def test_health_json_matches_model_dump(self, tmp_path: Path) -> None:
        from dd_agents.knowledge.base import DealKnowledgeBase

        kb = DealKnowledgeBase(tmp_path)
        kb.ensure_dirs()

        r = CliRunner().invoke(main, ["health", "--data-room", str(tmp_path), "--json"])
        assert r.exit_code == 0, r.output
        out = json.loads(r.output)
        assert "total_issues" in out
        assert "knowledge_base_stats" in out

    def test_health_no_json_prints_console_chrome(self, tmp_path: Path) -> None:
        from dd_agents.knowledge.base import DealKnowledgeBase

        kb = DealKnowledgeBase(tmp_path)
        kb.ensure_dirs()

        r = CliRunner().invoke(main, ["health", "--data-room", str(tmp_path)])
        assert r.exit_code == 0, r.output
        # Rich panel/table output, not raw JSON.
        assert not r.output.lstrip().startswith("{")


class TestMemoJson:
    def _run_with_findings(self, tmp_path: Path) -> Path:
        run = tmp_path / "_dd" / "forensic-dd" / "runs" / "run_x"
        merged = run / "findings" / "merged"
        merged.mkdir(parents=True)
        (merged / "northwind.json").write_text(
            json.dumps(
                {
                    "subject": "Northwind",
                    "findings": [
                        {
                            "title": "CoC auto-termination",
                            "severity": "P0",
                            "description": "Customer MSA terminates on change of control.",
                            "citations": [{"exact_quote": "terminates on change of control", "location": "§12.3"}],
                        }
                    ],
                    "gaps": [],
                }
            )
        )
        return run

    def test_memo_json_has_go_no_go_and_risks(self, tmp_path: Path) -> None:
        run = self._run_with_findings(tmp_path)
        r = CliRunner().invoke(main, ["memo", "--report", str(run), "--json"])
        assert r.exit_code == 0, r.output
        out = json.loads(r.output)
        assert "go_no_go" in out
        assert "top_risks" in out
        assert "recommendations" in out
        assert out["total_findings"] == 1

    def test_memo_json_skips_markdown_html_writes(self, tmp_path: Path) -> None:
        run = self._run_with_findings(tmp_path)
        CliRunner().invoke(main, ["memo", "--report", str(run), "--json"])
        assert not (run / "report" / "ic_memo.md").exists()

    def test_memo_json_missing_findings_emits_error_json(self, tmp_path: Path) -> None:
        run = tmp_path / "run_empty"
        (run / "findings" / "merged").mkdir(parents=True)
        r = CliRunner().invoke(main, ["memo", "--report", str(run), "--json"])
        assert r.exit_code == 1
        assert "error" in json.loads(r.output)


class TestDiffCommand:
    def _write_run(self, run_dir: Path, subject: str, findings: list[dict]) -> None:  # type: ignore[type-arg]
        merged = run_dir / "findings" / "merged"
        merged.mkdir(parents=True, exist_ok=True)
        (merged / f"{subject}.json").write_text(
            json.dumps({"subject": subject, "subject_safe_name": subject, "findings": findings}),
        )
        (merged / "gaps").mkdir(exist_ok=True)

    def test_diff_json_reports_new_finding(self, tmp_path: Path) -> None:
        run_a = tmp_path / "run_a"  # current: has the new finding
        run_b = tmp_path / "run_b"  # prior: empty
        self._write_run(run_a, "acme", [{"category": "ip", "citations": [{"source_path": "a.pdf"}]}])
        self._write_run(run_b, "acme", [])

        r = CliRunner().invoke(main, ["diff", str(run_a), str(run_b), "--json"])
        assert r.exit_code == 0, r.output
        out = json.loads(r.output)
        assert out["summary"]["new_findings"] == 1
        assert out["summary"]["resolved_findings"] == 0
        assert out["current_run_id"] == "run_a"
        assert out["prior_run_id"] == "run_b"

    def test_diff_human_summary_shows_panel(self, tmp_path: Path) -> None:
        run_a = tmp_path / "run_a"
        run_b = tmp_path / "run_b"
        self._write_run(run_a, "acme", [])
        self._write_run(run_b, "acme", [])

        r = CliRunner().invoke(main, ["diff", str(run_a), str(run_b)])
        assert r.exit_code == 0, r.output
        assert "Run Diff" in r.output

    def test_diff_output_writes_standalone_html(self, tmp_path: Path) -> None:
        run_a = tmp_path / "run_a"
        run_b = tmp_path / "run_b"
        self._write_run(run_a, "acme", [{"category": "ip", "citations": [{"source_path": "a.pdf"}]}])
        self._write_run(run_b, "acme", [])
        out_html = tmp_path / "diff.html"

        r = CliRunner().invoke(main, ["diff", str(run_a), str(run_b), "--output", str(out_html)])
        assert r.exit_code == 0, r.output
        assert out_html.exists()
        html = out_html.read_text(encoding="utf-8")
        assert html.startswith("<!DOCTYPE html>")
        assert "Run-over-Run Changes" in html

    def test_diff_nonexistent_run_dir_errors(self, tmp_path: Path) -> None:
        run_a = tmp_path / "run_a"
        self._write_run(run_a, "acme", [])
        r = CliRunner().invoke(main, ["diff", str(run_a), str(tmp_path / "missing")])
        assert r.exit_code != 0
