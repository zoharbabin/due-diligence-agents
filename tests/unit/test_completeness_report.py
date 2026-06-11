"""Tests for surfacing request-list completeness + model-integrity in reports (Issue #238).

Covers the HTML CompletenessRenderer (request-list + formula audit, XSS-safe,
parity when absent) and the two Excel data handlers.
"""

from __future__ import annotations

from typing import Any

from dd_agents.reporting.computed_metrics import ReportDataComputer
from dd_agents.reporting.excel import ExcelReportGenerator
from dd_agents.reporting.html_completeness import CompletenessRenderer


def _computed() -> Any:
    return ReportDataComputer().compute({})


def _render(run_metadata: dict[str, Any] | None) -> str:
    config = {"_run_metadata": run_metadata}
    return CompletenessRenderer(_computed(), {}, config).render()


class TestCompletenessRendererParity:
    def test_empty_when_no_metadata(self) -> None:
        assert _render(None) == ""
        assert _render({}) == ""

    def test_empty_when_no_artifacts(self) -> None:
        assert _render({"llm_provider": "anthropic"}) == ""

    def test_formula_audit_with_no_formulas_renders_nothing(self) -> None:
        assert _render({"formula_audit": {"files_with_formulas": 0, "issues": []}}) == ""


class TestRequestListSection:
    def test_renders_received_and_missing(self) -> None:
        rl = {
            "summary": "2/3 expected items received, 1 required missing.",
            "received": ["Signed MSA", "Cap Table"],
            "missing_required": ["Audited Financials"],
            "missing_optional": [],
            "unexpected_count": 4,
        }
        html = _render({"request_list": rl})
        assert "Request-List Completeness" in html
        assert "Signed MSA" in html
        assert "Audited Financials" in html
        assert "id='sec-completeness'" in html
        # Unexpected count surfaced.
        assert "4" in html

    def test_request_list_escapes_items(self) -> None:
        rl = {
            "summary": "x",
            "received": ["<script>alert(1)</script>"],
            "missing_required": [],
            "missing_optional": [],
            "unexpected_count": 0,
        }
        html = _render({"request_list": rl})
        assert "<script>alert(1)</script>" not in html
        assert "&lt;script&gt;" in html


class TestModelIntegritySection:
    def test_renders_issues_citing_cells(self) -> None:
        fa = {
            "files_with_formulas": 2,
            "files_with_issues": 1,
            "total_issues": 2,
            "by_kind": {"hardcoded_override": 1, "circular_reference": 1},
            "issues": [
                {
                    "file": "Acme/model.xlsx",
                    "sheet": "P&L",
                    "cell": "B5",
                    "kind": "hardcoded_override",
                    "detail": "Hardcoded value 1234",
                },
                {
                    "file": "Acme/model.xlsx",
                    "sheet": "P&L",
                    "cell": "C3",
                    "kind": "circular_reference",
                    "detail": "Cell references itself",
                },
            ],
            "truncated": False,
        }
        html = _render({"formula_audit": fa})
        assert "Model Integrity (2)" in html
        assert "P&amp;L!B5" in html
        assert "Hardcoded override" in html

    def test_clean_audit_shows_good_alert(self) -> None:
        fa = {"files_with_formulas": 3, "files_with_issues": 0, "total_issues": 0, "by_kind": {}, "issues": []}
        html = _render({"formula_audit": fa})
        assert "No formula-integrity issues" in html

    def test_model_integrity_escapes_detail(self) -> None:
        fa = {
            "files_with_formulas": 1,
            "files_with_issues": 1,
            "total_issues": 1,
            "by_kind": {"error_literal": 1},
            "issues": [{"file": "<x>", "sheet": "S", "cell": "A1", "kind": "error_literal", "detail": "<img src=x>"}],
            "truncated": False,
        }
        html = _render({"formula_audit": fa})
        assert "<img src=x>" not in html
        assert "&lt;img" in html


class TestExcelHandlers:
    def test_request_list_handler_rows(self) -> None:
        gen = ExcelReportGenerator()
        rl = {"received": ["A", "B"], "missing_required": ["C"], "missing_optional": ["D"], "unexpected_count": 2}
        rows = gen._data_request_list({}, {}, {"request_list": rl})
        statuses = {r["status"]: r for r in rows}
        assert statuses["Received"]["count"] == 2
        assert statuses["Missing — required"]["count"] == 1
        assert statuses["Unexpected files"]["count"] == 2

    def test_request_list_handler_empty_when_absent(self) -> None:
        gen = ExcelReportGenerator()
        assert gen._data_request_list({}, {}, {}) == []

    def test_model_integrity_handler_rows(self) -> None:
        gen = ExcelReportGenerator()
        fa = {
            "files_with_formulas": 1,
            "issues": [{"file": "m.xlsx", "sheet": "S", "cell": "A1", "kind": "error_literal", "detail": "d"}],
        }
        rows = gen._data_model_integrity({}, {}, {"formula_audit": fa})
        assert len(rows) == 1
        assert rows[0]["cell"] == "A1"

    def test_model_integrity_handler_empty_when_no_formulas(self) -> None:
        gen = ExcelReportGenerator()
        assert gen._data_model_integrity({}, {}, {"formula_audit": {"files_with_formulas": 0}}) == []
        assert gen._data_model_integrity({}, {}, {}) == []


class TestEmptyRoomParity:
    """The 2 new sheets stay strictly headers-only for a generic room (HTML/Excel parity).

    Regression guard for the audit finding: Model_Integrity must NOT write a
    phantom TOTAL footer when there is no formula data (it would imply an audit
    ran), matching the HTML CompletenessRenderer which renders nothing.
    """

    def _generate_empty(self, tmp_path: Any) -> Any:
        import pathlib

        from dd_agents.models.reporting import ReportSchema

        repo_root = pathlib.Path(__file__).resolve().parents[2]
        schema = ReportSchema.model_validate_json((repo_root / "config" / "report_schema.json").read_text())
        out = tmp_path / "r.xlsx"
        ExcelReportGenerator().generate(
            {"s": {"subject": "S", "findings": [], "gaps": []}},
            schema,
            out,
            deal_config={},
            run_metadata={},
        )
        from openpyxl import load_workbook

        return load_workbook(out)

    def test_new_sheets_are_headers_only_when_empty(self, tmp_path: Any) -> None:
        wb = self._generate_empty(tmp_path)
        # max_row == 1 means headers only (no data, no summary footer).
        assert wb["Model_Integrity"].max_row == 1
        assert wb["Request_List_Completeness"].max_row == 1

    def test_html_renders_nothing_for_same_empty_room(self) -> None:
        # Parity counterpart: HTML section is absent for a generic room.
        assert CompletenessRenderer(_computed(), {}, {"_run_metadata": {}}).render() == ""
