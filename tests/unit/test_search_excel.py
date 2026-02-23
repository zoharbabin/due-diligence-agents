"""Unit tests for search Excel report generation."""

from __future__ import annotations

from typing import TYPE_CHECKING

import pytest
from openpyxl import load_workbook

from dd_agents.models.search import (
    SearchCitation,
    SearchColumn,
    SearchColumnResult,
    SearchCustomerResult,
    SearchPrompts,
)
from dd_agents.search.excel_writer import SearchExcelWriter

if TYPE_CHECKING:
    from pathlib import Path


def _make_prompts() -> SearchPrompts:
    return SearchPrompts(
        name="Test",
        columns=[
            SearchColumn(name="Q1", prompt="Question 1 -- long enough to pass validation."),
            SearchColumn(name="Q2", prompt="Question 2 -- long enough to pass validation."),
        ],
    )


def _make_results() -> list[SearchCustomerResult]:
    return [
        SearchCustomerResult(
            customer_name="Acme Corp",
            group="GroupA",
            files_analyzed=2,
            total_files=3,
            chunks_analyzed=1,
            skipped_files=["GroupA/Acme Corp/missing.pdf"],
            columns={
                "Q1": SearchColumnResult(
                    answer="YES",
                    confidence="HIGH",
                    citations=[
                        SearchCitation(
                            file_path="GroupA/Acme Corp/msa.pdf",
                            page="3",
                            section_ref="Section 12",
                            exact_quote="Relevant clause text here.",
                        ),
                    ],
                ),
                "Q2": SearchColumnResult(
                    answer="NOT_ADDRESSED",
                    confidence="HIGH",
                    citations=[],
                ),
            },
        ),
        SearchCustomerResult(
            customer_name="Globex Inc",
            group="GroupA",
            files_analyzed=1,
            total_files=1,
            chunks_analyzed=0,
            error="API error after 3 retries: timeout",
        ),
        SearchCustomerResult(
            customer_name="Partial Corp",
            group="GroupB",
            files_analyzed=2,
            total_files=2,
            chunks_analyzed=3,
            incomplete_columns=["Q2"],
            columns={
                "Q1": SearchColumnResult(
                    answer="NO",
                    confidence="HIGH",
                    citations=[],
                ),
                "Q2": SearchColumnResult(
                    answer="INCOMPLETE — not returned by model",
                    confidence="",
                    citations=[],
                ),
            },
            error="Incomplete response — missing columns: Q2",
        ),
    ]


class TestSearchExcelWriter:
    """Tests for SearchExcelWriter."""

    def test_generates_xlsx(self, tmp_path: Path) -> None:
        output = tmp_path / "report.xlsx"
        writer = SearchExcelWriter()
        result_path = writer.write(_make_results(), _make_prompts(), output)

        assert result_path.exists()
        assert str(result_path).endswith(".xlsx")

    def test_summary_sheet_structure(self, tmp_path: Path) -> None:
        output = tmp_path / "report.xlsx"
        writer = SearchExcelWriter()
        writer.write(_make_results(), _make_prompts(), output)

        wb = load_workbook(str(output))
        assert "Summary" in wb.sheetnames

        ws = wb["Summary"]
        # Header row: Customer, Group, Files Analyzed, Chunks, Files Skipped, Q1, Q2, Error
        headers = [ws.cell(row=1, column=c).value for c in range(1, 9)]
        assert headers == ["Customer", "Group", "Files Analyzed", "Chunks", "Files Skipped", "Q1", "Q2", "Error"]

        # 3 data rows (one per customer).
        assert ws.cell(row=2, column=1).value == "Acme Corp"
        assert ws.cell(row=3, column=1).value == "Globex Inc"
        assert ws.cell(row=4, column=1).value == "Partial Corp"

        # Answer value check (Q1 is now column 6).
        assert ws.cell(row=2, column=6).value == "YES"

        # Files analyzed shows ratio.
        assert ws.cell(row=2, column=3).value == "2/3"

    def test_details_sheet_has_citations(self, tmp_path: Path) -> None:
        output = tmp_path / "report.xlsx"
        writer = SearchExcelWriter()
        writer.write(_make_results(), _make_prompts(), output)

        wb = load_workbook(str(output))
        assert "Details" in wb.sheetnames

        ws = wb["Details"]
        # Header check.
        assert ws.cell(row=1, column=1).value == "Customer"
        assert ws.cell(row=1, column=9).value == "Exact Quote"

        # Header should say "Question", not "Column" (user-friendly).
        assert ws.cell(row=1, column=3).value == "Question"

        # Find a citation row for Acme Q1.
        found_citation = False
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == "Acme Corp" and ws.cell(row=row, column=3).value == "Q1":
                assert ws.cell(row=row, column=6).value == "GroupA/Acme Corp/msa.pdf"
                assert ws.cell(row=row, column=9).value == "Relevant clause text here."
                found_citation = True
                break
        assert found_citation

    def test_freeze_panes_applied(self, tmp_path: Path) -> None:
        output = tmp_path / "report.xlsx"
        writer = SearchExcelWriter()
        writer.write(_make_results(), _make_prompts(), output)

        wb = load_workbook(str(output))
        assert wb["Summary"].freeze_panes == "A2"
        assert wb["Details"].freeze_panes == "A2"

    def test_error_customers_in_both_sheets(self, tmp_path: Path) -> None:
        output = tmp_path / "report.xlsx"
        writer = SearchExcelWriter()
        writer.write(_make_results(), _make_prompts(), output)

        wb = load_workbook(str(output))

        # Summary: Globex row should have error.
        ws_summary = wb["Summary"]
        globex_error = ws_summary.cell(row=3, column=8).value
        assert globex_error is not None
        assert "API error" in globex_error

        # Details: Globex should appear with error.
        ws_details = wb["Details"]
        found_globex = False
        for row in range(2, ws_details.max_row + 1):
            if ws_details.cell(row=row, column=1).value == "Globex Inc":
                found_globex = True
                break
        assert found_globex

    def test_incomplete_columns_marked(self, tmp_path: Path) -> None:
        """Incomplete columns should be visually distinct from NOT_ADDRESSED."""
        output = tmp_path / "report.xlsx"
        writer = SearchExcelWriter()
        writer.write(_make_results(), _make_prompts(), output)

        wb = load_workbook(str(output))
        ws_summary = wb["Summary"]

        # Partial Corp is row 4, Q2 is column 7.
        q2_cell = ws_summary.cell(row=4, column=7)
        assert q2_cell.value == "INCOMPLETE"

    def test_chunks_column(self, tmp_path: Path) -> None:
        """Chunks column appears between Files Analyzed and Files Skipped."""
        output = tmp_path / "report.xlsx"
        writer = SearchExcelWriter()
        writer.write(_make_results(), _make_prompts(), output)

        wb = load_workbook(str(output))
        ws = wb["Summary"]

        # Header check.
        assert ws.cell(row=1, column=4).value == "Chunks"

        # Values: Acme=1, Globex=0, Partial=3.
        assert ws.cell(row=2, column=4).value == 1
        assert ws.cell(row=3, column=4).value == 0
        assert ws.cell(row=4, column=4).value == 3

        # Multi-chunk highlight: Partial Corp (3 chunks) should have fill.
        partial_fill = ws.cell(row=4, column=4).fill
        assert partial_fill.start_color.rgb is not None
        assert partial_fill.start_color.rgb != "00000000"  # Not default/no fill

        # Single-chunk (Acme, 1 chunk) should NOT have the highlight.
        acme_fill = ws.cell(row=2, column=4).fill
        assert acme_fill.start_color.rgb in (None, "00000000")

    def test_skipped_files_shown(self, tmp_path: Path) -> None:
        """Skipped files should be visible in the Summary sheet."""
        output = tmp_path / "report.xlsx"
        writer = SearchExcelWriter()
        writer.write(_make_results(), _make_prompts(), output)

        wb = load_workbook(str(output))
        ws_summary = wb["Summary"]

        # Acme Corp (row 2) has a skipped file.
        skipped_cell = ws_summary.cell(row=2, column=5)
        assert "missing.pdf" in str(skipped_cell.value)

    def test_summary_answer_normalization(self, tmp_path: Path) -> None:
        """Verbose LLM answers should be normalized to YES/NO/NOT_ADDRESSED in Summary."""
        results = [
            SearchCustomerResult(
                customer_name="Verbose Corp",
                group="GroupA",
                files_analyzed=1,
                total_files=1,
                chunks_analyzed=1,
                columns={
                    "Q1": SearchColumnResult(
                        answer="YES. Section 12.1 requires prior written consent for any change of control.",
                        confidence="HIGH",
                        citations=[],
                    ),
                    "Q2": SearchColumnResult(
                        answer="NO. The agreement does not contain a notice provision.",
                        confidence="HIGH",
                        citations=[],
                    ),
                },
            ),
        ]

        output = tmp_path / "report.xlsx"
        writer = SearchExcelWriter()
        writer.write(results, _make_prompts(), output)

        wb = load_workbook(str(output))
        ws = wb["Summary"]

        # Summary should show canonical YES/NO, not the full paragraph.
        assert ws.cell(row=2, column=6).value == "YES"
        assert ws.cell(row=2, column=7).value == "NO"

    def test_detail_preserves_full_answer(self, tmp_path: Path) -> None:
        """Details sheet preserves the full verbose answer (not normalized)."""
        verbose_answer = "YES. Section 12.1 requires prior written consent for any change of control."
        results = [
            SearchCustomerResult(
                customer_name="Verbose Corp",
                group="GroupA",
                files_analyzed=1,
                total_files=1,
                chunks_analyzed=1,
                columns={
                    "Q1": SearchColumnResult(
                        answer=verbose_answer,
                        confidence="HIGH",
                        citations=[],
                    ),
                    "Q2": SearchColumnResult(
                        answer="NOT_ADDRESSED",
                        confidence="HIGH",
                        citations=[],
                    ),
                },
            ),
        ]

        output = tmp_path / "report.xlsx"
        writer = SearchExcelWriter()
        writer.write(results, _make_prompts(), output)

        wb = load_workbook(str(output))
        ws = wb["Details"]

        # Find the Q1 row — should have the FULL answer, not normalized.
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=3).value == "Q1":
                assert ws.cell(row=row, column=4).value == verbose_answer
                break
        else:
            pytest.fail("Q1 row not found in Details sheet")

    def test_summary_not_addressed_normalization(self, tmp_path: Path) -> None:
        """Verbose NOT_ADDRESSED answers are normalized in Summary."""
        results = [
            SearchCustomerResult(
                customer_name="Partial Corp",
                group="GroupA",
                files_analyzed=1,
                total_files=1,
                chunks_analyzed=1,
                columns={
                    "Q1": SearchColumnResult(
                        answer="NOT_ADDRESSED. The reviewed portions do not contain relevant clauses.",
                        confidence="HIGH",
                        citations=[],
                    ),
                    "Q2": SearchColumnResult(
                        answer="YES",
                        confidence="HIGH",
                        citations=[],
                    ),
                },
            ),
        ]

        output = tmp_path / "report.xlsx"
        writer = SearchExcelWriter()
        writer.write(results, _make_prompts(), output)

        wb = load_workbook(str(output))
        ws = wb["Summary"]

        assert ws.cell(row=2, column=6).value == "NOT_ADDRESSED"
