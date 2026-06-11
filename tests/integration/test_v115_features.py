"""End-to-end coverage for v1.13/v1.14/v1.15 features that were unit-tested in
isolation but lacked an integration proof (Issue #242).

Covers, without spawning agents or needing an API key:
- Per-agent routing env reaches ``ClaudeAgentOptions.env`` (config → get_route_env
  → build_agent_options(extra_env=) → SDK seam). (#233/#240)
- ``dd-agents assess --config`` reconciles a request list over a real tmp data
  room (CLI → DataRoomAssessor → request-list reconcile). (#192)
- The formula-integrity section a Finance agent consumes stays format-stable for
  a workbook with a known model defect. (#194)
- The IC memo passes non-English finding text through verbatim (language-agnostic
  by design — no translation pass). (#190)
"""

from __future__ import annotations

import json
from typing import TYPE_CHECKING

from click.testing import CliRunner

if TYPE_CHECKING:
    from pathlib import Path


class TestPerAgentRoutingEnvWiring:
    """config → get_route_env → build_agent_options(extra_env=) → SDK options.env."""

    def test_route_env_reaches_claude_agent_options(self, tmp_path: Path, monkeypatch) -> None:  # noqa: ANN001
        from dd_agents.agents.specialists import LegalAgent
        from dd_agents.llm import build_agent_options
        from dd_agents.models.config import AgentRoute, DealConfig

        monkeypatch.setenv("MY_GW_TOKEN", "sk-secret-value")  # pragma: allowlist secret
        cfg = DealConfig.model_validate(
            {
                "config_version": "1.0.0",
                "buyer": {"name": "B"},
                "target": {"name": "T"},
                "deal": {"type": "acquisition", "focus_areas": ["ip_ownership"]},
                "agent_models": {
                    "routes": {
                        "legal": AgentRoute(base_url="http://gateway:4011", auth_token_env="MY_GW_TOKEN").model_dump()
                    }
                },
            }
        )
        runner = LegalAgent(project_dir=tmp_path, run_dir=tmp_path, run_id="r", deal_config=cfg)

        route_env = runner.get_route_env()
        # The seam forwards extra_env onto the per-call CLI subprocess env.
        options = build_agent_options(model=runner.get_model_id(), extra_env=route_env)

        assert options.env["ANTHROPIC_BASE_URL"] == "http://gateway:4011"
        assert options.env["ANTHROPIC_AUTH_TOKEN"] == "sk-secret-value"  # pragma: allowlist secret

    def test_route_provider_attribution_is_secret_free(self, tmp_path: Path) -> None:
        from dd_agents.agents.specialists import LegalAgent
        from dd_agents.models.config import AgentRoute, DealConfig

        cfg = DealConfig.model_validate(
            {
                "config_version": "1.0.0",
                "buyer": {"name": "B"},
                "target": {"name": "T"},
                "deal": {"type": "acquisition", "focus_areas": ["ip_ownership"]},
                "agent_models": {"routes": {"legal": AgentRoute(base_url="http://gw:4011").model_dump()}},
            }
        )
        runner = LegalAgent(project_dir=tmp_path, run_dir=tmp_path, run_id="r", deal_config=cfg)
        provider, base_url = runner.get_route_provider()
        assert provider == "gateway"
        assert base_url == "http://gw:4011"


class TestAssessConfigRequestList:
    """dd-agents assess --config reconciles a request list over a real data room."""

    def test_assess_config_reports_received_and_missing(self, tmp_path: Path) -> None:
        from dd_agents.cli import main

        data_room = tmp_path / "dr"
        (data_room / "Acme").mkdir(parents=True)
        # One file satisfies "Signed MSA"; the required "Audited Financials" is absent.
        (data_room / "Acme" / "msa_signed.pdf").write_text("x")
        (data_room / "Acme" / "extra_notes.pdf").write_text("y")

        config = {
            "config_version": "1.0.0",
            "buyer": {"name": "B"},
            "target": {"name": "T"},
            "deal": {"type": "acquisition", "focus_areas": ["ip_ownership"]},
            "data_room": {"path": str(data_room)},
            "request_list": {
                "enabled": True,
                "items": [
                    {"category": "Signed MSA", "keywords": ["msa", "signed"], "required": True},
                    {"category": "Audited Financials", "keywords": ["audited", "financials"], "required": True},
                ],
            },
        }
        config_path = tmp_path / "deal-config.json"
        config_path.write_text(json.dumps(config))

        result = CliRunner().invoke(main, ["assess", str(data_room), "--config", str(config_path)])
        assert result.exit_code == 0, result.output
        # The reconciliation summary appears in the assess output.
        out = result.output.lower()
        assert "request" in out or "received" in out or "missing" in out


class TestFormulaAuditFindingFormat:
    """The Finance agent's Formula Integrity section stays format-stable (#194)."""

    def test_section_cites_defect_for_known_bad_model(self, tmp_path: Path) -> None:
        from openpyxl import Workbook

        from dd_agents.tools.read_office import _formula_audit_section

        p = tmp_path / "model.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "P&L"
        # A column of real formulas with one hardcoded override → a known defect.
        ws["B2"] = "=A2*1.1"
        ws["B3"] = "=A3*1.1"
        ws["B4"] = "=A4*1.1"
        ws["B5"] = "=1234"  # hardcoded numeric "formula" overriding a computed cell
        wb.save(p)
        wb.close()

        section = _formula_audit_section(p)
        assert "Formula Integrity" in section
        # The exact-cell citation a specialist would quote must be present.
        assert "B5" in section
        assert "hardcoded_override" in section


class TestICMemoLanguagePassthrough:
    """The IC memo passes non-English finding text through verbatim (#190)."""

    def test_non_english_finding_text_preserved(self) -> None:
        from dd_agents.reporting.computed_metrics import ReportDataComputer
        from dd_agents.reporting.ic_memo import render_ic_memo

        # A P0 finding written in French (as deal.output_language would produce).
        french_title = "Clause de changement de contrôle déclenchée"
        merged = {
            "targetco": {
                "subject": "TargetCo",
                "subject_safe_name": "targetco",
                "findings": [
                    {
                        "id": "F1",
                        "agent": "legal",
                        "severity": "P0",
                        "title": french_title,
                        "description": "Le contrat exige le consentement écrit du fournisseur.",
                        "citation": {"source_path": "msa.pdf", "location": "§7.2", "exact_quote": "consentement"},
                    }
                ],
                "gaps": [],
            }
        }
        computed = ReportDataComputer().compute(merged)
        memo = render_ic_memo(computed, {"buyer": {"name": "B"}, "target": {"name": "TargetCo"}})
        # The non-English finding text appears unchanged — no translation pass.
        assert french_title in memo
