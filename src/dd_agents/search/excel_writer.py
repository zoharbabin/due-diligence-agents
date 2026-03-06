"""Excel report generation for search results (Summary + Details sheets)."""

from __future__ import annotations

import logging
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from pathlib import Path

    from openpyxl.cell.cell import Cell
    from openpyxl.worksheet.worksheet import Worksheet

    from dd_agents.models.search import SearchCitation, SearchCustomerResult, SearchPrompts

logger = logging.getLogger(__name__)

# Colour fills for answer cells.
_GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # YES
_BLUE_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")  # NO
_YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # NOT_ADDRESSED
_RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Error
_ORANGE_FILL = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")  # INCOMPLETE
_CHUNKS_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")  # Multi-chunk highlight
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_BODY_FONT = Font(size=10)
_WARNING_FONT = Font(size=10, italic=True, color="CC0000")


class SearchExcelWriter:
    """Write search results to an Excel workbook with Summary and Details sheets."""

    def write(
        self,
        results: list[SearchCustomerResult],
        prompts: SearchPrompts,
        output_path: Path,
    ) -> Path:
        """Generate the Excel report and save to *output_path*.

        Returns the resolved output path.
        """
        wb = Workbook()
        # Remove default sheet.
        default_ws = wb.active
        if default_ws is not None:
            wb.remove(default_ws)

        self._write_summary(wb, results, prompts)
        self._write_details(wb, results, prompts)

        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output_path))
        logger.info("Wrote search report to %s", output_path)
        return output_path

    # ------------------------------------------------------------------
    # Summary sheet
    # ------------------------------------------------------------------

    def _write_summary(
        self,
        wb: Workbook,
        results: list[SearchCustomerResult],
        prompts: SearchPrompts,
    ) -> None:
        ws: Worksheet = wb.create_sheet(title="Summary")

        # Build header row.
        headers = ["Customer", "Group", "Files Analyzed", "Chunks", "Files Skipped"]
        headers.extend(col.name for col in prompts.columns)
        headers.append("Error")

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT

        # Data rows.
        for row_idx, result in enumerate(results, start=2):
            ws.cell(row=row_idx, column=1, value=result.customer_name).font = _BODY_FONT
            ws.cell(row=row_idx, column=2, value=result.group).font = _BODY_FONT
            files_cell = ws.cell(
                row=row_idx,
                column=3,
                value=f"{result.files_analyzed}/{result.total_files}",
            )
            files_cell.font = _BODY_FONT
            if result.skipped_files:
                files_cell.fill = _ORANGE_FILL

            chunks_cell = ws.cell(row=row_idx, column=4, value=result.chunks_analyzed)
            chunks_cell.font = _BODY_FONT
            if result.chunks_analyzed > 1:
                chunks_cell.fill = _CHUNKS_FILL

            skipped_val = ", ".join(result.skipped_files) if result.skipped_files else ""
            skipped_cell = ws.cell(row=row_idx, column=5, value=skipped_val)
            skipped_cell.font = _BODY_FONT
            if result.skipped_files:
                skipped_cell.fill = _ORANGE_FILL

            for col_offset, search_col in enumerate(prompts.columns):
                col_idx = 6 + col_offset
                col_result = result.columns.get(search_col.name)
                if col_result is None:
                    # Column completely missing (no result at all).
                    cell = ws.cell(row=row_idx, column=col_idx, value="[NO DATA]")
                    cell.font = _WARNING_FONT
                    cell.fill = _RED_FILL
                elif search_col.name in result.incomplete_columns:
                    # Column was missing from Claude's response.
                    cell = ws.cell(row=row_idx, column=col_idx, value="INCOMPLETE")
                    cell.font = _WARNING_FONT
                    cell.fill = _ORANGE_FILL
                else:
                    answer = self._normalize_summary_answer(col_result.answer)
                    cell = ws.cell(row=row_idx, column=col_idx, value=answer)
                    cell.font = _BODY_FONT
                    self._apply_answer_fill(cell, answer)

            error_col = 6 + len(prompts.columns)
            error_cell = ws.cell(row=row_idx, column=error_col, value=result.error or "")
            error_cell.font = _BODY_FONT
            if result.error:
                error_cell.fill = _RED_FILL

        # Formatting.
        ws.freeze_panes = "A2"
        if ws.max_row and ws.max_row > 1:
            last_col = get_column_letter(ws.max_column)
            ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"

        # Column widths.
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 16
        ws.column_dimensions["D"].width = 10
        ws.column_dimensions["E"].width = 40
        for i in range(len(prompts.columns)):
            ws.column_dimensions[get_column_letter(6 + i)].width = 40
        ws.column_dimensions[get_column_letter(6 + len(prompts.columns))].width = 40

    # ------------------------------------------------------------------
    # Details sheet
    # ------------------------------------------------------------------

    def _write_details(
        self,
        wb: Workbook,
        results: list[SearchCustomerResult],
        prompts: SearchPrompts,
    ) -> None:
        ws: Worksheet = wb.create_sheet(title="Details")

        headers = [
            "Customer",
            "Group",
            "Question",
            "Answer",
            "Confidence",
            "File Path",
            "Page",
            "Section",
            "Exact Quote",
            "Quote Verified",
            "Match Score",
        ]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT

        row_idx = 2
        for result in results:
            for search_col in prompts.columns:
                col_result = result.columns.get(search_col.name)
                if col_result is None:
                    # Write a single row for error/missing customers.
                    ws.cell(row=row_idx, column=1, value=result.customer_name).font = _BODY_FONT
                    ws.cell(row=row_idx, column=2, value=result.group).font = _BODY_FONT
                    ws.cell(row=row_idx, column=3, value=search_col.name).font = _BODY_FONT
                    error_val = result.error or "[NO DATA — column missing from response]"
                    error_cell = ws.cell(row=row_idx, column=4, value=error_val)
                    error_cell.font = _WARNING_FONT
                    error_cell.fill = _RED_FILL
                    row_idx += 1
                    continue

                # Mark incomplete columns distinctly.
                is_incomplete = search_col.name in result.incomplete_columns

                citations = col_result.citations or []
                if not citations:
                    # Write answer row even without citations.
                    ws.cell(row=row_idx, column=1, value=result.customer_name).font = _BODY_FONT
                    ws.cell(row=row_idx, column=2, value=result.group).font = _BODY_FONT
                    ws.cell(row=row_idx, column=3, value=search_col.name).font = _BODY_FONT
                    answer_cell = ws.cell(row=row_idx, column=4, value=col_result.answer)
                    if is_incomplete:
                        answer_cell.font = _WARNING_FONT
                        answer_cell.fill = _ORANGE_FILL
                    else:
                        answer_cell.font = _BODY_FONT
                        self._apply_answer_fill(answer_cell, col_result.answer)
                    ws.cell(row=row_idx, column=5, value=col_result.confidence).font = _BODY_FONT
                    row_idx += 1
                else:
                    for cit in citations:
                        ws.cell(row=row_idx, column=1, value=result.customer_name).font = _BODY_FONT
                        ws.cell(row=row_idx, column=2, value=result.group).font = _BODY_FONT
                        ws.cell(row=row_idx, column=3, value=search_col.name).font = _BODY_FONT
                        answer_cell = ws.cell(row=row_idx, column=4, value=col_result.answer)
                        answer_cell.font = _BODY_FONT
                        self._apply_answer_fill(answer_cell, col_result.answer)
                        ws.cell(row=row_idx, column=5, value=col_result.confidence).font = _BODY_FONT
                        ws.cell(row=row_idx, column=6, value=cit.file_path).font = _BODY_FONT
                        ws.cell(row=row_idx, column=7, value=cit.page).font = _BODY_FONT
                        ws.cell(row=row_idx, column=8, value=cit.section_ref).font = _BODY_FONT
                        quote_cell = ws.cell(row=row_idx, column=9, value=cit.exact_quote)
                        quote_cell.font = _BODY_FONT
                        from copy import copy

                        aligned = copy(quote_cell.alignment)
                        aligned.wrapText = True
                        quote_cell.alignment = aligned
                        # Citation verification columns (Issue #5).
                        self._write_verification_cells(ws, row_idx, cit)
                        row_idx += 1

        # Formatting.
        ws.freeze_panes = "A2"
        if ws.max_row and ws.max_row > 1:
            last_col = get_column_letter(ws.max_column)
            ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"

        widths = [25, 15, 30, 40, 12, 40, 8, 15, 60, 14, 12]
        for i, w in enumerate(widths):
            ws.column_dimensions[get_column_letter(i + 1)].width = w

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _write_verification_cells(ws: Worksheet, row_idx: int, cit: SearchCitation) -> None:
        """Write citation verification columns (Issue #5)."""
        if cit.quote_verified is None:
            # Not verified (no verification data available).
            ws.cell(row=row_idx, column=10, value="").font = _BODY_FONT
            ws.cell(row=row_idx, column=11, value="").font = _BODY_FONT
        elif cit.quote_verified:
            verified_cell = ws.cell(row=row_idx, column=10, value="YES")
            verified_cell.font = _BODY_FONT
            verified_cell.fill = _GREEN_FILL
            score_cell = ws.cell(row=row_idx, column=11, value=round(cit.quote_match_score))
            score_cell.font = _BODY_FONT
        else:
            verified_cell = ws.cell(row=row_idx, column=10, value="NO")
            verified_cell.font = _WARNING_FONT
            verified_cell.fill = _ORANGE_FILL
            score_cell = ws.cell(row=row_idx, column=11, value=round(cit.quote_match_score))
            score_cell.font = _WARNING_FONT

    @staticmethod
    def _apply_answer_fill(cell: Cell, answer: str) -> None:
        """Apply conditional colour fill based on answer text.

        Uses starts-with matching so that verbose LLM answers like
        ``"YES. Section 12 requires..."`` still receive correct fill.
        NOT_ADDRESSED is checked first because it starts with "NO".
        """
        upper = answer.strip().upper()
        if upper.startswith("NOT_ADDRESSED") or upper.startswith("NOT ADDRESSED"):
            cell.fill = _YELLOW_FILL
        elif upper.startswith("YES"):
            cell.fill = _GREEN_FILL
        elif upper.startswith("NO"):
            cell.fill = _BLUE_FILL

    @staticmethod
    def _normalize_summary_answer(answer: str) -> str:
        """Normalize verbose answers to canonical YES/NO/NOT_ADDRESSED for Summary.

        LLMs often return ``"YES. Section 12 requires..."`` or multi-sentence
        explanations.  The Summary sheet shows only the canonical verdict;
        the full text is preserved in the Details sheet.
        """
        upper = answer.strip().upper()
        if upper.startswith("NOT_ADDRESSED") or upper.startswith("NOT ADDRESSED"):
            return "NOT_ADDRESSED"
        if upper.startswith("YES"):
            return "YES"
        if upper.startswith("NO"):
            return "NO"
        # Free-text that doesn't start with a canonical keyword — truncate.
        if len(answer) > 200:
            return answer[:197] + "..."
        return answer
