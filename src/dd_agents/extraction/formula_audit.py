"""Excel formula extraction + model-integrity audit (Issue #194).

The bulk read path (`tools/read_office.py`) loads workbooks with
``data_only=True`` — it sees computed *values* but not the *formulas* that
produced them. In M&A, financial-model errors materially move valuation:
hardcoded overrides of formula cells, circular references, broken external
links, inconsistent row formulas. This module adds a second, formula-aware
pass and a pure detector so the Finance specialist can flag model-integrity
issues citing an exact cell.

Pure + dependency-light: uses openpyxl (already a core dependency) with
``data_only=False`` and ``read_only=True`` (bounded memory). The detector
functions take a plain formula map and return plain dataclasses, so they are
unit-testable without a workbook.

Contract:
- :func:`extract_formula_map` — workbook path → ``{sheet: {cell: formula}}``.
- :func:`audit_formulas` — formula map → list[:class:`FormulaIssue`].
- :func:`format_formula_audit` — render a compact, delimited markdown section
  for the agent prompt (empty string when nothing notable).
"""

from __future__ import annotations

import re
from collections import defaultdict
from dataclasses import dataclass
from typing import TYPE_CHECKING, TypedDict

if TYPE_CHECKING:
    from pathlib import Path


class FormulaAuditReport(TypedDict):
    """Serializable data-room formula-audit report (Issue #238)."""

    files_scanned: int
    files_with_formulas: int
    files_with_issues: int
    total_issues: int
    by_kind: dict[str, int]
    issues: list[dict[str, str]]
    truncated: bool


# A formula map: sheet name -> {cell coordinate (e.g. "B7") -> formula text incl. leading "="}.
FormulaMap = dict[str, dict[str, str]]

# Cap how many cells we scan per sheet so a pathological workbook can't blow up
# memory/time. Far above any real financial model's formula density.
_MAX_CELLS_PER_SHEET: int = 50_000

# Excel error literals that indicate a broken model.
_ERROR_LITERALS: tuple[str, ...] = ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A", "#NULL!", "#NUM!")

# A column letter + row number, e.g. "AB12" → ("AB", 12).
_CELL_RE = re.compile(r"^([A-Z]+)(\d+)$")
# External-workbook link markers in a formula, e.g. =[1]Sheet1!A1 or ='[Book.xlsx]Sheet'!A1.
_EXTERNAL_LINK_RE = re.compile(r"\[\d+\]|\[[^\]]+\.xlsx?\]", re.IGNORECASE)
# A bare numeric literal (int/float, optional sign), used to spot hardcoded overrides.
_NUMERIC_LITERAL_RE = re.compile(r"^[+-]?\d+(\.\d+)?$")


@dataclass(frozen=True)
class FormulaIssue:
    """One detected model-integrity issue, citable to an exact cell.

    Attributes
    ----------
    kind:
        Stable machine key: ``hardcoded_override`` | ``error_literal`` |
        ``broken_external_link`` | ``circular_reference``.
    sheet, cell:
        Location for the citation (e.g. ``"P&L"``, ``"C12"``).
    detail:
        Human-readable explanation including the offending formula/value.
    """

    kind: str
    sheet: str
    cell: str
    detail: str


def _split_cell(coord: str) -> tuple[str, int] | None:
    """Split ``"AB12"`` → ``("AB", 12)``; None if not a simple cell ref."""
    m = _CELL_RE.match(coord)
    if not m:
        return None
    return m.group(1), int(m.group(2))


def _col_to_num(col: str) -> int:
    """Convert an Excel column letter to its 1-based number (A=1, Z=26, AA=27).

    Used for sorting so multi-letter columns order correctly (B < Z < AA),
    rather than lexicographically (where "AA" would wrongly sort before "B").
    """
    n = 0
    for ch in col:
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n


def _sort_key(cell: str) -> tuple[int, int]:
    """Stable (column-number, row) sort key for a cell coordinate."""
    parts = _split_cell(cell)
    if parts is None:
        return (0, 0)
    return (_col_to_num(parts[0]), parts[1])


def extract_formula_map(path: Path, max_cells_per_sheet: int = _MAX_CELLS_PER_SHEET) -> FormulaMap:
    """Return ``{sheet: {cell: formula}}`` for every formula cell in *path*.

    Loads with ``data_only=False`` (formula text) and ``read_only=True``
    (bounded memory). Non-formula cells are skipped. Best-effort: returns an
    empty map for a non-.xlsx file or an unreadable workbook (the caller's
    value pass still runs).
    """
    if path.suffix.lower() != ".xlsx":
        return {}

    from openpyxl import load_workbook

    try:
        wb = load_workbook(path, read_only=True, data_only=False)
    except Exception:  # noqa: BLE001 — formula audit is best-effort; never block the value read
        return {}

    formula_map: FormulaMap = {}
    try:
        for name in wb.sheetnames:
            ws = wb[name]
            sheet_formulas: dict[str, str] = {}
            seen = 0
            for cell_row in ws.iter_rows(values_only=False):
                for cell in cell_row:
                    seen += 1
                    if seen > max_cells_per_sheet:
                        break
                    value = getattr(cell, "value", None)
                    # In data_only=False mode, a formula cell's value is its
                    # formula string (starts with "="). openpyxl read-only also
                    # exposes data_type == "f" for formula cells.
                    if isinstance(value, str) and value.startswith("="):
                        coord = getattr(cell, "coordinate", None)
                        if coord:
                            sheet_formulas[coord] = value
                if seen > max_cells_per_sheet:
                    break
            if sheet_formulas:
                formula_map[name] = sheet_formulas
    finally:
        wb.close()

    return formula_map


def audit_formulas(formula_map: FormulaMap) -> list[FormulaIssue]:
    """Detect model-integrity issues from a formula map. Pure + deterministic.

    Detects:
    - ``error_literal`` — a formula referencing an Excel error (``#REF!`` etc.),
      i.e. a broken reference or division surfaced in the model.
    - ``broken_external_link`` — a formula linking to an external workbook
      (``[1]Sheet!A1`` / ``[Book.xlsx]``), which silently breaks on hand-off.
    - ``circular_reference`` — a formula that references its own cell coordinate
      (the circular pattern detectable from formula text alone).
    - ``hardcoded_override`` — a cell holding a bare numeric literal as its
      "formula" (``=42``) inside a column whose other cells are real formulas,
      i.e. an analyst overrode a computed cell with a typed-in number.

    Returns issues sorted by (sheet, column, row) — columns ordered
    numerically (B < Z < AA) — for stable, citable output.
    """
    issues: list[FormulaIssue] = []

    for sheet, cells in formula_map.items():
        # Group formula cells by column to spot a hardcoded literal among formulas.
        by_column: dict[str, list[tuple[int, str, str]]] = defaultdict(list)
        for coord, formula in cells.items():
            parts = _split_cell(coord)
            if parts is not None:
                col, row = parts
                by_column[col].append((row, coord, formula))

            # Error literals + external links + self-reference are per-cell.
            upper = formula.upper()
            for err in _ERROR_LITERALS:
                if err in upper:
                    issues.append(
                        FormulaIssue(
                            kind="error_literal",
                            sheet=sheet,
                            cell=coord,
                            detail=f"Formula references {err}: {formula}",
                        )
                    )
                    break
            if _EXTERNAL_LINK_RE.search(formula):
                issues.append(
                    FormulaIssue(
                        kind="broken_external_link",
                        sheet=sheet,
                        cell=coord,
                        detail=f"Formula links to an external workbook (breaks on hand-off): {formula}",
                    )
                )
            # Circular: the cell's own coordinate appears in its formula. Word
            # boundary avoids matching A1 inside A12 (different rows/cols).
            if re.search(rf"(?<![A-Z0-9]){re.escape(coord)}(?![0-9])", upper):
                issues.append(
                    FormulaIssue(
                        kind="circular_reference",
                        sheet=sheet,
                        cell=coord,
                        detail=f"Cell references itself (circular): {formula}",
                    )
                )

        # Hardcoded-override: in a column with >=3 formula cells, a cell whose
        # "formula" is just a numeric literal is a typed-in override of a
        # computed cell. The >=3 floor avoids flagging genuinely constant columns.
        for col, entries in by_column.items():
            if len(entries) < 3:
                continue
            literals = [
                (row, coord, f) for (row, coord, f) in entries if _NUMERIC_LITERAL_RE.match(f.lstrip("=").strip())
            ]
            real_formulas = len(entries) - len(literals)
            # Only flag when the column is predominantly real formulas (so the
            # literal stands out as an override, not a constant column).
            if real_formulas >= 2 and literals:
                for _row, coord, f in literals:
                    literal = f.lstrip("=").strip()
                    issues.append(
                        FormulaIssue(
                            kind="hardcoded_override",
                            sheet=sheet,
                            cell=coord,
                            detail=f"Hardcoded value {literal} in column {col} of otherwise-computed cells",
                        )
                    )

    issues.sort(key=lambda i: (i.sheet, _sort_key(i.cell)))
    return issues


def audit_data_room(
    paths: list[Path],
    *,
    max_files: int = 200,
    max_issues: int = 500,
) -> FormulaAuditReport:
    """Audit every ``.xlsx`` in *paths* and return a serializable report (Issue #238).

    Reuses the pure per-workbook detector (:func:`extract_formula_map` +
    :func:`audit_formulas`) so the report and the agent prompt share one audit.
    The result is JSON-safe (persisted to ``formula_audit.json`` and surfaced in
    the HTML/Excel report). Best-effort and parity-safe: a data room with no
    spreadsheet formulas yields ``files_with_formulas == 0`` and an empty
    ``issues`` list, so nothing renders. Bounded by *max_files* / *max_issues*;
    any overflow is reported in ``truncated`` so silence never masks a cap.

    Each issue row carries ``file`` (the path as given), ``sheet``, ``cell``,
    ``kind``, and ``detail`` — enough to cite ``file → Sheet!Cell`` in the report.
    """
    xlsx = [p for p in paths if p.suffix.lower() == ".xlsx"]
    files_truncated = len(xlsx) > max_files
    xlsx = xlsx[:max_files]

    issue_rows: list[dict[str, str]] = []
    by_kind: dict[str, int] = defaultdict(int)
    files_with_formulas = 0
    files_with_issues = 0

    for path in xlsx:
        formula_map = extract_formula_map(path)
        if not formula_map:
            continue
        files_with_formulas += 1
        issues = audit_formulas(formula_map)
        if issues:
            files_with_issues += 1
        for issue in issues:
            issue_rows.append(
                {
                    "file": str(path),
                    "sheet": issue.sheet,
                    "cell": issue.cell,
                    "kind": issue.kind,
                    "detail": issue.detail,
                }
            )
            by_kind[issue.kind] += 1

    issues_truncated = len(issue_rows) > max_issues
    issue_rows = issue_rows[:max_issues]

    return FormulaAuditReport(
        files_scanned=len(xlsx),
        files_with_formulas=files_with_formulas,
        files_with_issues=files_with_issues,
        total_issues=sum(by_kind.values()),
        by_kind=dict(sorted(by_kind.items())),
        issues=issue_rows,
        truncated=files_truncated or issues_truncated,
    )


def format_formula_audit(formula_map: FormulaMap, issues: list[FormulaIssue], *, max_issues: int = 40) -> str:
    """Render a compact markdown section for the agent prompt.

    Empty string when there are no formulas (so non-model spreadsheets add
    nothing). Lists detected integrity issues first (citable to a cell), then
    a small per-sheet formula count for context. Capped to keep the prompt
    bounded.
    """
    if not formula_map:
        return ""

    lines: list[str] = ["", "## Formula Integrity (model audit)"]
    if issues:
        lines.append("")
        lines.append(f"Detected {len(issues)} potential model-integrity issue(s) — cite the exact cell:")
        for issue in issues[:max_issues]:
            lines.append(f"- [{issue.kind}] {issue.sheet}!{issue.cell}: {issue.detail}")
        if len(issues) > max_issues:
            lines.append(f"- … and {len(issues) - max_issues} more (truncated)")
    else:
        lines.append("")
        lines.append("No formula-integrity issues detected by the deterministic audit.")

    counts = ", ".join(f"{sheet} ({len(cells)})" for sheet, cells in sorted(formula_map.items()))
    lines.append("")
    lines.append(f"Formula cell counts by sheet: {counts}")
    return "\n".join(lines)
