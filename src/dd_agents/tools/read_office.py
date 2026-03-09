"""read_office MCP tool.

Reads binary Office files (.xlsx, .xls, .docx, .doc, .pptx, .ppt) and
returns structured text content.  Uses openpyxl for Excel and markitdown
for Word/PowerPoint.  Falls back to pre-extracted markdown in index/text/
when primary reading fails.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

_MAX_OUTPUT_CHARS: int = 150_000


class _SheetNotFoundError(Exception):
    """Raised when a requested Excel sheet does not exist."""


# openpyxl only supports .xlsx (OOXML), not legacy .xls (BIFF).
# Legacy .xls/.doc/.ppt are routed through markitdown which handles them.
_OPENPYXL_EXTENSIONS: frozenset[str] = frozenset({".xlsx"})
_MARKITDOWN_EXTENSIONS: frozenset[str] = frozenset({".xls", ".docx", ".doc", ".pptx", ".ppt"})
_ALL_OFFICE_EXTENSIONS: frozenset[str] = _OPENPYXL_EXTENSIONS | _MARKITDOWN_EXTENSIONS


def read_office(
    file_path: str,
    sheet_name: str | None = None,
    text_dir: str | None = None,
) -> dict[str, Any]:
    """Read a binary Office file and return its content as structured text.

    Args:
        file_path: Path to the Office file.
        sheet_name: For Excel files — specific sheet to read.  ``None`` reads all.
        text_dir: Optional path to the extracted text directory (``index/text/``).
            Used as fallback when primary reading fails.

    Returns:
        ``{"status": "ok", "content": "...", "method": "..."}`` on success, or
        ``{"status": "error", "reason": "..."}`` on failure.
    """
    path = Path(file_path)

    if not path.exists():
        return {"status": "error", "reason": f"File not found: {file_path}"}

    suffix = path.suffix.lower()
    if suffix not in _ALL_OFFICE_EXTENSIONS:
        return {
            "status": "error",
            "reason": f"Unsupported file type '{suffix}'. "
            f"read_office handles: {', '.join(sorted(_ALL_OFFICE_EXTENSIONS))}",
        }

    # Primary read attempt
    try:
        if suffix in _OPENPYXL_EXTENSIONS:
            content = _read_excel(path, sheet_name)
            method = "openpyxl"
        else:
            if sheet_name:
                logger.debug("sheet_name='%s' ignored for non-.xlsx file %s", sheet_name, path.name)
            content = _read_with_markitdown(path)
            method = "markitdown"

        content = _truncate(content)
        return {"status": "ok", "content": content, "method": method}

    except _SheetNotFoundError as exc:
        # Deterministic errors (invalid sheet name) — return immediately, no fallback
        return {"status": "error", "reason": str(exc)}

    except Exception as exc:
        logger.debug("Primary read failed for %s: %s", file_path, exc)

    # Fallback: pre-extracted text from index/text/
    fallback = _try_extracted_fallback(file_path, text_dir)
    if fallback is not None:
        return {
            "status": "ok",
            "content": _truncate(fallback),
            "method": "extracted_text_fallback",
        }

    return {
        "status": "error",
        "reason": f"Could not read '{path.name}'. The file may be corrupted or password-protected.",
    }


def _read_excel(path: Path, sheet_name: str | None) -> str:
    """Read an Excel file using openpyxl and return markdown tables.

    Streams rows directly into the output to avoid loading the entire
    spreadsheet into memory.  Stops reading once the output exceeds
    ``_MAX_OUTPUT_CHARS`` (truncation is applied by the caller).
    """
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                raise _SheetNotFoundError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
            sheet_names = [sheet_name]
        else:
            sheet_names = list(wb.sheetnames)

        parts: list[str] = []
        total_chars = 0
        budget_exceeded = False

        for name in sheet_names:
            if budget_exceeded:
                break

            ws = wb[name]

            # First pass: read rows to determine dimensions.
            # Use a capped read to avoid unbounded memory on huge sheets.
            rows: list[list[str]] = []
            for raw_row in ws.iter_rows(values_only=True):
                rows.append([_sanitize_cell(str(cell)) if cell is not None else "" for cell in raw_row])
                # Check periodically — every 500 rows — whether we've read enough
                if len(rows) % 500 == 0 and len(rows) > 0:
                    est_chars = sum(sum(len(c) + 3 for c in r) for r in rows)
                    if est_chars > _MAX_OUTPUT_CHARS * 2:
                        break

            num_rows = len(rows)
            num_cols = max((len(r) for r in rows), default=0)
            parts.append(f"## Sheet: {name} ({num_rows} rows, {num_cols} columns)\n")
            total_chars += len(parts[-1])

            if num_rows == 0:
                parts.append("(empty sheet)\n")
                continue

            # Generate column headers (A, B, C, ...) so real data isn't
            # misinterpreted as headers when row 1 contains data values.
            col_headers = [_col_letter(i) for i in range(num_cols)]
            parts.append("| " + " | ".join(col_headers) + " |")
            parts.append("| " + " | ".join("---" for _ in col_headers) + " |")
            for row in rows:
                padded = list(row)
                while len(padded) < num_cols:
                    padded.append("")
                line = "| " + " | ".join(padded) + " |"
                parts.append(line)
                total_chars += len(line) + 1  # +1 for \n join
                if total_chars > _MAX_OUTPUT_CHARS:
                    budget_exceeded = True
                    break
            parts.append("")

        return "\n".join(parts)
    finally:
        wb.close()


def _sanitize_cell(value: str) -> str:
    """Sanitize a cell value for safe embedding in a markdown table row.

    - Replaces newlines with spaces (newlines break table row boundaries)
    - Escapes pipe characters (pipes break column boundaries)
    """
    return value.replace("\r\n", " ").replace("\r", " ").replace("\n", " ").replace("|", "\\|")


def _col_letter(index: int) -> str:
    """Convert 0-based column index to Excel-style letter (A, B, ..., Z, AA, ...)."""
    result = ""
    i = index
    while True:
        result = chr(65 + i % 26) + result
        i = i // 26 - 1
        if i < 0:
            break
    return result


def _read_with_markitdown(path: Path) -> str:
    """Read Word/PowerPoint using markitdown."""
    from markitdown import MarkItDown

    result = MarkItDown().convert(str(path))
    text: str = result.text_content
    if not text or not text.strip():
        raise ValueError(f"markitdown returned empty content for {path.name}")
    return text


def _try_extracted_fallback(
    file_path: str,
    text_dir: str | None,
) -> str | None:
    """Try to read pre-extracted markdown from the text index directory."""
    if not text_dir:
        return None

    text_dir_path = Path(text_dir)
    if not text_dir_path.is_dir():
        return None

    # Try full-path convention first (matches extraction pipeline)
    from dd_agents.extraction.pipeline import ExtractionPipeline

    safe_name = ExtractionPipeline._safe_text_name(file_path)
    extracted = text_dir_path / safe_name
    if extracted.exists():
        return extracted.read_text(encoding="utf-8")

    # Try filename-only (handles absolute paths passed at runtime)
    basename = Path(file_path).name
    simple = text_dir_path / f"{basename}.md"
    if simple.exists():
        return simple.read_text(encoding="utf-8")

    return None


def _truncate(content: str) -> str:
    """Truncate content to _MAX_OUTPUT_CHARS with a notice."""
    if len(content) <= _MAX_OUTPUT_CHARS:
        return content
    return content[:_MAX_OUTPUT_CHARS] + "\n\n[... output truncated at 150K characters]"
