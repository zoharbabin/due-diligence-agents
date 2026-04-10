"""6-layer numerical audit.

Implements the blocking gate between analysis completion and Excel
generation. Every number in the pipeline must be traceable, re-derivable,
cross-source consistent, format-consistent, semantically reasonable,
and -- for financial citations -- present in the referenced source documents.
"""

from __future__ import annotations

import json
import logging
import random
import re
from pathlib import Path
from typing import TYPE_CHECKING, Any

from dd_agents.models.audit import AuditCheck

if TYPE_CHECKING:
    from dd_agents.models.numerical import ManifestEntry, NumericalManifest

logger = logging.getLogger(__name__)


def _manifest_get(manifest: NumericalManifest, entry_id: str) -> ManifestEntry | None:
    """Retrieve a manifest entry by ID."""
    for n in manifest.numbers:
        if n.id == entry_id:
            return n
    return None


class NumericalAuditor:
    """6-layer numerical auditor -- BLOCKING gate before Excel generation.

    Layers
    ------
    1. Source traceability  -- every number traces to a source file.
    2. Arithmetic           -- re-derive values from source.
    3. Cross-source         -- ``subjects.csv`` vs ``counts.json`` vs findings.
    4. Cross-format parity  -- spot-check Excel vs JSON (post-generation only).
    5. Semantic              -- flag implausible values.
    6. Financial citation   -- dollar amounts in P0/P1 findings match source docs.
    """

    def __init__(
        self,
        run_dir: Path,
        inventory_dir: Path | None = None,
        prior_manifest: NumericalManifest | None = None,
    ) -> None:
        self.run_dir = run_dir
        self.inventory_dir = inventory_dir or run_dir
        self.prior_manifest = prior_manifest
        self._findings_cache: list[dict[str, Any]] | None = None

    # ------------------------------------------------------------------ #
    # full audit
    # ------------------------------------------------------------------ #

    def run_full_audit(self, manifest: NumericalManifest, *, text_dir: Path | None = None) -> list[AuditCheck]:
        """Run all 6 layers and return a list of AuditCheck results.

        Layers 1-3 and 5 run pre-generation.
        Layer 4 is skipped when no Excel path is provided.
        Layer 6 runs when *text_dir* is provided; non-blocking when absent.
        """
        checks: list[AuditCheck] = [
            self.check_source_traceability(manifest),
            self.check_arithmetic(manifest),
            self.check_cross_source_consistency(manifest),
            self.check_semantic_reasonableness(manifest),
        ]
        # Layer 6: financial citation verification (non-blocking when text_dir unavailable).
        checks.append(self.check_financial_citations(text_dir=text_dir))
        return checks

    # ------------------------------------------------------------------ #
    # Layer 1 -- Source Traceability
    # ------------------------------------------------------------------ #

    def check_source_traceability(self, manifest: NumericalManifest) -> AuditCheck:
        """Every number must trace to a specific file and derivation method."""
        failures: list[str] = []
        for entry in manifest.numbers:
            resolved = self._resolve_path(entry.source_file)
            if not self._source_exists(resolved):
                # A glob pattern matching 0 files is valid when the manifest
                # value is also 0 (e.g. zero gaps → no gap JSON files).
                if "*" in entry.source_file and entry.value == 0:
                    logger.debug(
                        "%s (%s): glob '%s' matched 0 files but value is 0 -- accepted",
                        entry.id,
                        entry.label,
                        entry.source_file,
                    )
                else:
                    failures.append(f"{entry.id} ({entry.label}): source_file '{entry.source_file}' does not exist")
            if not entry.derivation:
                failures.append(f"{entry.id} ({entry.label}): missing derivation")
        return AuditCheck(
            passed=len(failures) == 0,
            dod_checks=[17],
            details={"layer": 1, "failures": failures},
            rule="Layer 1: every number maps to a file that exists.",
        )

    # ------------------------------------------------------------------ #
    # Layer 2 -- Arithmetic Verification
    # ------------------------------------------------------------------ #

    def check_arithmetic(self, manifest: NumericalManifest) -> AuditCheck:
        """Re-derive every number from source.

        For known IDs (N001-N010) this uses deterministic rederivation.
        For unknown IDs, the value is accepted as-is (verification flag set).
        """
        failures: list[str] = []
        for entry in manifest.numbers:
            rederived = self._rederive(entry)
            if rederived is not None and rederived != entry.value:
                failures.append(f"{entry.id} ({entry.label}): manifest={entry.value}, rederived={rederived}")
            entry.verified = True
        return AuditCheck(
            passed=len(failures) == 0,
            dod_checks=[17],
            details={"layer": 2, "failures": failures},
            rule="Layer 2: re-derive every number from source.",
        )

    # ------------------------------------------------------------------ #
    # Layer 3 -- Cross-Source Consistency
    # ------------------------------------------------------------------ #

    def check_cross_source_consistency(self, manifest: NumericalManifest) -> AuditCheck:
        """Numbers that appear in multiple sources must agree."""
        failures: list[str] = []

        # Cross-check: subjects.csv row count == counts.json total_subjects
        csv_count = self._count_csv_rows("subjects.csv")
        counts_total = self._read_counts_json_field("total_subjects")
        n001 = _manifest_get(manifest, "N001")

        if csv_count is not None and counts_total is not None and csv_count != counts_total:
            failures.append(f"subjects.csv rows ({csv_count}) != counts.json total_subjects ({counts_total})")

        if csv_count is not None and n001 is not None and csv_count != n001.value:
            failures.append(f"subjects.csv rows ({csv_count}) != manifest N001 ({n001.value})")

        # Cross-check: severity sum == total findings
        n003 = _manifest_get(manifest, "N003")
        n004 = _manifest_get(manifest, "N004")
        n005 = _manifest_get(manifest, "N005")
        n006 = _manifest_get(manifest, "N006")
        n007 = _manifest_get(manifest, "N007")

        if n003 is not None and n004 is not None and n005 is not None and n006 is not None and n007 is not None:
            severity_sum = n004.value + n005.value + n006.value + n007.value
            if severity_sum != n003.value:
                failures.append(f"N004+N005+N006+N007 ({severity_sum}) != N003 ({n003.value})")

        # Ratio consistency: each severity count / total must be in [0, 1]
        if n003 is not None and n003.value > 0:
            for nid, entry in [("N004", n004), ("N005", n005), ("N006", n006), ("N007", n007)]:
                if entry is not None:
                    ratio = entry.value / n003.value
                    if not (0.0 <= ratio <= 1.0):
                        failures.append(f"{nid}/N003 ratio {ratio:.4f} is outside [0, 1]")

        return AuditCheck(
            passed=len(failures) == 0,
            dod_checks=[17],
            details={"layer": 3, "failures": failures},
            rule="Layer 3: numbers across files must agree.",
        )

    # ------------------------------------------------------------------ #
    # Layer 4 -- Cross-Format Parity (post-generation)
    # ------------------------------------------------------------------ #

    def check_cross_format_parity(
        self,
        excel_path: Path,
        manifest: NumericalManifest,
        sample_count: int = 3,
    ) -> AuditCheck:
        """Spot-check that Excel cell values match manifest values.

        Parameters
        ----------
        excel_path:
            Path to the generated Excel workbook.
        manifest:
            The numerical manifest to verify against.
        sample_count:
            Number of manifest entries to spot-check (default 3).
        """
        failures: list[str] = []

        if not excel_path.exists():
            return AuditCheck(
                passed=False,
                dod_checks=[17],
                details={"layer": 4, "failures": ["Excel file does not exist"]},
                rule="Layer 4: Excel cells match manifest values.",
            )

        try:
            import openpyxl

            wb = openpyxl.load_workbook(excel_path, data_only=True)
        except Exception as exc:
            return AuditCheck(
                passed=False,
                dod_checks=[17],
                details={"layer": 4, "failures": [f"Cannot open Excel: {exc}"]},
                rule="Layer 4: Excel cells match manifest values.",
            )

        # Spot-check random entries
        entries = manifest.numbers
        sample = random.sample(entries, min(sample_count, len(entries)))
        for entry in sample:
            # Check if value appears in any sheet
            found = self._find_value_in_workbook(wb, entry)
            if not found:
                failures.append(f"{entry.id} ({entry.label}): value {entry.value} not found in Excel workbook")

        return AuditCheck(
            passed=len(failures) == 0,
            dod_checks=[17],
            details={
                "layer": 4,
                "samples_checked": len(sample),
                "failures": failures,
            },
            rule="Layer 4: Excel cells match manifest values.",
        )

    # ------------------------------------------------------------------ #
    # Layer 5 -- Semantic Reasonableness
    # ------------------------------------------------------------------ #

    def check_semantic_reasonableness(self, manifest: NumericalManifest) -> AuditCheck:
        """Flag numbers that are implausible."""
        failures: list[str] = []

        # No negative values
        for entry in manifest.numbers:
            if isinstance(entry.value, (int, float)) and entry.value < 0:
                failures.append(f"{entry.id} ({entry.label}): negative value {entry.value}")

        n003 = _manifest_get(manifest, "N003")
        n004 = _manifest_get(manifest, "N004")

        # P0 count cannot exceed total findings
        if n003 and n004 and n004.value > n003.value:
            failures.append("P0 findings count > total findings count")

        # Subject count change >20% between runs
        if self.prior_manifest:
            prior_n001 = _manifest_get(self.prior_manifest, "N001")
            curr_n001 = _manifest_get(manifest, "N001")
            if prior_n001 and curr_n001:
                pct_change = abs(curr_n001.value - prior_n001.value) / max(prior_n001.value, 1)
                if pct_change > 0.20:
                    failures.append(
                        f"Subject count changed by {pct_change:.0%} ({prior_n001.value} -> {curr_n001.value})"
                    )

        # Gap count decreased between runs
        if self.prior_manifest:
            prior_n009 = _manifest_get(self.prior_manifest, "N009")
            curr_n009 = _manifest_get(manifest, "N009")
            if prior_n009 and curr_n009 and curr_n009.value < prior_n009.value:
                failures.append(
                    f"Gap count decreased ({prior_n009.value} -> {curr_n009.value}). "
                    f"Gaps should only increase or stay same unless data added."
                )

        return AuditCheck(
            passed=len(failures) == 0,
            dod_checks=[17],
            details={"layer": 5, "failures": failures},
            rule="Layer 5: flag implausible numbers.",
        )

    # ------------------------------------------------------------------ #
    # Layer 6 -- Financial Citation Verification
    # ------------------------------------------------------------------ #

    def check_financial_citations(self, text_dir: Path | None = None) -> AuditCheck:
        """Spot-check that financial values in P0/P1 findings appear in their cited source.

        For each P0/P1 finding that contains dollar amounts, this layer:
        1. Extracts all dollar amounts from the finding's exact_quote and description.
        2. Loads the referenced source file's extracted text.
        3. Verifies that the cited amounts appear in the source (with +/-5% tolerance).

        Parameters
        ----------
        text_dir:
            Directory containing extracted text files (``_dd/text/``).
            When ``None``, skips verification (non-blocking).
        """
        if text_dir is None or not text_dir.exists():
            return AuditCheck(
                passed=True,
                dod_checks=[17],
                details={"layer": 6, "skipped": True, "reason": "text_dir not available"},
                rule="Layer 6: financial values in P0/P1 findings match source documents.",
            )

        findings = self._load_merged_findings()
        failures: list[str] = []
        verified_count = 0
        checked_count = 0

        dollar_pattern = re.compile(
            r"\$[\d,]+(?:\.\d+)?(?:\s*[MBKmk](?:illion|illion)?)?",
        )

        for f in findings:
            severity = f.get("severity", "P3")
            if severity not in ("P0", "P1"):
                continue

            # Extract dollar amounts from the finding.
            description = f.get("description", "")
            finding_amounts = self._extract_dollar_amounts(description, dollar_pattern)

            # Also check exact_quote in citations.
            for cit in f.get("citations", []):
                if isinstance(cit, dict):
                    quote = cit.get("exact_quote", "")
                    if quote:
                        finding_amounts.update(self._extract_dollar_amounts(quote, dollar_pattern))

            if not finding_amounts:
                continue  # No financial values to verify

            checked_count += 1

            # Load cited source text.
            source_text = self._load_cited_source_text(f, text_dir)
            if not source_text:
                continue  # Source not available -- non-blocking

            # Verify each amount appears in source.
            source_amounts = self._extract_dollar_amounts(source_text, dollar_pattern)
            unmatched = []
            for amount in finding_amounts:
                if not self._amount_matches_any(amount, source_amounts, tolerance=0.05):
                    unmatched.append(amount)

            if unmatched:
                failures.append(
                    f"Finding '{f.get('title', 'untitled')[:60]}' ({severity}): "
                    f"values {unmatched} not found in source (+-5% tolerance)"
                )
            else:
                verified_count += 1

        return AuditCheck(
            passed=len(failures) == 0,
            dod_checks=[17],
            details={
                "layer": 6,
                "checked": checked_count,
                "verified": verified_count,
                "failures": failures,
            },
            rule="Layer 6: financial values in P0/P1 findings match source documents.",
        )

    @staticmethod
    def _extract_dollar_amounts(text: str, pattern: re.Pattern[str] | None = None) -> set[float]:
        """Extract normalized dollar amounts from text.

        Handles formats like: $1.2M, $1,200,000, $1.98M, $12.53M, $904K.
        Returns a set of float values in dollars.
        """
        if pattern is None:
            pattern = re.compile(r"\$[\d,]+(?:\.\d+)?(?:\s*[MBKmk](?:illion|illion)?)?")

        amounts: set[float] = set()
        for match in pattern.finditer(text):
            raw = match.group(0).replace("$", "").replace(",", "").strip()
            multiplier = 1.0
            # Check for suffix multipliers.
            lower = raw.lower()
            if lower.endswith("m") or "million" in lower:
                raw = re.sub(r"[MmBbKk](?:illion)?$", "", raw).strip()
                multiplier = 1_000_000.0
            elif lower.endswith("b") or "billion" in lower:
                raw = re.sub(r"[MmBbKk](?:illion)?$", "", raw).strip()
                multiplier = 1_000_000_000.0
            elif lower.endswith("k"):
                raw = re.sub(r"[MmBbKk]$", "", raw).strip()
                multiplier = 1_000.0
            try:
                amounts.add(float(raw) * multiplier)
            except ValueError:
                continue
        return amounts

    @staticmethod
    def _amount_matches_any(
        target: float,
        candidates: set[float],
        tolerance: float = 0.05,
    ) -> bool:
        """Check if target amount matches any candidate within tolerance."""
        if not candidates:
            return False
        for c in candidates:
            if c == 0 and target == 0:
                return True
            if c != 0 and abs(target - c) / abs(c) <= tolerance:
                return True
        return False

    def _load_cited_source_text(self, finding: dict[str, Any], text_dir: Path) -> str:
        """Load extracted text for the first citation's source file."""
        citations = finding.get("citations", [])
        if not citations:
            return ""
        cit = citations[0] if isinstance(citations[0], dict) else {}
        source_path = cit.get("source_path", "")
        if not source_path or source_path.startswith("["):
            return ""

        # Try to find the extracted text file.
        # Source path might be like "1. Due Diligence/Finance/file.xlsx"
        # Extracted text would be at text_dir/subject/file.xlsx.md
        basename = source_path.rsplit("/", 1)[-1] if "/" in source_path else source_path
        # Search text_dir recursively for the basename (with .md suffix).
        for suffix in [".md", ""]:
            for txt_file in text_dir.rglob(f"{basename}{suffix}"):
                try:
                    return txt_file.read_text(errors="replace")[:500_000]  # Cap at 500K chars
                except OSError:
                    continue
        return ""

    # ------------------------------------------------------------------ #
    # private helpers
    # ------------------------------------------------------------------ #

    def _resolve_path(self, source_file: str) -> Path:
        """Resolve ``{RUN_DIR}`` placeholder and glob patterns."""
        resolved = source_file.replace("{RUN_DIR}", str(self.run_dir))
        # If the path contains a glob wildcard, check if at least one match exists
        if "*" in resolved:
            return Path(resolved)
        return Path(resolved)

    def _source_exists(self, path: Path) -> bool:
        path_str = str(path)
        if "*" in path_str:
            from glob import glob as _glob

            return len(_glob(path_str)) > 0
        return path.exists()

    def _rederive(self, entry: ManifestEntry) -> int | float | None:
        """Attempt to re-derive a manifest entry value from source."""
        match entry.id:
            case "N001":
                return self._count_csv_rows("subjects.csv")
            case "N002":
                return self._count_file_lines("files.txt")
            case "N003":
                return self._count_merged_findings()
            case "N004":
                return self._count_findings_by_severity("P0")
            case "N005":
                return self._count_findings_by_severity("P1")
            case "N006":
                return self._count_findings_by_severity("P2")
            case "N007":
                return self._count_findings_by_severity("P3")
            case "N008":
                return self._count_clean_results()
            case "N009":
                return self._count_total_gaps()
            case "N010":
                return self._count_reference_files()
            case _:
                # For entries beyond N001-N010, accept the manifest value
                return None

    def _cached_findings(self) -> list[dict[str, Any]]:
        """Load ALL findings (including clean results) once and cache them."""
        if self._findings_cache is not None:
            return self._findings_cache
        merged_dir = self.run_dir / "findings" / "merged"
        if not merged_dir.exists():
            self._findings_cache = []
            return self._findings_cache
        all_findings: list[dict[str, Any]] = []
        for jf in sorted(merged_dir.glob("*.json")):
            try:
                data = json.loads(jf.read_text(encoding="utf-8"))
                findings = data.get("findings", [])
                all_findings.extend(findings)
            except (json.JSONDecodeError, OSError):
                continue
        self._findings_cache = all_findings
        return self._findings_cache

    def _load_merged_findings(self) -> list[dict[str, Any]]:
        """Load findings excluding ``domain_reviewed_no_issues`` (clean results)."""
        return [f for f in self._cached_findings() if f.get("category") != "domain_reviewed_no_issues"]

    def _count_merged_findings(self) -> int:
        """N003: count total findings from merged dir, excluding domain_reviewed_no_issues."""
        return len(self._load_merged_findings())

    def _count_findings_by_severity(self, severity: str) -> int:
        """N004-N007: count findings matching a specific severity level.

        Excludes clean results (``domain_reviewed_no_issues``) so N007 (P3)
        does not double-count them.
        """
        return sum(1 for f in self._load_merged_findings() if f.get("severity") == severity)

    def _count_clean_results(self) -> int:
        """N008: count findings with category='domain_reviewed_no_issues'."""
        return sum(1 for f in self._cached_findings() if f.get("category") == "domain_reviewed_no_issues")

    def _count_total_gaps(self) -> int:
        """N009: count total gaps from merged subject files.

        Gaps are stored inside each merged subject JSON (``data.gaps[]``),
        not as separate files in ``merged/gaps/``.  This counts all gaps
        across all merged subject files.
        """
        merged_dir = self.run_dir / "findings" / "merged"
        if not merged_dir.exists():
            return 0
        total = 0
        for jf in sorted(merged_dir.glob("*.json")):
            try:
                data = json.loads(jf.read_text(encoding="utf-8"))
                total += len(data.get("gaps", []))
            except (json.JSONDecodeError, OSError):
                continue
        return total

    def _count_reference_files(self) -> int | None:
        """N010: count reference files from reference_files.json."""
        path = self.inventory_dir / "reference_files.json"
        if not path.exists():
            return None
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
            if isinstance(data, list):
                return len(data)
            if isinstance(data, dict):
                return len(data.get("files", []))
        except (json.JSONDecodeError, OSError):
            return None
        return None

    def _count_csv_rows(self, filename: str) -> int | None:
        """Count data rows (excluding header) in a CSV file."""
        path = self.inventory_dir / filename
        if not path.exists():
            return None
        lines = path.read_text(encoding="utf-8").strip().splitlines()
        # Subtract header row
        return max(0, len(lines) - 1) if lines else 0

    def _count_file_lines(self, filename: str) -> int | None:
        """Count lines in a text file."""
        path = self.inventory_dir / filename
        if not path.exists():
            return None
        lines = path.read_text(encoding="utf-8").strip().splitlines()
        return len(lines)

    def _read_counts_json_field(self, field: str) -> Any:
        """Read a field from counts.json."""
        path = self.inventory_dir / "counts.json"
        if not path.exists():
            return None
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
            return data.get(field)
        except (json.JSONDecodeError, OSError):
            return None

    def _find_value_in_workbook(self, wb: Any, entry: ManifestEntry) -> bool:
        """Search for a manifest entry's value in the workbook."""
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(values_only=True):
                for cell_value in row:
                    if cell_value == entry.value:
                        return True
        return False
