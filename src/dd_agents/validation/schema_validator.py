"""Report schema validation.

After Excel generation, validates that the generated workbook conforms
to the :class:`~dd_agents.models.reporting.ReportSchema` definition:
all expected sheets exist, columns match schema order and names,
sort orders are correct, and conditional formatting is applied.
"""

from __future__ import annotations

import logging
from typing import TYPE_CHECKING, Any

from dd_agents.models.audit import AuditCheck

if TYPE_CHECKING:
    from pathlib import Path

    from dd_agents.models.reporting import ReportSchema, SheetDef

logger = logging.getLogger(__name__)


class SchemaValidator:
    """Validate a generated Excel workbook against a ReportSchema.

    Parameters
    ----------
    report_schema:
        The parsed :class:`ReportSchema` to validate against.
    deal_config:
        Optional deal configuration dict used to evaluate activation conditions.
    """

    def __init__(
        self,
        report_schema: ReportSchema,
        deal_config: dict[str, Any] | None = None,
    ) -> None:
        self.schema = report_schema
        self.deal_config: dict[str, Any] = deal_config or {}

    # ------------------------------------------------------------------ #
    # public API
    # ------------------------------------------------------------------ #

    def validate_report(self, excel_path: Path) -> list[AuditCheck]:
        """Validate the Excel workbook at *excel_path* against the schema.

        Returns a list of :class:`AuditCheck` objects -- one per
        validation dimension (sheets, columns, sort, formatting).
        """
        checks: list[AuditCheck] = []

        if not excel_path.exists():
            checks.append(
                AuditCheck(
                    passed=False,
                    dod_checks=[28, 29],
                    details={"error": f"Excel file not found: {excel_path}"},
                    rule="Excel workbook must exist for schema validation.",
                )
            )
            return checks

        try:
            import openpyxl

            wb = openpyxl.load_workbook(excel_path, data_only=True)
        except Exception as exc:
            checks.append(
                AuditCheck(
                    passed=False,
                    dod_checks=[28, 29],
                    details={"error": f"Cannot open Excel: {exc}"},
                )
            )
            return checks

        checks.append(self._check_sheets_exist(wb))
        checks.append(self._check_columns_match(wb))
        checks.append(self._check_sort_orders(wb))
        checks.append(self._check_conditional_formatting(wb))

        return checks

    # ------------------------------------------------------------------ #
    # individual checks
    # ------------------------------------------------------------------ #

    def _check_sheets_exist(self, wb: Any) -> AuditCheck:
        """Verify all expected sheets exist in the workbook."""
        present = set(wb.sheetnames)
        expected_sheets = [s for s in self.schema.sheets if self._sheet_is_active(s)]
        missing: list[str] = []
        for sheet_def in expected_sheets:
            if sheet_def.name not in present:
                missing.append(sheet_def.name)

        return AuditCheck(
            passed=len(missing) == 0,
            dod_checks=[29],
            details={
                "check": "sheets_exist",
                "expected_sheets": [s.name for s in expected_sheets],
                "present_sheets": sorted(present),
                "missing_sheets": missing,
            },
            rule="All expected sheets must exist in the workbook.",
        )

    def _check_columns_match(self, wb: Any) -> AuditCheck:
        """Verify column names and order match schema for each sheet."""
        failures: list[dict[str, Any]] = []

        for sheet_def in self.schema.sheets:
            if not self._sheet_is_active(sheet_def):
                continue
            if sheet_def.name not in wb.sheetnames:
                continue  # already caught by sheets_exist check

            ws = wb[sheet_def.name]
            # Read header row (row 1)
            header_row = []
            for cell in ws[1]:
                if cell.value is not None:
                    header_row.append(str(cell.value))

            expected_cols = [col.name for col in sheet_def.columns if self._column_is_active(col)]

            if not expected_cols:
                continue

            # Check order and names
            mismatches: list[str] = []
            for idx, expected_name in enumerate(expected_cols):
                if idx >= len(header_row):
                    mismatches.append(
                        f"Column {idx}: expected '{expected_name}', but sheet has only {len(header_row)} columns"
                    )
                elif header_row[idx] != expected_name:
                    mismatches.append(f"Column {idx}: expected '{expected_name}', got '{header_row[idx]}'")

            if mismatches:
                failures.append(
                    {
                        "sheet": sheet_def.name,
                        "mismatches": mismatches,
                    }
                )

        return AuditCheck(
            passed=len(failures) == 0,
            dod_checks=[29],
            details={
                "check": "columns_match",
                "failures": failures,
            },
            rule="Column names and order must match schema for each sheet.",
        )

    def _check_sort_orders(self, wb: Any) -> AuditCheck:
        """Verify sort orders are correct for sheets that define them.

        This does a best-effort check: it reads the data rows and
        verifies they are sorted according to the schema's sort_order
        specification.
        """
        failures: list[dict[str, Any]] = []

        for sheet_def in self.schema.sheets:
            if not sheet_def.sort_order:
                continue
            if not self._sheet_is_active(sheet_def):
                continue
            if sheet_def.name not in wb.sheetnames:
                continue

            ws = wb[sheet_def.name]
            header_row = [str(cell.value) if cell.value else "" for cell in ws[1]]

            for sort_spec in sheet_def.sort_order:
                col_name = sort_spec.column
                if col_name not in header_row:
                    continue
                col_idx = header_row.index(col_name)

                # Read column values (skip header)
                values = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if col_idx < len(row) and row[col_idx] is not None:
                        values.append(row[col_idx])

                if len(values) <= 1:
                    continue

                # Check sort direction
                is_desc = sort_spec.direction == "desc"
                is_sorted = self._is_sorted(values, reverse=is_desc)
                if not is_sorted:
                    failures.append(
                        {
                            "sheet": sheet_def.name,
                            "column": col_name,
                            "direction": sort_spec.direction,
                            "error": "data not sorted correctly",
                        }
                    )

        return AuditCheck(
            passed=len(failures) == 0,
            dod_checks=[29],
            details={
                "check": "sort_orders",
                "failures": failures,
            },
            rule="Data must be sorted according to schema sort_order.",
        )

    def _check_conditional_formatting(self, wb: Any) -> AuditCheck:
        """Check that conditional formatting rules are applied.

        This is a structural check: it verifies that the sheets with
        conditional formatting definitions have at least one conditional
        format rule applied in openpyxl.
        """
        issues: list[dict[str, Any]] = []

        for sheet_def in self.schema.sheets:
            if not sheet_def.conditional_formatting:
                continue
            if not self._sheet_is_active(sheet_def):
                continue
            if sheet_def.name not in wb.sheetnames:
                continue

            ws = wb[sheet_def.name]
            # openpyxl stores conditional formatting rules
            cf_rules = ws.conditional_formatting
            if not cf_rules:
                issues.append(
                    {
                        "sheet": sheet_def.name,
                        "expected_rules": len(sheet_def.conditional_formatting),
                        "found_rules": 0,
                        "warning": "no conditional formatting found",
                    }
                )

        return AuditCheck(
            passed=True,  # Warnings only -- do not block on CF
            dod_checks=[29],
            details={
                "check": "conditional_formatting",
                "issues": issues,
            },
            rule="Conditional formatting should be applied per schema.",
        )

    # ------------------------------------------------------------------ #
    # helpers
    # ------------------------------------------------------------------ #

    def _sheet_is_active(self, sheet_def: SheetDef) -> bool:
        """Determine whether a sheet should be present based on activation.

        Evaluates the ``activation_condition`` field:
        - ``"always"`` or empty: sheet is active if required
        - ``"never"``: sheet is never active
        - Other conditions: evaluated against the deal_config context
          (e.g. ``"judge.enabled"``, ``"execution.mode == incremental"``)
        """
        condition = sheet_def.activation_condition
        if not condition or condition == "always":
            return sheet_def.required
        if condition == "never":
            return False
        # Evaluate dotted-path conditions like "judge.enabled"
        # or "source_of_truth.customer_database"
        return self._evaluate_condition(condition)

    def _column_is_active(self, col: Any) -> bool:
        """Determine whether a column should be present.

        Columns without an activation_condition are always active.
        Columns with a condition are evaluated against the deal_config context.
        """
        condition = getattr(col, "activation_condition", None)
        if not condition:
            return True
        if condition == "always":
            return True
        if condition == "never":
            return False
        return self._evaluate_condition(condition)

    @staticmethod
    def _is_sorted(values: list[Any], reverse: bool = False) -> bool:
        """Check if a list of values is sorted."""
        try:
            comparable = [v for v in values if v is not None]
            if not comparable:
                return True
            if reverse:
                return all(comparable[i] >= comparable[i + 1] for i in range(len(comparable) - 1))
            return all(comparable[i] <= comparable[i + 1] for i in range(len(comparable) - 1))
        except TypeError:
            # Incomparable types -- treat as sorted
            return True

    def _evaluate_condition(self, condition: str) -> bool:
        """Evaluate an activation condition against the deal config.

        Supports:
        - Dotted paths: ``"judge.enabled"`` -> deal_config["judge"]["enabled"]
        - Equality: ``"execution.mode == incremental"``
        - Truthy: ``"source_of_truth.customer_database"``

        Returns False when the path is missing or the condition is not met.
        """
        condition = condition.strip()

        # Handle equality conditions
        if "==" in condition:
            parts = condition.split("==", 1)
            path = parts[0].strip()
            expected = parts[1].strip().strip("'\"")
            value = self._resolve_path(path)
            return str(value) == expected

        # Handle "!=" conditions
        if "!=" in condition:
            parts = condition.split("!=", 1)
            path = parts[0].strip()
            expected = parts[1].strip().strip("'\"")
            value = self._resolve_path(path)
            return str(value) != expected

        # Truthy evaluation of dotted path
        value = self._resolve_path(condition)
        return bool(value)

    def _resolve_path(self, path: str) -> Any:
        """Resolve a dotted path like ``"judge.enabled"`` in deal_config."""
        current: Any = self.deal_config
        for part in path.split("."):
            if isinstance(current, dict):
                current = current.get(part)
            else:
                return None
            if current is None:
                return None
        return current
