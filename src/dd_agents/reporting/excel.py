"""Schema-driven Excel report generator.

Reads ``report_schema.json`` at runtime and produces the 14-sheet
(10 always-active + 4 conditional) Excel workbook using openpyxl.
"""

from __future__ import annotations

import contextlib
import logging
from collections.abc import Callable
from typing import TYPE_CHECKING, Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from dd_agents.models.reporting import (
    ColumnDef,
    GlobalFormatting,
    ReportSchema,
    SheetDef,
    SummaryFormulaEntry,
)
from dd_agents.reporting.computed_metrics import ReportDataComputer
from dd_agents.utils.constants import SEVERITY_ORDER

if TYPE_CHECKING:
    from pathlib import Path

    from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


def _recalibrate_merged(merged: dict[str, dict[str, Any]]) -> dict[str, dict[str, Any]]:
    """Return a shallow copy of merged data with all findings recalibrated.

    Ensures Excel output matches the HTML report's recalibrated severity
    values.  Findings are shallow-copied; the original dicts are not mutated.

    Handles both plain dicts and Pydantic model instances (e.g.
    ``MergedCustomerOutput``) by normalizing via ``model_dump()`` first.
    """
    out: dict[str, dict[str, Any]] = {}
    for csn, data in merged.items():
        # Normalize Pydantic model instances to dicts
        if hasattr(data, "model_dump"):
            data = data.model_dump()
        elif not isinstance(data, dict):
            out[csn] = data  # type: ignore[assignment]
            continue
        raw_findings = data.get("findings", [])
        recalibrated = [
            ReportDataComputer._recalibrate_severity(f if isinstance(f, dict) else dict(f)) for f in raw_findings
        ]
        out[csn] = {**data, "findings": recalibrated}
    return out


# Overall risk rating priority for sorting
RISK_RANK: dict[str, int] = {
    "Critical": 0,
    "High": 1,
    "Medium": 2,
    "Low": 3,
    "Clean": 4,
}


def compute_overall_risk(
    findings: list[dict[str, Any]],
    gaps: list[dict[str, Any]],
) -> str:
    """Compute overall risk rating for a customer.

    Uses softened thresholds (Issue #113):
    - Critical: P0 count >= 3
    - High: any P0 (1-2), or P1 count >= 3
    - Medium: any P1, or P2 count >= 5
    - Low: P2 or P3 present
    - Clean: no material findings and no gaps
    """
    from collections import Counter

    sev_counter: Counter[str] = Counter()
    for f in findings:
        if f.get("category") != "domain_reviewed_no_issues":
            sev_counter[f.get("severity", f.get("priority", ""))] += 1
    for g in gaps:
        sev_counter[g.get("priority", "")] += 1

    p0 = sev_counter.get("P0", 0)
    p1 = sev_counter.get("P1", 0)
    p2 = sev_counter.get("P2", 0)

    if p0 >= 3:
        return "Critical"
    if p0 > 0:
        return "High"
    if p1 >= 3:
        return "High"
    if p1 > 0 or p2 >= 5:
        return "Medium"
    if p2 > 0 or sev_counter.get("P3", 0) > 0:
        return "Low"
    return "Clean"


class ExcelReportGenerator:
    """Generate the Excel workbook from ``report_schema.json`` and merged data."""

    def __init__(self) -> None:
        self._wb: Workbook | None = None

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def generate(
        self,
        merged_findings: dict[str, Any],
        report_schema: ReportSchema,
        output_path: Path,
        deal_config: dict[str, Any] | None = None,
        run_metadata: dict[str, Any] | None = None,
    ) -> Path:
        """Create the Excel workbook and save it to *output_path*.

        Parameters
        ----------
        merged_findings:
            ``{customer_safe_name: MergedCustomerOutput | dict}``
        report_schema:
            Parsed ``ReportSchema`` model from ``report_schema.json``.
        output_path:
            Destination ``.xlsx`` path.
        deal_config:
            The deal configuration dict (used for activation conditions
            and the _Metadata sheet).
        run_metadata:
            Additional run metadata key/value pairs for _Metadata.
        """
        deal_config = deal_config or {}
        run_metadata = run_metadata or {}

        # ------------------------------------------------------------------
        # Guard: reject schemas with zero sheet definitions (Issue #35).
        # ------------------------------------------------------------------
        if not report_schema.sheets:
            raise ValueError(
                "Cannot generate Excel report: report_schema has zero sheet "
                "definitions. Provide a valid report_schema.json with at "
                "least one sheet."
            )

        # Apply severity recalibration so Excel matches the HTML report.
        merged_findings = _recalibrate_merged(merged_findings)  # type: ignore[arg-type]

        self._wb = Workbook()
        # Remove default sheet
        default_ws = self._wb.active
        if default_ws is not None:
            self._wb.remove(default_ws)

        fmt = report_schema.global_formatting

        # Track per-sheet row counts to warn about empty sheets later
        sheet_row_counts: dict[str, int] = {}

        for sheet_def in report_schema.sheets:
            if not self._is_activated(sheet_def, deal_config):
                continue

            ws = self._wb.create_sheet(title=sheet_def.name)

            # Filter columns by activation condition
            active_cols = self._active_columns(sheet_def, deal_config)

            data_rows = self._prepare_sheet_data(
                sheet_def.name,
                merged_findings,
                deal_config,
                run_metadata,
            )

            sheet_row_counts[sheet_def.name] = len(data_rows)

            self._write_sheet(ws, sheet_def, active_cols, data_rows, fmt)
            self._apply_formatting(ws, sheet_def, active_cols, fmt)

        # ------------------------------------------------------------------
        # Warn about empty non-Summary sheets (Issue #53).
        # ------------------------------------------------------------------
        for sheet_name, row_count in sheet_row_counts.items():
            if row_count == 0 and sheet_name != "Summary":
                logger.warning(
                    "Sheet '%s' has only headers and zero data rows",
                    sheet_name,
                )

        output_path.parent.mkdir(parents=True, exist_ok=True)
        self._wb.save(str(output_path))
        return output_path

    # ------------------------------------------------------------------
    # Sheet writing
    # ------------------------------------------------------------------

    def _write_sheet(
        self,
        ws: Worksheet,
        sheet_def: SheetDef,
        columns: list[ColumnDef],
        data_rows: list[dict[str, Any]],
        fmt: GlobalFormatting,
    ) -> None:
        """Write headers and data rows into the worksheet."""
        # Headers
        for col_idx, col_def in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_def.name)
            cell.fill = PatternFill(
                start_color=fmt.header_bg_color.lstrip("#"),
                end_color=fmt.header_bg_color.lstrip("#"),
                fill_type="solid",
            )
            cell.font = Font(
                bold=fmt.header_bold,
                color=fmt.header_font_color.lstrip("#"),
                size=fmt.header_font_size,
            )

        # Sort data rows
        data_rows = self._sort_rows(data_rows, sheet_def)

        # Data
        for row_idx, row_data in enumerate(data_rows, start=2):
            for col_idx, col_def in enumerate(columns, start=1):
                value = row_data.get(col_def.key, "")
                # Handle list values -- join with "; " so that
                # Excel cells never contain raw Python repr strings
                # like ['item1', 'item2'] (Issue #53).
                if isinstance(value, list):
                    value = "; ".join(str(v) for v in value)
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = Font(size=fmt.body_font_size)

        # Summary formulas
        self._write_summary_row(ws, sheet_def, columns, data_rows, fmt)

    # ------------------------------------------------------------------
    # Formatting
    # ------------------------------------------------------------------

    def _apply_formatting(
        self,
        ws: Worksheet,
        sheet_def: SheetDef,
        columns: list[ColumnDef],
        fmt: GlobalFormatting,
    ) -> None:
        """Apply freeze panes, auto-filter, column widths, conditional formatting."""
        # Freeze panes
        if fmt.freeze_panes:
            ws.freeze_panes = f"A{fmt.freeze_row + 1}"

        # Auto-filter
        if fmt.auto_filter and ws.max_row and ws.max_row > 1:
            last_col = get_column_letter(ws.max_column)
            ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"

        # Column widths
        for col_idx, col_def in enumerate(columns, start=1):
            width = min(col_def.width, fmt.max_column_width)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Conditional formatting (cell-level colour)
        col_key_to_idx = {c.key: i + 1 for i, c in enumerate(columns)}
        for cf_rule in sheet_def.conditional_formatting:
            cf_col_idx = col_key_to_idx.get(cf_rule.column)
            if cf_col_idx is None:
                continue
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=cf_col_idx)
                if self._rule_matches(cf_rule.rule, cell.value):
                    cell.fill = PatternFill(
                        start_color=cf_rule.format.bg.lstrip("#"),
                        end_color=cf_rule.format.bg.lstrip("#"),
                        fill_type="solid",
                    )
                    cell.font = Font(
                        size=fmt.body_font_size,
                        color=cf_rule.format.font.lstrip("#"),
                    )

    # ------------------------------------------------------------------
    # Data preparation per sheet
    # ------------------------------------------------------------------

    def _prepare_sheet_data(
        self,
        sheet_name: str,
        merged_findings: dict[str, Any],
        deal_config: dict[str, Any],
        run_metadata: dict[str, Any],
    ) -> list[dict[str, Any]]:
        """Prepare data rows for a given sheet name."""
        # Normalise MergedCustomerOutput instances to dicts
        normalised: dict[str, dict[str, Any]] = {}
        for csn, mco in merged_findings.items():
            if hasattr(mco, "model_dump"):
                normalised[csn] = mco.model_dump()
            elif isinstance(mco, dict):
                normalised[csn] = mco
            else:
                normalised[csn] = dict(mco)

        _sheet_handler_t = Callable[
            [dict[str, dict[str, Any]], dict[str, Any], dict[str, Any]],
            list[dict[str, Any]],
        ]
        dispatch: dict[str, _sheet_handler_t] = {
            "Summary": self._data_summary,
            "Wolf_Pack": self._data_wolf_pack,
            "Legal_Risks": lambda m, *a: self._data_agent_findings(m, "legal"),
            "Commercial_Data": lambda m, *a: self._data_agent_findings(m, "commercial"),
            "Financials": lambda m, *a: self._data_agent_findings(m, "finance"),
            "Product_Scope": lambda m, *a: self._data_agent_findings(m, "producttech"),
            "Data_Reconciliation": self._data_reconciliation,
            "Missing_Docs_Gaps": self._data_gaps,
            "Contract_Date_Reconciliation": self._data_contract_dates,
            "Reference_Files_Index": self._data_reference_files,
            "Entity_Resolution_Log": self._data_entity_log,
            "Quality_Audit": self._data_quality_audit,
            "Run_Diff": self._data_run_diff,
            "_Metadata": self._data_metadata,
        }

        handler = dispatch.get(sheet_name)
        if handler:
            return handler(normalised, deal_config, run_metadata)
        return []

    # -- Summary -------------------------------------------------------

    def _data_summary(
        self,
        merged: dict[str, dict[str, Any]],
        deal_config: dict[str, Any],
        run_metadata: dict[str, Any],
    ) -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []
        for _csn, data in sorted(merged.items()):
            findings = data.get("findings", [])
            gaps = data.get("gaps", [])

            sev_counts = {"P0": 0, "P1": 0, "P2": 0, "P3": 0}
            for f in findings:
                if f.get("category") == "domain_reviewed_no_issues":
                    continue
                sev = f.get("severity", "P3")
                if sev in sev_counts:
                    sev_counts[sev] += 1

            row = {
                "customer": data.get("customer", _csn),
                "overall_risk_rating": compute_overall_risk(findings, gaps),
                "p0_count": sev_counts["P0"],
                "p1_count": sev_counts["P1"],
                "p2_count": sev_counts["P2"],
                "p3_count": sev_counts["P3"],
                "total_findings": len(findings),
                "gap_count": len(gaps),
                "files_analyzed": data.get("files_analyzed", 0),
                "governance_resolved_pct": data.get("governance_resolved_pct", 0.0),
            }
            rows.append(row)
        return rows

    # -- Wolf Pack (P0 + P1) -------------------------------------------

    def _data_wolf_pack(
        self,
        merged: dict[str, dict[str, Any]],
        deal_config: dict[str, Any],
        run_metadata: dict[str, Any],
    ) -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []
        for _csn, data in sorted(merged.items()):
            for f in data.get("findings", []):
                sev = f.get("severity", "P3")
                if sev in ("P0", "P1"):
                    rows.append(self._finding_to_row(f, data))
        return rows

    # -- Agent-specific sheets -----------------------------------------

    def _data_agent_findings(
        self,
        merged: dict[str, dict[str, Any]],
        agent: str,
    ) -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []
        for _csn, data in sorted(merged.items()):
            for f in data.get("findings", []):
                if f.get("agent") == agent:
                    rows.append(self._finding_to_row(f, data))
        return rows

    # -- Data Reconciliation -------------------------------------------

    def _data_reconciliation(
        self,
        merged: dict[str, dict[str, Any]],
        deal_config: dict[str, Any],
        run_metadata: dict[str, Any],
    ) -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []
        for _csn, data in sorted(merged.items()):
            customer = data.get("customer", _csn)
            for cr in data.get("cross_references", []):
                cr_dict = cr if isinstance(cr, dict) else dict(cr)
                row: dict[str, Any] = {"customer": customer}
                row["data_type"] = cr_dict.get("data_type", "")
                row["data_point"] = cr_dict.get("data_point", "")
                row["contract_value"] = cr_dict.get("contract_value", "")
                cs = cr_dict.get("contract_source", {})
                row["contract_source_file"] = cs.get("file", "") if isinstance(cs, dict) else ""
                row["reference_value"] = cr_dict.get("reference_value", "")
                rs = cr_dict.get("reference_source", {})
                row["reference_source_file"] = rs.get("file", "") if isinstance(rs, dict) else ""
                row["match_status"] = cr_dict.get("match_status", "")
                row["variance"] = cr_dict.get("variance", "")
                row["severity"] = cr_dict.get("severity", "")
                row["interpretation"] = cr_dict.get("interpretation", "")
                rows.append(row)
        return rows

    # -- Gaps ----------------------------------------------------------

    def _data_gaps(
        self,
        merged: dict[str, dict[str, Any]],
        deal_config: dict[str, Any],
        run_metadata: dict[str, Any],
    ) -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []
        for _csn, data in sorted(merged.items()):
            for g in data.get("gaps", []):
                g_dict = g if isinstance(g, dict) else dict(g)
                rows.append(g_dict)
        return rows

    # -- Contract dates ------------------------------------------------

    def _data_contract_dates(
        self,
        merged: dict[str, dict[str, Any]],
        deal_config: dict[str, Any],
        run_metadata: dict[str, Any],
    ) -> list[dict[str, Any]]:
        recon = run_metadata.get("contract_date_reconciliation", {})
        result: list[dict[str, Any]] = recon.get("entries", [])
        return result

    # -- Reference files -----------------------------------------------

    def _data_reference_files(
        self,
        merged: dict[str, dict[str, Any]],
        deal_config: dict[str, Any],
        run_metadata: dict[str, Any],
    ) -> list[dict[str, Any]]:
        result: list[dict[str, Any]] = run_metadata.get("reference_files", [])
        return result

    # -- Entity resolution log -----------------------------------------

    def _data_entity_log(
        self,
        merged: dict[str, dict[str, Any]],
        deal_config: dict[str, Any],
        run_metadata: dict[str, Any],
    ) -> list[dict[str, Any]]:
        result: list[dict[str, Any]] = run_metadata.get("entity_matches", [])
        return result

    # -- Quality audit -------------------------------------------------

    def _data_quality_audit(
        self,
        merged: dict[str, dict[str, Any]],
        deal_config: dict[str, Any],
        run_metadata: dict[str, Any],
    ) -> list[dict[str, Any]]:
        result: list[dict[str, Any]] = run_metadata.get("quality_scores", {}).get("spot_checks", [])
        return result

    # -- Run diff ------------------------------------------------------

    def _data_run_diff(
        self,
        merged: dict[str, dict[str, Any]],
        deal_config: dict[str, Any],
        run_metadata: dict[str, Any],
    ) -> list[dict[str, Any]]:
        diff = run_metadata.get("report_diff", {})
        result: list[dict[str, Any]] = diff.get("changes", [])
        return result

    # -- Metadata ------------------------------------------------------

    def _data_metadata(
        self,
        merged: dict[str, dict[str, Any]],
        deal_config: dict[str, Any],
        run_metadata: dict[str, Any],
    ) -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []
        # Flatten deal_config top-level keys
        for k, v in deal_config.items():
            rows.append({"property": k, "value": str(v)})
        for k, v in run_metadata.items():
            if k not in (
                "contract_date_reconciliation",
                "reference_files",
                "entity_matches",
                "quality_scores",
                "report_diff",
            ):
                rows.append({"property": k, "value": str(v)})
        return rows

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _finding_to_row(finding: dict[str, Any], customer_data: dict[str, Any]) -> dict[str, Any]:
        """Flatten a finding dict into a row dict."""
        cit = (finding.get("citations") or [{}])[0]
        if isinstance(cit, dict):
            source_path = cit.get("source_path", "")
            location = cit.get("location", "")
            exact_quote = cit.get("exact_quote", "")
        else:
            source_path = getattr(cit, "source_path", "")
            location = getattr(cit, "location", "")
            exact_quote = getattr(cit, "exact_quote", "")

        return {
            "analysis_unit": finding.get("analysis_unit", customer_data.get("customer", "")),
            "severity": finding.get("severity", ""),
            "agent": finding.get("agent", ""),
            "category": finding.get("category", ""),
            "title": finding.get("title", ""),
            "description": finding.get("description", ""),
            "citation_source_path": source_path,
            "citation_location": location,
            "citation_exact_quote": exact_quote,
            "confidence": finding.get("confidence", ""),
            "id": finding.get("id", ""),
        }

    @staticmethod
    def _sort_rows(rows: list[dict[str, Any]], sheet_def: SheetDef) -> list[dict[str, Any]]:
        """Sort rows according to the sheet's sort_order definition."""
        if not sheet_def.sort_order:
            return rows

        def sort_key(row: dict[str, Any]) -> tuple[Any, ...]:
            parts: list[Any] = []
            for so in sheet_def.sort_order:
                val = row.get(so.column, "")
                # Handle special ranking for known columns
                rank: int | str
                if so.column == "severity" or so.column == "priority":
                    rank = SEVERITY_ORDER.get(str(val), 9)
                elif so.column == "overall_risk_rating":
                    rank = RISK_RANK.get(str(val), 9)
                else:
                    rank = str(val).lower() if val is not None else ""
                if so.direction == "desc" and isinstance(rank, int):
                    rank = -rank
                parts.append(rank)
            return tuple(parts)

        return sorted(rows, key=sort_key)

    @staticmethod
    def _is_activated(sheet_def: SheetDef, deal_config: dict[str, Any]) -> bool:
        """Always activate all sheets so the workbook has all 14 tabs.

        Conditional sheets (Quality_Audit, Run_Diff, _Metadata,
        Contract_Date_Reconciliation) may have zero data rows, but they
        still appear with headers so the report structure is complete.
        """
        return True

    @staticmethod
    def _active_columns(
        sheet_def: SheetDef,
        deal_config: dict[str, Any],
    ) -> list[ColumnDef]:
        """Filter columns whose activation_condition is met."""
        result: list[ColumnDef] = []
        for col in sheet_def.columns:
            cond = col.activation_condition
            if cond is None or cond == "always":
                result.append(col)
            elif "judge.enabled" in cond:
                if deal_config.get("judge", {}).get("enabled", False):
                    result.append(col)
            else:
                result.append(col)
        return result

    @staticmethod
    def _rule_matches(rule: str, value: Any) -> bool:
        """Evaluate a conditional formatting rule against a cell value."""
        if value is None:
            return False
        val_str = str(value)
        if rule.startswith("== "):
            return val_str == rule[3:]
        if rule.startswith("contains "):
            return rule[9:].lower() in val_str.lower()
        if rule.startswith("< "):
            try:
                return float(val_str) < float(rule[2:])
            except (ValueError, TypeError):
                return False
        if rule.startswith("> "):
            try:
                return float(val_str) > float(rule[2:])
            except (ValueError, TypeError):
                return False
        return False

    def _write_summary_row(
        self,
        ws: Worksheet,
        sheet_def: SheetDef,
        columns: list[ColumnDef],
        data_rows: list[dict[str, Any]],
        fmt: GlobalFormatting,
    ) -> None:
        """Write summary formula rows if defined in the schema."""
        formulas = sheet_def.summary_formulas
        if not formulas:
            return

        col_key_to_idx = {c.key: i + 1 for i, c in enumerate(columns)}

        for position, entries in formulas.items():
            if position != "bottom":
                continue

            summary_row = ws.max_row + 1

            for entry_raw in entries:
                entry = SummaryFormulaEntry.model_validate(entry_raw) if isinstance(entry_raw, dict) else entry_raw

                col_idx = col_key_to_idx.get(entry.column)
                if col_idx is None:
                    continue

                if entry.value is not None:
                    cell = ws.cell(row=summary_row, column=col_idx, value=entry.value)
                    cell.font = Font(bold=True, size=fmt.body_font_size)
                elif entry.formula:
                    result = self._eval_formula(entry.formula, entry.column, data_rows)
                    cell = ws.cell(row=summary_row, column=col_idx, value=result)
                    cell.font = Font(bold=True, size=fmt.body_font_size)

    @staticmethod
    def _eval_formula(
        formula: str,
        column_key: str,
        data_rows: list[dict[str, Any]],
    ) -> int | float | str:
        """Evaluate a pseudo-formula against the data rows."""
        if formula.startswith("SUM("):
            target = formula[4:-1]
            total = 0.0
            for row in data_rows:
                val = row.get(target, 0)
                with contextlib.suppress(ValueError, TypeError):
                    total += float(val)
            return int(total) if total == int(total) else total

        if formula.startswith("COUNTIF("):
            # COUNTIF(column, value)
            inner = formula[8:-1]
            parts = [p.strip() for p in inner.split(",", 1)]
            if len(parts) == 2:
                col, match_val = parts
                count = sum(1 for row in data_rows if str(row.get(col, "")) == match_val)
                return count
            return 0

        if formula.startswith("COUNTA_UNIQUE("):
            target = formula[14:-1]
            unique: set[str] = set()
            for row in data_rows:
                val = row.get(target, "")
                if val:
                    unique.add(str(val))
            return len(unique)

        if formula.startswith("COUNTA("):
            target = formula[7:-1]
            count = sum(1 for row in data_rows if row.get(target))
            return count

        return ""
